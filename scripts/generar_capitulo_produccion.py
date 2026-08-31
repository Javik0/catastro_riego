# -*- coding: utf-8 -*-
"""
Capítulo del informe técnico: "Producción agropecuaria".

Cubre la sección 4 de la ficha (Producción) y sus dos tablas relacionadas:
12.984 registros de cultivos y 9.819 de especies pecuarias.

CRITERIOS DE ANÁLISIS
---------------------
· Universo: los PREDIOS (fichas principales y adicionales). A diferencia de los
  capítulos de encuesta —donde se cuentan personas— aquí se mide producción, y
  un titular con tres predios produce en los tres.
· Los nombres de cultivo y especie vienen con dos escrituras (mayúsculas y
  normal) por cambios en el formulario: 'CEBOLLA' y 'Cebolla' son lo mismo. Se
  unifican; sin eso, la cebolla aparecería partida en dos filas.
· Cuando el técnico eligió "Otros" y detalló el nombre, se usa ese detalle.
· El ganado se reporta con y sin el caso de la granja avícola de ASOCIACIÓN
  ROSALÍA: seis fichas de la misma familia, sobre el mismo predio, declaran
  10.000 gallinas cada una. Es la misma explotación contada seis veces y por sí
  sola representa el 35 % de las cabezas del sistema.

SALIDAS
  docs/CAPITULO-produccion-agropecuaria.html
  build_entrega/Produccion_Agropecuaria.xlsx
"""

import os
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica, nombre_publico, normalizar  # noqa: E402
import informe_estilo as E  # noqa: E402

GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
T = 'Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e'
C = 'Cultivos_Agricolas_ebc9efb2_1fb3_459f_9538_6ecb946d1632'
AN = 'Animales_Especies_74c54436_56a5_45e4_aa36_20830a4c33f5'
BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
HTML = os.path.join(BASE, 'docs', 'CAPITULO-produccion-agropecuaria.html')
XLSX = os.path.join(BASE, 'build_entrega', 'Produccion_Agropecuaria.xlsx')

# Explotación cuyo ganado está declarado varias veces sobre el mismo predio.
GRANJA_DUPLICADA = ('CEVALLOS GORDON', 10000)
MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
         'octubre noviembre diciembre').split()


def num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def pct(a, b):
    return 100.0 * a / b if b else 0.0


def titulo(t):
    """Unifica 'CEBOLLA' y 'Cebolla'; respeta siglas y barras."""
    t = re.sub(r'\s+', ' ', (t or '').strip())
    if not t:
        return ''
    return ' / '.join(p.strip().capitalize() for p in t.split('/'))


def cargar():
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(f'SELECT * FROM "{T}"')
    fichas = {r['id']: dict(r) for r in cur.fetchall()}
    cur.execute(f'SELECT * FROM "{C}"')
    cultivos = [dict(r) for r in cur.fetchall()]
    cur.execute(f'SELECT * FROM "{AN}"')
    animales = [dict(r) for r in cur.fetchall()]
    cur.execute(f'SELECT MAX(fecha_creacion), MAX(fecha_completado) FROM "{T}"')
    f1, f2 = cur.fetchone()
    cur.execute(f'SELECT COUNT(*) FROM "{T}" WHERE es_ficha_hija = 1 AND '
                f'coalesce(estado_investigacion, "pendiente_produccion") != "completada"')
    pendientes = cur.fetchone()[0]
    con.close()

    from generar_capas_sectores_comunidades import COM_A_SECTOR
    vistos = defaultdict(Counter)
    for p in fichas.values():
        crudo = p.get('comunidad') or ''
        p['_comk'] = canonica(crudo) or '(sin comunidad)'
        vistos[p['_comk']][nombre_publico(crudo) or '(sin comunidad)'] += 1
        p['_sec'] = (p.get('sector_investigacion') or '').strip()
    display = {}
    for k, c in vistos.items():
        val = [(n_, v) for n_, v in c.most_common() if normalizar(n_) == k]
        display[k] = val[0][0] if val else k
    for p in fichas.values():
        p['_com'] = display[p['_comk']]
        if p['_sec'] in ('', 'None'):
            p['_sec'] = COM_A_SECTOR.get(p['_comk'], '(sin sector)')

    corte = max(str(f1 or '')[:10], str(f2 or '')[:10])
    corte_txt = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                 if corte else 'la fecha de generación')
    return fichas, cultivos, animales, corte_txt, pendientes


def main():
    fichas, cultivos, animales, corte_txt, pendientes = cargar()

    # ── cultivos ──
    sup, frec = defaultdict(float), Counter()
    destino_c = Counter()
    for x in cultivos:
        t = titulo(x.get('tipo_cultivo'))
        if t.lower() in ('otro', 'otros') and (x.get('tipo_cultivo_otro') or '').strip():
            t = titulo(x['tipo_cultivo_otro'])
        if not t:
            continue
        frec[t] += 1
        sup[t] += num(x.get('superficie_m2'))
        for campo, et in (('es_autoconsumo', 'Autoconsumo'), ('es_mercado', 'Mercado'),
                          ('es_agroindustria', 'Agroindustria'), ('es_exportacion', 'Exportación')):
            if str(x.get(campo) or '') in ('1', 'True'):
                destino_c[et] += 1
    sup_total = sum(sup.values())
    predios_c = len({x['ficha_id'] for x in cultivos})

    # agrupación funcional para la lectura del informe
    GRUPOS = {
        'Pastos y forraje': ('pasto', 'alfalfa', 'forraje'),
        'Hortalizas y tubérculos': ('cebolla', 'papa', 'hortaliza', 'melloco', 'zanahoria',
                                    'brocoli', 'brócoli', 'lechuga', 'zambo', 'zapallo'),
        'Cereales y leguminosas': ('maíz', 'maiz', 'cebada', 'trigo', 'haba', 'frijol',
                                   'chocho', 'quinua', 'arveja', 'lenteja'),
        'Flores': ('flor', 'rosa'),
        'Frutales': ('frutal', 'mora', 'tomate de árbol', 'aguacate', 'durazno', 'capuli'),
        'Bosque y áreas no cultivadas': ('bosque', 'monte', 'baldío', 'baldio', 'eucalipto'),
    }

    def grupo(nombre):
        n = unicodedata.normalize('NFD', nombre.lower())
        n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
        for g, claves in GRUPOS.items():
            if any(k in n for k in claves):
                return g
        return 'Otros cultivos'

    sup_grupo = defaultdict(float)
    for t, s in sup.items():
        sup_grupo[grupo(t)] += s

    # ── ganado ──
    cabezas, reg_esp = defaultdict(int), Counter()
    destino_a = Counter()
    cabezas_dup = 0
    for x in animales:
        e = titulo(x.get('especie'))
        if e.lower() in ('otro', 'otros') and (x.get('especie_otro') or '').strip():
            e = titulo(x['especie_otro'])
        if not e:
            continue
        f = fichas.get(x['ficha_id'], {})
        apel = (f.get('apellidos') or '').upper()
        n = int(num(x.get('cantidad')))
        if GRANJA_DUPLICADA[0] in apel and n >= GRANJA_DUPLICADA[1]:
            cabezas_dup += n
            continue                      # se excluye del cómputo general
        reg_esp[e] += 1
        cabezas[e] += n
        for campo, et in (('es_autoconsumo', 'Autoconsumo'), ('es_mercado', 'Mercado'),
                          ('es_agroindustria', 'Agroindustria'), ('es_exportacion', 'Exportación')):
            if str(x.get(campo) or '') in ('1', 'True'):
                destino_a[et] += 1
    total_cab = sum(cabezas.values())
    predios_a = len({x['ficha_id'] for x in animales})

    # ── soberanía alimentaria (campo de la ficha) ──
    sob = [num(p.get('soberania_aliment_pct')) for p in fichas.values()
           if p.get('es_ficha_hija') not in (1, True)
           and p.get('soberania_aliment_pct') not in (None, '')]
    sob_media = sum(sob) / len(sob) if sob else 0

    # ── por sector ──
    por_sector = defaultdict(lambda: {'sup': 0.0, 'cab': 0, 'predios': set()})
    idx_f = {k: v for k, v in fichas.items()}
    for x in cultivos:
        f = idx_f.get(x['ficha_id'])
        if not f:
            continue
        por_sector[f['_sec']]['sup'] += num(x.get('superficie_m2'))
        por_sector[f['_sec']]['predios'].add(x['ficha_id'])
    for x in animales:
        f = idx_f.get(x['ficha_id'])
        if not f:
            continue
        apel = (f.get('apellidos') or '').upper()
        n = int(num(x.get('cantidad')))
        if GRANJA_DUPLICADA[0] in apel and n >= GRANJA_DUPLICADA[1]:
            continue
        por_sector[f['_sec']]['cab'] += n

    # La acuicultura no es hato ganadero: se declara aparte para no llamar
    # "cabezas" a los peces ni mezclar dos actividades distintas.
    ACUICOLA = ('trucha', 'tilapia', 'pez', 'peces')
    acuicola = {e: n for e, n in cabezas.items()
                if any(k in e.lower() for k in ACUICOLA)}
    cab_acui = sum(acuicola.values())
    total_pec = total_cab - cab_acui

    # ── documento ──
    B = []
    A = B.append
    A(E.cabecera('Producción agropecuaria',
                 'Cultivos, ganadería y destino de la producción · '
                 'Capítulo del informe técnico'))
    A(E.aviso_corte(corte_txt, len([1 for p in fichas.values()
                                    if p.get('es_ficha_hija') not in (1, True)]), pendientes))
    A(E.kpis([
        (f'{sup_total / 10000:,.0f} ha', 'superficie cultivada'),
        (f'{len(sup):,}', 'cultivos distintos'),
        (f'{total_pec:,}', 'animales de granja'),
        (f'{pct(destino_c["Autoconsumo"], sum(destino_c.values())):.0f}%', 'destino autoconsumo'),
    ]))

    A('<h2>1. Alcance</h2>')
    A(f'<p>Se registraron <b>{len(cultivos):,} declaraciones de cultivo</b> en '
      f'<b>{predios_c:,} predios</b> y <b>{len(animales):,} declaraciones de ganado</b> '
      f'en <b>{predios_a:,} predios</b>. A diferencia de los capítulos de encuesta, '
      'aquí la unidad de análisis es el <b>predio</b> y no la persona: un titular '
      'con varias parcelas produce en todas ellas.</p>')

    A('<h2>2. Uso agrícola del suelo</h2>')
    A(f'<p>La superficie declarada en cultivos asciende a '
      f'<b>{sup_total / 10000:,.1f} hectáreas</b>, distribuidas así:</p>')
    A('<table class="evitar-corte"><tr><th>Grupo de cultivo</th>'
      '<th class="n">Superficie (ha)</th><th>Peso</th></tr>')
    for g, s in sorted(sup_grupo.items(), key=lambda x: -x[1]):
        A(f'<tr><td>{g}</td><td class="n">{s / 10000:,.1f}</td>'
          f'<td>{E.barra(pct(s, sup_total))}</td></tr>')
    A('</table>')
    past = sup_grupo['Pastos y forraje']
    A(f'<p>El dato dominante es que <b>{pct(past, sup_total):.1f} % de la superficie '
      f'cultivada son pastos</b> ({past / 10000:,.1f} ha), destinados a sostener la '
      'ganadería. La agricultura de consumo y venta ocupa el resto, con la cebolla y '
      'la papa como cultivos comerciales de referencia.</p>')
    A('<h3>Cultivos individuales de mayor superficie</h3>')
    A('<table class="evitar-corte"><tr><th>Cultivo</th><th class="n">Superficie (ha)</th>'
      '<th class="n">Predios</th><th class="n">Superficie media (m²)</th></tr>')
    for t, s in sorted(sup.items(), key=lambda x: -x[1])[:12]:
        A(f'<tr><td>{t}</td><td class="n">{s / 10000:,.1f}</td>'
          f'<td class="n">{frec[t]:,}</td><td class="n">{s / frec[t]:,.0f}</td></tr>')
    A('</table>')

    A('<h2>3. Ganadería</h2>')
    A(f'<p>Se contabilizan <b>{total_pec:,} animales</b> en {predios_a:,} predios. '
      'La composición del hato revela una ganadería <b>de traspatio</b>, orientada '
      'al consumo familiar y a la venta de excedentes:</p>')
    A('<table class="evitar-corte"><tr><th>Especie</th><th class="n">Animales</th>'
      '<th class="n">Predios</th><th>Peso</th></tr>')
    for e, n in sorted(((e, n) for e, n in cabezas.items() if e not in acuicola),
                       key=lambda x: -x[1])[:12]:
        A(f'<tr><td>{e}</td><td class="n">{n:,}</td><td class="n">{reg_esp[e]:,}</td>'
          f'<td>{E.barra(pct(n, total_pec))}</td></tr>')
    A('</table>')
    if acuicola:
        det = ', '.join(f'{n:,} {e.lower()}' for e, n in
                        sorted(acuicola.items(), key=lambda x: -x[1]))
        A(f'<p>Se registra además <b>producción acuícola</b> —{det}— en '
          f'{sum(reg_esp[e] for e in acuicola)} predios. Se contabiliza aparte por '
          'tratarse de una actividad de naturaleza distinta a la ganadería.</p>')
    if cabezas_dup:
        A(f'<div class="alerta"><b>Dato excluido, a verificar en campo.</b> '
          f'Seis fichas de una misma familia en ASOCIACIÓN ROSALÍA declaran '
          f'<b>10.000 gallinas cada una</b> —{cabezas_dup:,} en total— sobre el '
          '<b>mismo predio</b> de 9,8 ha. Se trata, con toda probabilidad, de una '
          'única granja avícola contabilizada una vez por cada titular. Incluirla '
          f'elevaría el hato a {total_pec + cabezas_dup:,} animales y haría que una '
          'sola explotación representara el 35 % del ganado del sistema. '
          '<b>Se excluye de las cifras de este capítulo</b> hasta confirmar cuántas '
          'aves existen realmente.</div>')

    A('<h2>4. Destino de la producción</h2>')
    A('<p>Cada declaración indica a qué se destina lo producido. Un mismo cultivo o '
      'especie puede tener más de un destino, de modo que los porcentajes se calculan '
      'sobre el total de menciones:</p>')
    A('<table class="evitar-corte"><tr><th>Destino</th>'
      '<th class="n">Cultivos</th><th class="n">Ganado</th><th>Peso agrícola</th></tr>')
    tot_dc, tot_da = sum(destino_c.values()), sum(destino_a.values())
    for et in ('Autoconsumo', 'Mercado', 'Agroindustria', 'Exportación'):
        A(f'<tr><td>{et}</td><td class="n">{destino_c[et]:,}</td>'
          f'<td class="n">{destino_a[et]:,}</td>'
          f'<td>{E.barra(pct(destino_c[et], tot_dc))}</td></tr>')
    A('</table>')
    A(f'<div class="hallazgo"><b>Hallazgo.</b> La producción del sistema es '
      f'predominantemente de <b>autoconsumo</b>: '
      f'{pct(destino_c["Autoconsumo"], tot_dc):.1f} % de las declaraciones agrícolas y '
      f'{pct(destino_a["Autoconsumo"], tot_da):.1f} % de las pecuarias se destinan a '
      'la alimentación familiar. El mercado es el segundo destino y la agroindustria '
      'y la exportación son marginales. <b>El riego sostiene aquí la seguridad '
      'alimentaria de las familias antes que una cadena comercial</b>, un dato central '
      'para dimensionar el impacto social del sistema.</div>')
    if sob:
        A(f'<p>Consistentemente, los titulares declaran destinar en promedio el '
          f'<b>{sob_media:.0f} % de su producción a la soberanía alimentaria</b> '
          f'(autoconsumo familiar), según el campo específico de la ficha '
          f'({len(sob):,} respuestas).</p>')

    A('<h2>5. Distribución territorial</h2>')
    A('<table class="evitar-corte"><tr><th>Sector</th><th class="n">Predios con cultivo</th>'
      '<th class="n">Superficie (ha)</th><th class="n">Cabezas</th></tr>')
    for sec in sorted(por_sector):
        if sec.startswith('('):
            continue
        d = por_sector[sec]
        A(f'<tr><td>{sec}</td><td class="n">{len(d["predios"]):,}</td>'
          f'<td class="n">{d["sup"] / 10000:,.1f}</td><td class="n">{d["cab"]:,}</td></tr>')
    A('</table>')

    A('<h2>6. Conclusiones</h2>')
    A('<ul>')
    A(f'<li>Se cultivan <b>{sup_total / 10000:,.0f} ha</b> con <b>{len(sup)} especies '
      f'vegetales</b> distintas, lo que indica un sistema productivo diversificado.</li>')
    A(f'<li><b>{pct(past, sup_total):.0f} % de la superficie son pastos</b>: el uso '
      'principal del agua es sostener la ganadería familiar.</li>')
    A(f'<li>El hato asciende a <b>{total_pec:,} animales</b>, dominado por especies '
      'menores (cuyes, aves, ovinos) propias de la producción de traspatio.</li>')
    A(f'<li>El <b>autoconsumo es el destino principal</b> '
      f'({pct(destino_c["Autoconsumo"], tot_dc):.0f} % de las declaraciones agrícolas): '
      'el sistema de riego sostiene la seguridad alimentaria de las familias.</li>')
    A('<li>Dos registros requieren verificación en campo antes de su uso oficial: la '
      'granja avícola de Asociación Rosalía y los cultivos declarados como "Otros".</li>')
    A('</ul>')
    A(E.pie(corte_txt))

    os.makedirs(os.path.dirname(HTML), exist_ok=True)
    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(E.documento('Producción agropecuaria — Padrón Guanguilquí–Porotog',
                            '\n'.join(B)))
    print(f'  capítulo: {os.path.relpath(HTML, BASE)}  (corte: {corte_txt})')

    # ── Excel ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    azul, blanco = PatternFill('solid', fgColor='1e4d8c'), Font(color='FFFFFF', bold=True)

    def hoja(nombre, cab, filas):
        ws = wb.create_sheet(nombre)
        ws.append(cab)
        for c in ws[1]:
            c.fill, c.font = azul, blanco
        for f_ in filas:
            ws.append(f_)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                12, min(40, max(len(str(c.value or '')) for c in col) + 2))

    hoja('Resumen', ['Indicador', 'Valor'], [
        ['Declaraciones de cultivo', len(cultivos)],
        ['Predios con cultivo', predios_c],
        ['Superficie cultivada (ha)', round(sup_total / 10000, 2)],
        ['Cultivos distintos', len(sup)],
        ['Declaraciones de ganado', len(animales)],
        ['Predios con ganado', predios_a],
        ['Animales de granja (sin la duplicada)', total_pec],
        ['Producción acuícola', cab_acui],
        ['Cabezas excluidas por duplicación', cabezas_dup],
        ['% autoconsumo (cultivos)', round(pct(destino_c['Autoconsumo'], tot_dc), 1)],
        ['% soberanía alimentaria declarada', round(sob_media, 1)],
    ])
    hoja('Cultivos', ['Cultivo', 'Superficie (ha)', 'Predios', 'Superficie media (m2)'],
         [[t, round(s / 10000, 3), frec[t], round(s / frec[t])]
          for t, s in sorted(sup.items(), key=lambda x: -x[1])])
    hoja('Grupos de cultivo', ['Grupo', 'Superficie (ha)', '% del total'],
         [[g, round(s / 10000, 2), round(pct(s, sup_total), 1)]
          for g, s in sorted(sup_grupo.items(), key=lambda x: -x[1])])
    hoja('Ganado', ['Especie', 'Cabezas', 'Predios'],
         [[e, n, reg_esp[e]] for e, n in sorted(cabezas.items(), key=lambda x: -x[1])])
    hoja('Destino', ['Destino', 'Menciones cultivos', 'Menciones ganado'],
         [[et, destino_c[et], destino_a[et]]
          for et in ('Autoconsumo', 'Mercado', 'Agroindustria', 'Exportación')])
    filas_sec = [[sec, len(d['predios']), round(d['sup'] / 10000, 2), d['cab']]
                 for sec, d in sorted(por_sector.items()) if not sec.startswith('(')]
    hoja('Por sector', ['Sector', 'Predios con cultivo', 'Superficie (ha)', 'Cabezas'], filas_sec)

    del wb['Sheet']
    os.makedirs(os.path.dirname(XLSX), exist_ok=True)
    wb.save(XLSX)
    # Sin esto, hay builds de Excel que heredan «sin relleno» del estilo
    # base y los colores no se pintan. Ver excel_compat.py.
    from excel_compat import aplicar_formatos
    aplicar_formatos(XLSX)
    print(f'  excel   : {os.path.relpath(XLSX, BASE)}')
    print(f'\n  {sup_total / 10000:,.0f} ha cultivadas | {total_cab:,} cabezas | '
          f'autoconsumo {pct(destino_c["Autoconsumo"], tot_dc):.0f}% | '
          f'excluidas {cabezas_dup:,} cabezas duplicadas')


if __name__ == '__main__':
    main()
