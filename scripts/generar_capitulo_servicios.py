# -*- coding: utf-8 -*-
"""
Capítulo del informe técnico: "Servicios básicos y hábitat".

Cubre la sección 3 de la ficha: agua de consumo, energía eléctrica, material de
la vivienda y altitud del predio.

APARTADO EN LEVANTAMIENTO
-------------------------
Es el único bloque de la ficha cuya cobertura está por debajo del 70 %: hay
1.395 fichas principales sin registro en ninguno de los tres campos. El
levantamiento de este apartado continúa, de modo que las cifras se presentan
siempre acompañadas de la base sobre la que se calculan: las viviendas.

Nunca se atribuye el vacío a personas o a la organización del trabajo: el
informe describe el estado del dato, no el desempeño de quien lo levanta.

SALIDAS
  docs/CAPITULO-servicios-basicos.html
  build_entrega/Servicios_Basicos.xlsx
"""

import os
import sqlite3
import statistics as st
import sys
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica, nombre_publico, normalizar  # noqa: E402
import informe_estilo as E  # noqa: E402

GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
T = 'Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e'
BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
HTML = os.path.join(BASE, 'docs', 'CAPITULO-servicios-basicos.html')
XLSX = os.path.join(BASE, 'build_entrega', 'Servicios_Basicos.xlsx')
MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
         'octubre noviembre diciembre').split()
SERV = ('agua_consumo', 'energia_electrica', 'material_construccion')


def lleno(v):
    return v not in (None, '') and str(v).strip() != ''


def si(v):
    return str(v).strip() in ('1', 'True', 'Sí', 'Si')


def pct(a, b):
    return 100.0 * a / b if b else 0.0


def main():
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(f'SELECT * FROM "{T}" WHERE es_ficha_hija IS NOT 1')
    pri = [dict(r) for r in cur.fetchall()]
    cur.execute(f'SELECT MAX(fecha_creacion), MAX(fecha_completado) FROM "{T}"')
    f1, f2 = cur.fetchone()
    cur.execute(f'SELECT COUNT(*) FROM "{T}" WHERE es_ficha_hija = 1 AND '
                f'coalesce(estado_investigacion, "pendiente_produccion") != "completada"')
    pendientes = cur.fetchone()[0]
    con.close()

    from generar_capas_sectores_comunidades import COM_A_SECTOR
    vistos = defaultdict(Counter)
    for p in pri:
        crudo = p.get('comunidad') or ''
        p['_comk'] = canonica(crudo) or '(sin comunidad)'
        vistos[p['_comk']][nombre_publico(crudo) or '(sin comunidad)'] += 1
        p['_sec'] = (p.get('sector_investigacion') or '').strip()
    display = {}
    for k, c in vistos.items():
        val = [(n_, v) for n_, v in c.most_common() if normalizar(n_) == k]
        display[k] = val[0][0] if val else k
    for p in pri:
        p['_com'] = display[p['_comk']]
        if p['_sec'] in ('', 'None'):
            p['_sec'] = COM_A_SECTOR.get(p['_comk'], '(sin sector)')

    corte = max(str(f1 or '')[:10], str(f2 or '')[:10])
    corte_txt = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                 if corte else 'la fecha de generación')

    N = len(pri)
    agua = [p for p in pri if lleno(p.get('agua_consumo'))]
    ener = [p for p in pri if lleno(p.get('energia_electrica'))]
    mat = Counter(str(p['material_construccion']).strip().title() for p in pri
                  if lleno(p.get('material_construccion')))
    n_mat = sum(mat.values())
    con_agua = sum(1 for p in agua if si(p['agua_consumo']))
    con_ener = sum(1 for p in ener if si(p['energia_electrica']))
    registrado = sum(1 for p in pri if any(lleno(p.get(k)) for k in SERV))
    cot = [float(p['cota_msnm']) for p in pri if lleno(p.get('cota_msnm'))]

    # cobertura por sector, para orientar el levantamiento restante
    cob_sector = {}
    for sec in sorted({p['_sec'] for p in pri if not p['_sec'].startswith('(')}):
        ps = [p for p in pri if p['_sec'] == sec]
        r = sum(1 for p in ps if any(lleno(p.get(k)) for k in SERV))
        cob_sector[sec] = (r, len(ps), pct(r, len(ps)))

    B = []
    A = B.append
    A(E.cabecera('Servicios básicos y hábitat',
                 'Agua de consumo, energía, vivienda y altitud · '
                 'Capítulo del informe técnico'))
    A(E.aviso_corte(corte_txt, N, pendientes))

    # El 64 % NO es avance de levantamiento (cerrado el 5-ago-2026): es la
    # proporción de predios CON VIVIENDA. Regla 2 del cliente (9-ago-2026):
    # sin material de construcción no hay vivienda, y entonces agua y luz
    # vacías son la respuesta correcta. Leerlo como cobertura pendiente
    # hacía parecer que a un tercio del padrón le falta el servicio.
    A('<div class="nota"><b>Sobre la base de cálculo.</b> Este bloque describe '
      f'la <b>vivienda</b> del predio: <b>{registrado:,} de {N:,} fichas '
      f'principales ({pct(registrado, N):.1f} %)</b> declaran una construcción. '
      'En los demás predios no hay vivienda, y por eso agua y energía figuran '
      'vacías: es la respuesta correcta, no un dato faltante (criterio del '
      'cliente, 9 de agosto de 2026). Los porcentajes de servicios se '
      'calculan <b>sobre las viviendas</b>, nunca sobre el total del padrón, '
      'y no deben presentarse como cobertura de servicios del sistema.</div>')

    A(E.kpis([
        (f'{pct(registrado, N):.0f}%', 'del padrón con este dato'),
        (f'{pct(con_agua, len(agua)):.1f}%', 'con agua de consumo'),
        (f'{pct(con_ener, len(ener)):.1f}%', 'con energía eléctrica'),
        (f'{st.median(cot):,.0f}', 'msnm (altitud mediana)'),
    ]))

    A('<h2>1. Estado del registro</h2>')
    A(f'<p>De los {N:,} predios con ficha principal, <b>{registrado:,} '
      f'({pct(registrado, N):.1f} %) tienen registrada al menos una de las tres '
      'variables</b> de este apartado. La cobertura por sector muestra dónde se '
      'concentra el levantamiento pendiente:</p>')
    A('<table class="evitar-corte"><tr><th>Sector</th><th class="n">Con registro</th>'
      '<th class="n">Predios</th><th>Cobertura</th></tr>')
    for sec, (r, t, p) in cob_sector.items():
        A(f'<tr><td>{sec}</td><td class="n">{r:,}</td><td class="n">{t:,}</td>'
          f'<td>{E.barra(p)}</td></tr>')
    A('</table>')

    A('<h2>2. Agua de consumo y energía eléctrica</h2>')
    A('<table class="evitar-corte"><tr><th>Servicio</th><th class="n">Dispone</th>'
      '<th class="n">No dispone</th><th class="n">Base</th><th>Cobertura</th></tr>')
    A(f'<tr><td>Agua de consumo</td><td class="n">{con_agua:,}</td>'
      f'<td class="n">{len(agua) - con_agua:,}</td><td class="n">{len(agua):,}</td>'
      f'<td>{E.barra(pct(con_agua, len(agua)))}</td></tr>')
    A(f'<tr><td>Energía eléctrica</td><td class="n">{con_ener:,}</td>'
      f'<td class="n">{len(ener) - con_ener:,}</td><td class="n">{len(ener):,}</td>'
      f'<td>{E.barra(pct(con_ener, len(ener)))}</td></tr>')
    A('</table>')
    A(f'<p>Sobre los predios ya registrados, la cobertura de ambos servicios es '
      f'prácticamente universal: <b>{pct(con_agua, len(agua)):.1f} % dispone de agua '
      f'de consumo</b> y <b>{pct(con_ener, len(ener)):.1f} % de energía eléctrica</b>. '
      f'Los casos sin servicio son {len(agua) - con_agua} y {len(ener) - con_ener} '
      'respectivamente, cifras reducidas pero identificables predio a predio para '
      'una eventual intervención focalizada.</p>')
    A('<div class="nota"><b>Alcance de la cifra.</b> Estos porcentajes describen a '
      'los predios <b>ya registrados</b> en este apartado. No pueden extrapolarse al '
      'total del padrón mientras el levantamiento siga en curso.</div>')

    A('<h2>3. Material de la vivienda</h2>')
    A(f'<p>Se registró el material predominante de la vivienda en {n_mat:,} '
      'predios:</p>')
    A('<table class="evitar-corte"><tr><th>Material</th><th class="n">Viviendas</th>'
      '<th>Peso</th></tr>')
    for k, n in mat.most_common():
        A(f'<tr><td>{k}</td><td class="n">{n:,}</td>'
          f'<td>{E.barra(pct(n, n_mat))}</td></tr>')
    A('</table>')
    trad = sum(n for k, n in mat.items() if k.lower() in ('tapia', 'adobe', 'madera'))
    A(f'<p>Predomina el <b>bloque</b> ({pct(mat.get("Bloque", 0), n_mat):.1f} %), '
      f'seguido del hormigón armado ({pct(mat.get("Hormigón Armado", 0), n_mat):.1f} %). '
      f'Las construcciones de materiales tradicionales —tapia, adobe y madera— '
      f'representan el {pct(trad, n_mat):.1f} % de las viviendas registradas.</p>')

    A('<h2>4. Altitud de los predios</h2>')
    A(f'<p>La cota está registrada en la totalidad de los predios '
      f'({len(cot):,} registros). El sistema se despliega entre los '
      f'<b>{min(cot):,.0f} y los {max(cot):,.0f} msnm</b>, con una mediana de '
      f'<b>{st.median(cot):,.0f} msnm</b>.</p>')
    tramos = [('Bajo 3.000 m', 0, 3000), ('3.000 – 3.200 m', 3000, 3200),
              ('3.200 – 3.400 m', 3200, 3400), ('3.400 – 3.600 m', 3400, 3600),
              ('Sobre 3.600 m', 3600, 9999)]
    A('<table class="evitar-corte"><tr><th>Franja altitudinal</th>'
      '<th class="n">Predios</th><th>Peso</th></tr>')
    for et, lo, hi in tramos:
        n = sum(1 for x in cot if lo <= x < hi)
        A(f'<tr><td>{et}</td><td class="n">{n:,}</td>'
          f'<td>{E.barra(pct(n, len(cot)))}</td></tr>')
    A('</table>')
    A('<p>El rango altitudinal de más de mil metros condiciona los cultivos '
      'posibles y los requerimientos de riego en cada franja, y explica la '
      'diversidad de especies descrita en el capítulo de producción.</p>')

    A('<h2>5. Conclusiones</h2>')
    A('<ul>')
    A(f'<li><b>{pct(registrado, N):.0f} % de las fichas principales declara '
      'una vivienda</b> en el predio; el resto son predios sin construcción y '
      'quedan fuera del cálculo de servicios.</li>')
    A(f'<li>Entre los predios ya registrados, la cobertura de <b>agua de consumo '
      f'({pct(con_agua, len(agua)):.1f} %) y energía eléctrica '
      f'({pct(con_ener, len(ener)):.1f} %) es prácticamente universal</b>.</li>')
    A(f'<li>La vivienda es mayoritariamente de <b>bloque</b> '
      f'({pct(mat.get("Bloque", 0), n_mat):.0f} %); los materiales tradicionales '
      f'persisten en el {pct(trad, n_mat):.0f} % de los casos.</li>')
    A(f'<li>El sistema abarca desde los {min(cot):,.0f} hasta los {max(cot):,.0f} '
      'msnm, un rango que condiciona la aptitud productiva de cada zona.</li>')
    A('</ul>')
    A(E.pie(corte_txt))

    os.makedirs(os.path.dirname(HTML), exist_ok=True)
    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(E.documento('Servicios básicos y hábitat — Padrón Guanguilquí–Porotog',
                            '\n'.join(B)))
    print(f'  capítulo: {os.path.relpath(HTML, BASE)}  (corte: {corte_txt})')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    azul, blanco = PatternFill('solid', fgColor='1e4d8c'), Font(color='FFFFFF', bold=True)

    def hoja(nombre, cab, filas):
        ws = wb.create_sheet(nombre)
        ws.append(cab)
        for c in ws[1]:
            c.fill, c.font = azul, blanco
        for f_ in filas:
            ws.append(f_)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                12, min(40, max(len(str(c.value or '')) for c in col) + 2))

    hoja('Resumen', ['Indicador', 'Valor'], [
        ['Predios con ficha principal', N],
        ['Con el apartado registrado', registrado],
        ['% de registro', round(pct(registrado, N), 1)],
        ['% con agua (sobre registrados)', round(pct(con_agua, len(agua)), 1)],
        ['% con energía (sobre registrados)', round(pct(con_ener, len(ener)), 1)],
        ['Altitud mediana (msnm)', round(st.median(cot))],
    ])
    hoja('Cobertura por sector', ['Sector', 'Con registro', 'Predios', '% registro'],
         [[s, r, t, round(p, 1)] for s, (r, t, p) in cob_sector.items()])
    hoja('Materiales', ['Material', 'Viviendas', '%'],
         [[k, n, round(pct(n, n_mat), 1)] for k, n in mat.most_common()])
    filas = []
    for com in sorted({p['_com'] for p in pri}):
        ps = [p for p in pri if p['_com'] == com]
        r = [p for p in ps if any(lleno(p.get(k)) for k in SERV)]
        a_ = [p for p in ps if lleno(p.get('agua_consumo'))]
        e_ = [p for p in ps if lleno(p.get('energia_electrica'))]
        filas.append([com, ps[0]['_sec'], len(ps), len(r), round(pct(len(r), len(ps)), 1),
                      round(pct(sum(1 for p in a_ if si(p['agua_consumo'])), len(a_)), 1) if a_ else None,
                      round(pct(sum(1 for p in e_ if si(p['energia_electrica'])), len(e_)), 1) if e_ else None])
    hoja('Por comunidad', ['Comunidad', 'Sector', 'Predios', 'Con registro',
                           '% registro', '% agua', '% energía'], filas)

    del wb['Sheet']
    os.makedirs(os.path.dirname(XLSX), exist_ok=True)
    wb.save(XLSX)
    # Sin esto, hay builds de Excel que heredan «sin relleno» del estilo
    # base y los colores no se pintan. Ver excel_compat.py.
    from excel_compat import aplicar_formatos
    aplicar_formatos(XLSX)
    print(f'  excel   : {os.path.relpath(XLSX, BASE)}')
    print(f'\n  registro {pct(registrado, N):.0f}% | agua {pct(con_agua, len(agua)):.1f}% | '
          f'energía {pct(con_ener, len(ener)):.1f}% (sobre registrados)')


if __name__ == '__main__':
    main()
