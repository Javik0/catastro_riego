# -*- coding: utf-8 -*-
"""
Hoja de trabajo en Excel para la revisión de campo, con seguimiento automático.

Por qué un Excel y por qué así
------------------------------
`REVISION-CAMPO.md` sirve para leer y planificar, pero no para trabajar: no se
filtra, no se ordena y no deja constancia de lo revisado. Esto es lo mismo en
formato de hoja de cálculo, con una regla de diseño que conviene entender antes
de tocarlo:

    **El Excel NO es donde se corrige. Se corrige en QField.**

Si el técnico llena el Excel, ese dato no llega a ninguna parte: el padrón sale
del `data.gpkg`, y lo que no esté ahí no existe para los informes. Peor aún,
tendríamos dos versiones del mismo dato sin saber cuál manda. Por eso las
columnas de trabajo son de *seguimiento* (¿ya fui?, ¿qué pasó?), no de datos.

Cómo evita que se pierdan
-------------------------
El seguimiento no depende de que nadie marque nada:

* Cada vez que se ejecuta, vuelve a leer el `data.gpkg`. Lo que ya se corrigió
  **desaparece de Pendientes y aparece en la hoja «Resueltos»**, con la fecha en
  que se detectó. El avance se ve solo.
* Las notas escritas a mano (Estado, Fecha, Observación) **se conservan** entre
  ejecuciones: se leen del Excel anterior y se vuelven a colocar en la fila que
  les corresponde. Se puede regenerar sin miedo a perder el trabajo del revisor.

Hojas
-----
1. **Instrucciones** — cómo se usa, en una pantalla.
2. **Resumen** — cuánto falta en cada comunidad, para decidir la ruta.
3. **Pendientes** — una fila por ficha y por dato faltante. Filtros activados.
4. **Fuera de campo** — lo que falta pero no lo resuelve el técnico: las fichas
   sin comunidad (oficina) y lo que espera una decisión de la coordinación.
5. **Resueltos** — lo que se ha ido corrigiendo, con la fecha.

Qué cuenta como pendiente
-------------------------
Lo define `generar_revision_campo.py` y este script lo importa: las reglas de
«no aplica» del 9 de agosto de 2026 viven en un solo sitio. Ver allí.

Cómo se ejecuta
---------------
Necesita `osgeo`, que en esta máquina solo está en el Python de OSGeo4W::

    "C:\\OSGeo4W\\bin\\python-qgis.bat" -X utf8 scripts/generar_excel_revision_campo.py

Salida
------
docs/REVISION-CAMPO.xlsx
"""
import json
import os
import sys
from datetime import datetime
from importlib import util as _util

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from osgeo import ogr

ogr.UseExceptions()

BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
SALIDA = os.path.join(BASE, 'docs', 'REVISION-CAMPO.xlsx')
HISTORICO = os.path.join(BASE, 'docs', '.revision-campo-historico.json')

# La definición de «qué es un pendiente» vive en un solo sitio: el script que
# genera el Markdown. Duplicarla aquí garantizaría que un día digan cosas
# distintas y nadie sepa cuál vale.
_spec = _util.spec_from_file_location(
    'revision_campo', os.path.join(os.path.dirname(__file__),
                                   'generar_revision_campo.py'))
_rc = _util.module_from_spec(_spec)
_spec.loader.exec_module(_rc)

ESTADOS = ['Pendiente', 'Corregido en QField', 'No aplica',
           'No se pudo (anotar por qué)', 'Predio ya no existe']

AZUL = 'FF1F4E79'
GRIS = 'FFF2F2F2'
VERDE = 'FFE2EFDA'
AMBAR = 'FFFFF2CC'
ROJO = 'FFFCE4E4'

FINA = Side(style='thin', color='FFBFBFBF')
BORDE = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)


def relleno(color):
    """Relleno sólido que se ve en cualquier visor.

    Con solo fgColor, el bgColor queda en negro-transparente y hay versiones
    de Excel/WPS que resuelven el sólido contra el fondo: la celda sale
    blanca aunque el color esté escrito. Fijar los dos colores lo pinta en
    todos.
    """
    return PatternFill(fill_type='solid', start_color=color, end_color=color)


def cabecera(ws, fila, titulos, anchos):
    for i, (t, a) in enumerate(zip(titulos, anchos), start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font = Font(bold=True, color='FFFFFFFF', size=10)
        c.fill = relleno(AZUL)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDE
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.row_dimensions[fila].height = 28


def leer_notas_previas():
    """Estado, fecha y observación que el revisor escribió en la versión anterior."""
    if not os.path.exists(SALIDA):
        return {}
    try:
        wb = load_workbook(SALIDA, read_only=True, data_only=True)
        if 'Pendientes' not in wb.sheetnames:
            return {}
        ws = wb['Pendientes']
        notas = {}
        encabezados = None
        for fila in ws.iter_rows(values_only=True):
            if encabezados is None:
                if fila and fila[0] == 'Comunidad':
                    encabezados = list(fila)
                continue
            if not fila or len(fila) < 10 or not fila[9]:
                continue
            clave = '{}|{}'.format(fila[9], fila[4])       # id | qué falta
            estado, fecha, obs = (fila[6] if len(fila) > 6 else None,
                                  fila[7] if len(fila) > 7 else None,
                                  fila[8] if len(fila) > 8 else None)
            if any([estado and estado != 'Pendiente', fecha, obs]):
                notas[clave] = (estado, fecha, obs)
        wb.close()
        return notas
    except Exception as e:
        print("      aviso: no se pudo leer el Excel anterior ({}); se generará "
              "uno limpio".format(e))
        return {}


def main():
    print("=" * 74)
    print(" HOJA DE TRABAJO EN EXCEL - REVISION DE CAMPO")
    print("=" * 74)

    if not os.path.exists(_rc.GPKG):
        print("ERROR: no se encuentra el data.gpkg de QField.")
        return 1

    notas = leer_notas_previas()
    print("\n  notas del revisor recuperadas de la version anterior: {}"
          .format(len(notas)))

    # El histórico solo vale si se escribió con el mismo tipo de clave. Si el
    # esquema cambió (pasó al cambiar de `codigo_final` al `id` de QField), las
    # claves viejas no se parecen a las nuevas y TODO figuraría como resuelto de
    # golpe. Mejor empezar de cero que anunciar un avance que no existe.
    #
    # El sufijo cambia también cuando cambia la DEFINICIÓN de pendiente, no solo
    # el identificador. Al aplicar las reglas del 9-ago-2026 desaparecieron más
    # de 8.000 filas de golpe: sin versionar esto, la hoja «Resueltos» habría
    # anunciado 8.000 correcciones que nadie hizo. No se resolvieron, se dejaron
    # de contar.
    ESQUEMA = 'id|falta|reglas-2026-08-09'
    previos = {}
    if os.path.exists(HISTORICO):
        try:
            with open(HISTORICO, encoding='utf-8') as f:
                cargado = json.load(f)
            if cargado.get('esquema_claves') == ESQUEMA:
                previos = cargado
            else:
                print("      el historico anterior usaba otro identificador; se "
                      "descarta para no marcar todo como resuelto")
        except Exception:
            previos = {}

    ds = ogr.Open(_rc.GPKG, 0)              # SOLO LECTURA
    t = '"{}"'.format(_rc.CAPA)
    corte = (_rc.consultar(ds, "SELECT MAX(fecha_creacion) FROM {}".format(t))[0][0]
             or '')[:10]
    total = _rc.consultar(ds, "SELECT COUNT(*) FROM {}".format(t))[0][0]

    # ── pendientes actuales ──
    # El identificador de seguimiento es el `id` de QField, no el «código» de la
    # ficha: `codigo_final` vale S-C-P001 en 5.529 de las 6.831 fichas (es el
    # valor por defecto del formulario y casi nunca se cambió), así que usarlo
    # mezclaría el estado de miles de fichas distintas.
    UID = ", COALESCE(id,'') uid"

    def recoger(cond, etq, universo=None):
        salida = []
        for com, clave, nombre, ced, tec, uid in _rc.fichas_de(
                ds, t, cond, universo, extra=UID):
            falta = etq.replace('Sin ', '')
            salida.append({
                'comunidad': com, 'clave': clave, 'regante': nombre,
                'cedula': ced, 'falta': falta, 'tecnico': tec, 'uid': uid,
                'clave_seg': '{}|{}'.format(uid or clave, falta),
            })
        return salida

    # Lo que sí es trabajo de campo. `DE_CAMPO` deja fuera las fichas hijas y
    # las comunidades que no son encuestas (ALPAKA).
    filas = []
    for campo, etq, cond, _listar in _rc.PENDIENTES:
        filas.extend(recoger(cond, etq))

    # Lo que falta pero no lo resuelve el técnico. Va en su propia hoja: si se
    # mezclara con lo anterior, volveríamos al problema que estas reglas
    # vinieron a corregir — mandar gente a campo a resolver lo que no está allá.
    fuera = []
    for f in recoger(_rc.OFICINA[2], _rc.OFICINA[1], _rc.PRINCIPALES):
        f['donde'] = 'OFICINA — cruce espacial'
        f['nota'] = ('Se asigna por traslape con la capa de comunidades. Ver '
                     'scripts/represa/06_capas_padron.py. Nunca cruzar por nombre.')
        fuera.append(f)
    for f in recoger(_rc.EN_ESPERA[2], _rc.EN_ESPERA[1], _rc.PRINCIPALES):
        f['donde'] = 'EN ESPERA — decisión del cliente'
        f['nota'] = 'Los datos de escritura se revisan después de la depuración.'
        fuera.append(f)
    for com, motivo in sorted(_rc.NO_SON_ENCUESTAS.items()):
        solo = "UPPER(TRIM(COALESCE(comunidad,'')))='{}'".format(com)
        for campo, etq, cond, _l in _rc.PENDIENTES:
            for f in recoger(cond, etq,
                             '{} AND {}'.format(_rc.PRINCIPALES, solo)):
                f['donde'] = 'EN ESPERA — decisión de dirección'
                f['nota'] = 'No salir a completarlas: {}.'.format(motivo)
                fuera.append(f)

    # Las coordenadas solo sirven para las de oficina, que es donde se hace el
    # cruce; para las demás la columna queda vacía a propósito.
    coords = {r[0]: (r[1], r[2]) for r in _rc.consultar(ds,
        "SELECT COALESCE(id,''), COALESCE(coord_x_utm,0), COALESCE(coord_y_utm,0) "
        "FROM {} WHERE {}".format(t, _rc.OFICINA[2]))}

    # Cuántos datos daba el criterio anterior (todo campo vacío = pendiente).
    # Se calcula, no se escribe a mano: si el gpkg cambia, la comparación de las
    # instrucciones sigue siendo cierta.
    anterior = sum(_rc.contar(ds, t, c, _rc.PRINCIPALES)
                   for c in _rc.CRITERIO_ANTERIOR)
    ds = None

    actuales = {f['clave_seg'] for f in filas}
    hoy = datetime.now().strftime('%d/%m/%Y')

    # ── resueltos: estaban en la corrida anterior y ya no están ──
    resueltos = dict(previos.get('resueltos', {}))
    antes = set(previos.get('pendientes', []))
    nuevos_resueltos = antes - actuales
    for k in nuevos_resueltos:
        resueltos.setdefault(k, hoy)
    print("  pendientes ahora        : {:,}".format(len(actuales)))
    print("  resueltos desde la ultima ejecucion: {:,}".format(len(nuevos_resueltos)))
    print("  resueltos acumulados    : {:,}".format(len(resueltos)))

    wb = Workbook()

    # ── 1. Instrucciones ──
    ws = wb.active
    ws.title = 'Instrucciones'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 110
    lineas = [
        ('Revisión de campo — Padrón Guanguilquí–Porotog', 'titulo'),
        ('Última ficha en el sistema: {}  ·  Generado: {}'.format(corte, hoy), 'sub'),
        ('', ''),
        ('LO PRIMERO, PORQUE ES LO QUE MÁS SE CONFUNDE', 'h'),
        ('Las correcciones se hacen EN QFIELD, no en este Excel. Lo que se escriba '
         'aquí no llega al padrón: los informes salen del archivo de QField.', ''),
        ('Este archivo sirve para saber qué falta, repartir el trabajo y dejar '
         'constancia de lo que se revisó.', ''),
        ('', ''),
        ('UN CAMPO VACÍO NO SIEMPRE ES UN PENDIENTE', 'h'),
        ('Desde el 9 de agosto de 2026 este archivo aplica las reglas de «no aplica» '
         'que acordó la coordinación del proyecto. Con el criterio anterior salían '
         '{:,} datos faltantes; el trabajo real de campo son {:,}.'
         .format(anterior, len(filas)).replace(',', '.'), ''),
        ('· Un predio SIN MATERIAL DE CONSTRUCCIÓN no tiene vivienda: que agua y luz '
         'estén vacías es la respuesta correcta. Solo se pide cuando hay material '
         'declarado y falta uno de los dos.', ''),
        ('· Un predio con ÁREA SIN RIEGO medida no tiene por qué tener área de riego: '
         'existe y no se riega.', ''),
        ('· El CAUDAL se hereda de la comuna, no se mide ficha por ficha.', ''),
        ('· La FOTO DEL PREDIO se eliminó del levantamiento desde el inicio.', ''),
        ('· CÉDULA y TELÉFONO sí se mantienen, aunque no se puedan obtener: la '
         'coordinación pidió dejar constancia del dato que falta.', ''),
        ('Estas reglas no se cambian sin consultar con la coordinación del proyecto.', ''),
        ('', ''),
        ('CÓMO SE USA', 'h'),
        ('1. En la hoja «Resumen» mira qué comunidades tienen más pendientes y arma '
         'la salida por ahí.', ''),
        ('2. En la hoja «Pendientes» filtra por tu comunidad (o por tu usuario en la '
         'columna «Levantó»). Cada fila es una ficha y un dato que le falta.', ''),
        ('3. Busca la ficha en QField por su CLAVE CATASTRAL y completa el dato ahí.', ''),
        ('   OJO: no busques por el «código» de la ficha. Ese campo vale S-C-P001 en '
         '5.529 de las 6.831 fichas (es el valor por defecto del formulario), así que '
         'no sirve para encontrar nada. La clave catastral sí está en todas.', ''),
        ('4. Vuelve a esta hoja y marca el ESTADO. Si no se pudo, escribe por qué en '
         'OBSERVACIÓN: un campo vacío no distingue entre «falta ir» y «no aplica».', ''),
        ('', ''),
        ('QUÉ PASA CUANDO SE VUELVE A GENERAR ESTE ARCHIVO', 'h'),
        ('Lo que ya se corrigió en QField desaparece de «Pendientes» y pasa a la hoja '
         '«Resueltos» con la fecha. No hay que marcar nada para que eso ocurra.', ''),
        ('Tus notas (Estado, Fecha, Observación) se conservan: se leen del archivo '
         'anterior y se vuelven a colocar. Se puede regenerar sin perder trabajo.', ''),
        ('', ''),
        ('LA HOJA «FUERA DE CAMPO»', 'h'),
        ('Son datos que faltan pero que el técnico no resuelve. Las fichas sin '
         'comunidad se asignan en OFICINA por cruce espacial (todas tienen '
         'coordenadas); la tenencia del predio y las fichas de ALPAKA esperan una '
         'decisión de la coordinación.', ''),
        ('Están en su propia hoja para que no se cuelen en la ruta de campo, pero a '
         'la vista para que nadie las dé por perdidas.', ''),
        ('', ''),
        ('LO QUE NO HAY QUE HACER', 'h'),
        ('· No crear una ficha nueva para completar datos: se duplica el predio.', ''),
        ('· No borrar filas de este archivo: se regenera solo.', ''),
        ('· Las fichas de la hoja «Fuera de campo» no se tocan sin consultar con la '
         'dirección del proyecto.', ''),
        ('', ''),
        ('Se regenera con:  python scripts/generar_excel_revision_campo.py', 'sub'),
    ]
    r = 2
    for texto, tipo in lineas:
        c = ws.cell(row=r, column=2, value=texto)
        if tipo == 'titulo':
            c.font = Font(bold=True, size=16, color=AZUL)
        elif tipo == 'h':
            c.font = Font(bold=True, size=11, color=AZUL)
        elif tipo == 'sub':
            c.font = Font(italic=True, size=9, color='FF808080')
        else:
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1

    # ── 2. Resumen por comunidad ──
    ws = wb.create_sheet('Resumen')
    cabecera(ws, 1, ['Comunidad', 'Datos por completar', 'Fichas afectadas', 'Aviso'],
             [38, 13, 17, 62])
    por_com = {}
    for f in filas:
        d = por_com.setdefault(f['comunidad'], {'n': 0, 'fichas': set()})
        d['n'] += 1
        d['fichas'].add(f['uid'] or f['clave'])
    r = 2
    for com in sorted(por_com, key=lambda c: -por_com[c]['n']):
        d = por_com[com]
        ws.cell(row=r, column=1, value=com).border = BORDE
        ws.cell(row=r, column=2, value=d['n']).border = BORDE
        ws.cell(row=r, column=3, value=len(d['fichas'])).border = BORDE
        nota = ''
        if com == '(sin comunidad)':
            nota = ('Primero hay que asignarles comunidad en oficina; hasta '
                    'entonces no se pueden meter en ninguna ruta. Ver la hoja '
                    '«Fuera de campo».')
        c = ws.cell(row=r, column=4, value=nota)
        c.border = BORDE
        c.alignment = Alignment(wrap_text=True, vertical='center')
        if nota:
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = relleno(AMBAR)
        r += 1

    # Las comunidades que no son encuestas ya no cuentan como pendiente, pero
    # tienen que seguir viéndose: si desaparecen de la vista, dentro de un mes
    # alguien las vuelve a contar como trabajo de campo.
    for com in sorted(_rc.NO_SON_ENCUESTAS):
        # solo lo que quedó fuera POR no ser encuesta; una ficha de ALPAKA a la
        # que además le falte la tenencia se cuenta en su propia categoría
        suyas = [f for f in fuera
                 if f['comunidad'].strip().upper() == com
                 and f['donde'].endswith('dirección')]
        if not suyas:
            continue
        ws.cell(row=r, column=1, value=com).border = BORDE
        ws.cell(row=r, column=2, value=len(suyas)).border = BORDE
        ws.cell(row=r, column=3,
                value=len({f['uid'] or f['clave'] for f in suyas})).border = BORDE
        c = ws.cell(row=r, column=4,
                    value='NO CUENTAN COMO PENDIENTE — ' +
                          _rc.NO_SON_ENCUESTAS[com] +
                          '. Están en la hoja «Fuera de campo».')
        c.border = BORDE
        c.alignment = Alignment(wrap_text=True, vertical='center')
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = relleno(ROJO)
        r += 1

    ws.auto_filter.ref = 'A1:D{}'.format(r - 1)
    ws.freeze_panes = 'A2'

    # ── 3. Pendientes ──
    ws = wb.create_sheet('Pendientes')
    titulos = ['Comunidad', 'Clave catastral', 'Regante', 'Cédula',
               'Qué falta', 'Levantó', 'ESTADO', 'Fecha revisión', 'OBSERVACIÓN',
               'id (no tocar)']
    cabecera(ws, 1, titulos, [30, 18, 34, 13, 24, 16, 24, 14, 46, 12])
    r = 2
    for f in sorted(filas, key=lambda x: (x['comunidad'], x['clave'], x['falta'])):
        estado, fecha, obs = notas.get(f['clave_seg'], ('Pendiente', None, None))
        vals = [f['comunidad'], f['clave'], f['regante'], f['cedula'],
                f['falta'], f['tecnico'], estado or 'Pendiente', fecha, obs,
                f['uid']]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDE
            c.font = Font(size=10)
            if i in (7, 8, 9):
                c.fill = relleno(GRIS)
        if estado and estado != 'Pendiente':
            ws.cell(row=r, column=7).fill = relleno(VERDE)
        r += 1

    dv = DataValidation(type='list', formula1='"{}"'.format(','.join(ESTADOS)),
                        allow_blank=True, showDropDown=False)
    dv.error = 'Elige una opción de la lista'
    ws.add_data_validation(dv)
    dv.add('G2:G{}'.format(max(r - 1, 2)))
    ws.auto_filter.ref = 'A1:J{}'.format(max(r - 1, 1))
    ws.freeze_panes = 'C2'
    # el id es lo que permite reencontrar cada fila al regenerar el archivo y
    # devolverle sus notas; se deja oculto para no invitar a editarlo
    ws.column_dimensions['J'].hidden = True

    # ── 4. Fuera de campo ──
    # Estas filas NO llevan columna de ESTADO a propósito: no son tareas del
    # técnico y marcarlas «corregido» daría una idea falsa de avance.
    ws = wb.create_sheet('Fuera de campo')
    cabecera(ws, 1, ['Se resuelve en', 'Comunidad', 'Clave catastral', 'Regante',
                     'Cédula', 'Qué falta', 'Levantó', 'X (UTM 17S)',
                     'Y (UTM 17S)', 'Por qué no es trabajo de campo'],
             [30, 26, 18, 32, 13, 22, 14, 13, 13, 60])
    r = 2
    for f in sorted(fuera, key=lambda x: (x['donde'], x['comunidad'],
                                          x['clave'], x['falta'])):
        x, y = coords.get(f['uid'], ('', ''))
        vals = [f['donde'], f['comunidad'], f['clave'], f['regante'], f['cedula'],
                f['falta'], f['tecnico'], x or '', y or '', f['nota']]
        color = AMBAR if f['donde'].startswith('OFICINA') else ROJO
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDE
            c.font = Font(size=10)
            if i in (1, 10):
                c.fill = relleno(color)
                c.alignment = Alignment(wrap_text=True, vertical='center')
            if i in (8, 9) and v:
                c.number_format = '#,##0'
        r += 1
    if r == 2:
        ws.cell(row=2, column=1, value='Nada fuera de campo.').font = Font(
            italic=True, color='FF808080')
    ws.auto_filter.ref = 'A1:J{}'.format(max(r - 1, 1))
    ws.freeze_panes = 'C2'

    # ── 5. Resueltos ──
    ws = wb.create_sheet('Resueltos')
    cabecera(ws, 1, ['Código', 'Qué faltaba', 'Detectado como resuelto el'],
             [18, 30, 26])
    r = 2
    for k in sorted(resueltos, key=lambda x: resueltos[x], reverse=True):
        cod, _, falta = k.partition('|')
        for i, v in enumerate([cod, falta, resueltos[k]], start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDE
            c.fill = relleno(VERDE)
            c.font = Font(size=10)
        r += 1
    if r == 2:
        ws.cell(row=2, column=1,
                value='Todavía no hay nada resuelto: esta hoja se llena sola '
                      'cuando se vuelva a generar el archivo.').font = Font(
            italic=True, color='FF808080')
    ws.auto_filter.ref = 'A1:C{}'.format(max(r - 1, 1))
    ws.freeze_panes = 'A2'

    wb.save(SALIDA)

    with open(HISTORICO, 'w', encoding='utf-8') as f:
        json.dump({'esquema_claves': ESQUEMA, 'generado': hoy, 'corte': corte,
                   'total_fichas': total, 'pendientes': sorted(actuales),
                   'resueltos': resueltos},
                  f, ensure_ascii=False, indent=1)

    print("\n  hojas    : Instrucciones · Resumen · Pendientes · Fuera de campo "
          "· Resueltos")
    print("  campo    : {:,} datos por completar en {} comunidades"
          .format(len(filas), len(por_com)))
    print("  fuera    : {:,} filas (oficina {:,} · en espera {:,})"
          .format(len(fuera),
                  sum(1 for f in fuera if f['donde'].startswith('OFICINA')),
                  sum(1 for f in fuera if f['donde'].startswith('EN ESPERA'))))
    print("  guardado : {} ({:,.0f} KB)"
          .format(os.path.relpath(SALIDA, BASE), os.path.getsize(SALIDA) / 1024))
    print("=" * 74)
    return 0


if __name__ == '__main__':
    sys.exit(main())
