# -*- coding: utf-8 -*-
"""
Informe por Sector de Investigación — material del sociólogo del proyecto.

QUÉ ES
------
El hermano del informe por comunidad (`generar_informe_sociologo.py`): UN solo
documento que reproduce LOS GRÁFICOS DEL DASHBOARD de la aplicación web —«los
gráficos del Dashboard van muy bien en el informe», dijo el cliente— tal como
se ven con el filtro de sector puesto, para el sistema completo y para cada
sector de investigación (Sector 1, 2 y 3). Encargo de JAVIKO, 31-ago-2026.

Los 10 gráficos reproducidos (mismos títulos, colores, universos y redondeos
que `src/components/dashboard/DashboardHome.tsx`):

  1. Uso del Suelo: Con Riego vs Sin Riego (ha)
  2. Especies Pecuarias Principales (Cabezas)
  3. Destino de la Producción Agrícola
  4. Nivel de Instrucción
  5. Hijos por Familia
  6. Represa y Capacitación
  7. Método de Riego (promedio %)
  8. Cultivos Más Frecuentes
  9. Tenencia del Predio
 10. Fichas por Parroquia

«Fichas por Técnico» y «Fichas Investigadas por Día» NO van: son seguimiento
interno del equipo (el Dashboard ya los restringe a admin/tecnico) y el
contratante ve el padrón, no cómo se repartió el trabajo.

DECISIONES DE JAVIKO (31-ago-2026) QUE ESTE SCRIPT APLICA
---------------------------------------------------------
· «Uso del Suelo» usa la medición CATASTRAL de las comunidades del sector,
  igual que la web con el filtro puesto, y lo etiqueta así. La superficie
  DECLARADA sigue siendo el universo de la narrativa (regla 12: las dos
  familias se citan, nunca se suman ni se mezclan sin nombre).
· La granja avícola de Asociación Rosalía (registros de ≥10.000 aves por
  titular sobre el mismo predio) se EXCLUYE del gráfico pecuario, con nota.
  El Dashboard web no la excluye: en el sector de Rosalía el gráfico difiere
  de la web a propósito, y la nota lo dice.
· El documento lleva un capítulo «Todo el sistema» (los mismos gráficos sin
  filtro) antes de los tres sectores.
· ASOCIACIÓN ROSALÍA cuenta en el SECTOR 3, como la web y como el catálogo de
  `constants.ts` («en campo son del Sector 3»). El bloque `sectores` de
  superficie_por_comunidad.json todavía la cuenta en el Sector 2 — bug
  heredado de la lista duplicada de generar_capas_sectores_comunidades.py,
  REPORTADO y pendiente de corregir allí; este script NO lo replica y avisa
  en consola de la diferencia esperada (47 fichas / 36,97 ha entre S2 y S3).

REGLAS DURAS DEL PROYECTO QUE ESTE SCRIPT RESPETA
-------------------------------------------------
· Regla 3: el caudal NO se suma ficha a ficha (fuente única
  caudal_por_comunidad.json; los heredados se muestran y no se suman).
· Regla 4: el nombre de comunidad se canoniza SOLO con comunidades_canon.py.
· Regla 6: personas ≠ predios. Instrucción, hijos, represa/capacitación y
  TENENCIA salen SOLO de fichas principales. (El gráfico de tenencia del
  Dashboard cuenta todas las fichas; aquí manda la regla 6 y la nota al pie
  documenta la diferencia con la web.)
· Regla 9: «sin riego», nunca «secano».
· Regla 12: superficie catastral y declarada no se mezclan; cada gráfico y
  cada cifra nombra su familia.
· Las fichas hijas PENDIENTES se excluyen (hoy 0; se filtran igual).
· NO se toca el data.gpkg: todo se lee de los GeoJSON de public/geo/.
· FECHA_CORTE editorial compartida: «19 de agosto de 2026».

SALIDAS
-------
  docs/INFORME-SOCIOLOGO-por-sector.html       documento imprimible (estilo casa)
  docs/INFORME-SOCIOLOGO-por-sector.md         fuente Markdown (para md_a_docx.py)
  docs/graficos-sociologo/*.png|.jpg           gráficos y mapas para el Word
  build_entrega/Informe_Sociologo_Sectores.xlsx  matrices crudas por gráfico

USO
---
  python -X utf8 scripts/generar_informe_sociologo_sector.py
  python scripts/md_a_docx.py docs/INFORME-SOCIOLOGO-por-sector.md

Se corre con el Python del PATH (C:\\Python314): nada de aquí lee el data.gpkg.
"""

import base64
import io
import os
import statistics as st
import sys
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import informe_estilo as E  # noqa: E402
# El motor de agregación y los mapas satelitales son los del informe por
# comunidad: aquí no se recalcula nada que ya exista allá.
from generar_informe_sociologo import (  # noqa: E402
    BASE, DIR_MAPAS, FECHA_CORTE, GRANJA_AVICOLA_MIN,
    agregar_todo, f0, f1, f2, generar_mapas, lleno, num, pct,
)

HTML_OUT = os.path.join(BASE, 'docs', 'INFORME-SOCIOLOGO-por-sector.html')
MD_OUT = os.path.join(BASE, 'docs', 'INFORME-SOCIOLOGO-por-sector.md')
XLSX_OUT = os.path.join(BASE, 'build_entrega', 'Informe_Sociologo_Sectores.xlsx')
# Carpeta (relativa a docs/) con los PNG de los gráficos y los JPG de los
# mapas: el HTML los lleva embebidos en base64, pero el Markdown —y por tanto
# el Word— necesita archivos a los que apuntar.
DIR_GRAF = 'graficos-sociologo'

SECTORES = ('Sector 1', 'Sector 2', 'Sector 3')

# La paleta del Dashboard (PIE_COLORS de DashboardHome.tsx), en el mismo orden.
PIE_COLORS = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444',
              '#8b5cf6', '#ec4899', '#06b6d4', '#84cc16']
COLORES_INSTRUCCION = {'Ninguno': '#94a3b8', 'Alfabetizado': '#22d3ee',
                       'Primaria': '#3b82f6', 'Secundaria': '#6366f1',
                       'Superior': '#8b5cf6'}
NIVELES_INSTRUCCION = ['Ninguno', 'Alfabetizado', 'Primaria',
                       'Secundaria', 'Superior']

NOTA_UNIV_TODAS = ('Universo: todas las fichas del corte (principales + '
                   'adicionales completadas). Cada ficha es un predio.')
NOTA_UNIV_PRI = ('Universo: fichas principales (una por titular entrevistado; '
                 'las adicionales duplicarían su respuesta).')


def es_hija(p):
    return p.get('es_ficha_hija') == 1


def fnum(n):
    """Números al estilo es-EC de la web: punto de miles, coma decimal."""
    return f0(n)


# ─── Los datos de cada gráfico, calcados de DashboardHome.tsx ────────────────

def datos_graficos(todas, pri, cultivos, animales, coms_sup):
    """`todas`/`pri`: fichas del corte (sector o sistema); `cultivos`/
    `animales`: sus registros; `coms_sup`: filas de superficie_por_comunidad
    de sus comunidades. Devuelve un dict con la matriz de cada gráfico."""
    d = {}

    # 1 — Uso del Suelo (medición CATASTRAL de las comunidades del corte,
    #     como la web con filtro: riego ajustado + resto del polígono)
    d['uso_suelo'] = {
        'riego': sum(c.get('riego_ajustado_ha') or 0 for c in coms_sup),
        'sin_riego': sum(c.get('sin_riego_catastral_ha') or 0 for c in coms_sup),
        'catastral': sum(c.get('superficie_catastral_ha') or 0 for c in coms_sup),
        'n_com': len(coms_sup),
    }

    # 2 — Especies Pecuarias Principales (especie cruda, top 5 por cabezas;
    #     granja avícola excluida con nota — la web NO la excluye)
    cab = defaultdict(int)
    granja = 0
    for a in animales:
        n = int(a.get('cantidad') or 0)
        if n <= 0:
            continue
        if n >= GRANJA_AVICOLA_MIN:
            granja += n
            continue
        cab[(a.get('especie') or 'Sin clasificar').strip()] += n
    d['pecuario'] = sorted(cab.items(), key=lambda kv: -kv[1])[:5]
    d['pecuario_granja_excluida'] = granja
    d['pecuario_total'] = sum(cab.values())
    d['pecuario_registros'] = sum(1 for a in animales
                                  if 0 < int(a.get('cantidad') or 0) < GRANJA_AVICOLA_MIN)

    # 3 — Destino de la Producción Agrícola (registros de cultivo que declaran
    #     cada destino; un registro puede declarar varios)
    d['destino'] = [(nom, sum(1 for c in cultivos if c.get(campo)), color)
                    for nom, campo, color in
                    (('Autoconsumo', 'es_autoconsumo', '#10b981'),
                     ('Mercado / Venta', 'es_mercado', '#3b82f6'),
                     ('Agroindustria', 'es_agroindustria', '#8b5cf6'),
                     ('Exportación', 'es_exportacion', '#ec4899'))]
    d['destino'] = [x for x in d['destino'] if x[1] > 0]

    # 4 — Nivel de Instrucción (principales; escala pedagógica y las erratas
    #     fuera de escala al final, como la web)
    ins = Counter()
    for p in pri:
        v = str(p.get('nivel_instruccion') or '').strip()
        if v:
            ins[v] += 1
    d['instruccion'] = ([(nv, ins[nv]) for nv in NIVELES_INSTRUCCION if ins[nv] > 0] +
                        [(nv, c) for nv, c in ins.items()
                         if nv not in NIVELES_INSTRUCCION])
    d['instruccion_con_dato'] = sum(ins.values())
    d['n_pri'] = len(pri)

    # 5 — Hijos por Familia (principales; basta un campo lleno para contar a
    #     la familia — dos hijos hombres y ninguna mujer es 2 y vacío)
    hh = hm = fam = 0
    for p in pri:
        th, tm = p.get('hijos_hombres'), p.get('hijos_mujeres')
        if th is None and tm is None:
            continue
        fam += 1
        hh += int(num(p, 'hijos_hombres'))
        hm += int(num(p, 'hijos_mujeres'))
    d['hijos'] = {'hombres': hh, 'mujeres': hm, 'familias': fam,
                  'promedio': (hh + hm) / fam if fam else 0.0}

    # 6 — Represa y Capacitación (principales; el campo llega como texto, con
    #     y sin tilde según el dispositivo: S.../N... como la web)
    def si_no(campo):
        si = no = 0
        for p in pri:
            v = str(p.get(campo) or '').strip().upper()
            if not v:
                continue
            if v.startswith('S'):
                si += 1
            elif v.startswith('N'):
                no += 1
        return si, no
    d['comunitaria'] = [(nom,) + si_no(campo) for nom, campo in
                        (('Conoce la represa', 'conoce_presa'),
                         ('Recibió capacitación', 'recibio_capacitacion'),
                         ('Quiere capacitarse', 'le_gustaria_cap'))]
    d['comunitaria'] = [x for x in d['comunitaria'] if x[1] + x[2] > 0]

    # 7 — Método de Riego (promedio % sobre TODAS las fichas del corte,
    #     incluidas las que no declaran método, redondeado a enteros — es
    #     exactamente lo que muestra el tablero web)
    nf = len(todas) or 1
    d['metodo'] = [(nom, round(sum(num(p, campo) for p in todas) / nf), color)
                   for nom, campo, color in
                   (('Aspersión', 'metodo_aspersion_pct', '#3b82f6'),
                    ('Gravedad', 'metodo_gravedad_pct', '#10b981'),
                    ('Goteo', 'metodo_goteo_pct', '#f59e0b'))]
    d['metodo'] = [x for x in d['metodo'] if x[1] > 0]

    # 8 — Cultivos Más Frecuentes (top 12 por número de registros, con la
    #     escritura unificada por mayúsculas como la web)
    etiquetas, conteo = {}, Counter()
    for c in cultivos:
        bruto = ' '.join(str(c.get('tipo_cultivo') or 'Sin dato').split())
        clave = bruto.upper()
        if clave not in etiquetas or etiquetas[clave] == etiquetas[clave].upper():
            etiquetas[clave] = bruto
        conteo[clave] += 1
    d['cultivos_frec'] = [(etiquetas[k], v) for k, v in conteo.most_common(12)]
    d['cultivos_registros'] = len(cultivos)

    # 9 — Tenencia del Predio. Regla 6: SOLO fichas principales (el tablero
    #     web cuenta todas las fichas; la nota al pie documenta la diferencia).
    ten = Counter((str(p.get('tenencia_predio') or '').strip() or 'Sin dato')
                  for p in pri)
    d['tenencia'] = ten.most_common()
    # el valor de la web (todas las fichas), solo para la verificación
    d['tenencia_web'] = Counter((str(p.get('tenencia_predio') or '').strip()
                                 or 'Sin dato') for p in todas).most_common()

    # 10 — Fichas por Parroquia (todas las fichas del corte)
    parr = Counter((str(p.get('parroquia') or '').strip() or 'Sin parroquia')
                   for p in todas)
    d['parroquias'] = parr.most_common()

    d['n_todas'] = len(todas)
    return d


# ─── Dibujo: los mismos gráficos, en matplotlib ──────────────────────────────

_ARCHIVOS = {}  # clave → ruta relativa a docs/, para el Markdown


def _guardar(fig, clave):
    """PNG en base64 para el HTML y archivo en disco para el Markdown/Word."""
    import matplotlib.pyplot as plt
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, facecolor='white',
                bbox_inches='tight')
    carpeta = os.path.join(BASE, 'docs', DIR_GRAF)
    os.makedirs(carpeta, exist_ok=True)
    ruta = os.path.join(carpeta, clave + '.png')
    fig.savefig(ruta, format='png', dpi=200, facecolor='white',
                bbox_inches='tight')
    _ARCHIVOS[clave] = f'{DIR_GRAF}/{clave}.png'
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def _ejes(fig, ax):
    ax.spines[:].set_visible(False)
    ax.tick_params(length=0, labelsize=8.5, colors='#556070')
    ax.set_axisbelow(True)


def g_donut(clave, series, fmt, centro=None):
    """Anillo como los Pie de recharts: (nombre, valor, color) con la leyenda
    de cifras debajo, no como etiquetas radiales."""
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(5.4, 3.4))
    vals = [v for _, v, _ in series]
    cols = [c for _, _, c in series]
    ax.pie(vals, colors=cols, startangle=90, counterclock=False,
           wedgeprops=dict(width=0.38, edgecolor='white', linewidth=2))
    if centro:
        ax.text(0, 0, centro, ha='center', va='center', fontsize=10,
                fontweight='bold', color='#24405e')
    ax.set(aspect='equal')
    leyenda = '    '.join(f'{n}: {fmt(v)}' for n, v, _ in series)
    fig.text(0.5, 0.02, leyenda, ha='center', fontsize=8.5, color='#374151')
    return _guardar(fig, clave)


def g_barras_h(clave, items, colores, fmt=fnum, alto=None):
    """Barras horizontales con la cifra al final, como los BarChart layout=
    vertical de recharts. `items`: [(nombre, valor)] de mayor a menor."""
    import matplotlib.pyplot as plt
    n = len(items)
    fig, ax = plt.subplots(figsize=(6.4, alto or max(1.6, 0.5 * n + 0.6)))
    nombres = [x[0] for x in items][::-1]
    vals = [x[1] for x in items][::-1]
    cols = ([colores(i, nom) for i, (nom, _) in enumerate(items)][::-1]
            if callable(colores) else list(colores)[::-1])
    barras = ax.barh(range(n), vals, color=cols, height=0.62)
    ax.set_yticks(range(n), nombres)
    ax.xaxis.grid(True, linestyle=(0, (3, 3)), color='#dbe3ee', linewidth=0.8)
    _ejes(fig, ax)
    vmax = max(vals) if vals else 1
    for b, v in zip(barras, vals):
        dentro = v > vmax * 0.18
        ax.text(v - vmax * 0.015 if dentro else v + vmax * 0.015,
                b.get_y() + b.get_height() / 2, fmt(v),
                ha='right' if dentro else 'left', va='center', fontsize=8,
                fontweight='bold', color='white' if dentro else '#556070')
    ax.margins(x=0.06)
    return _guardar(fig, clave)


def g_barras_v(clave, items, colores, fmt=fnum, rot=0):
    """Barras verticales con la cifra dentro, como los BarChart de recharts."""
    import matplotlib.pyplot as plt
    n = len(items)
    fig, ax = plt.subplots(figsize=(max(4.6, 0.62 * n + 1.6), 3.2))
    nombres = [x[0] for x in items]
    vals = [x[1] for x in items]
    cols = ([colores(i, nom) for i, (nom, _) in enumerate(items)]
            if callable(colores) else list(colores))
    barras = ax.bar(range(n), vals, color=cols, width=0.62)
    ax.set_xticks(range(n), nombres, rotation=rot,
                  ha='right' if rot else 'center', fontsize=8)
    ax.yaxis.grid(True, linestyle=(0, (3, 3)), color='#dbe3ee', linewidth=0.8)
    _ejes(fig, ax)
    vmax = max(vals) if vals else 1
    for b, v in zip(barras, vals):
        dentro = v > vmax * 0.12
        ax.text(b.get_x() + b.get_width() / 2,
                v - vmax * 0.02 if dentro else v + vmax * 0.02, fmt(v),
                ha='center', va='top' if dentro else 'bottom', fontsize=8,
                fontweight='bold', color='white' if dentro else '#556070')
    ax.margins(y=0.08)
    return _guardar(fig, clave)


def g_si_no(clave, filas):
    """Pares Sí/No en horizontal (verde Sí, rojo No), como «Represa y
    Capacitación». `filas`: [(nombre, sí, no)]."""
    import matplotlib.pyplot as plt
    n = len(filas)
    fig, ax = plt.subplots(figsize=(6.4, 0.9 * n + 0.7))
    ys = range(n)
    si = [f[1] for f in filas][::-1]
    no = [f[2] for f in filas][::-1]
    nombres = [f[0] for f in filas][::-1]
    b1 = ax.barh([y + 0.19 for y in ys], si, height=0.34, color='#10b981',
                 label='Sí')
    b2 = ax.barh([y - 0.19 for y in ys], no, height=0.34, color='#ef4444',
                 label='No')
    ax.set_yticks(list(ys), nombres, fontsize=8.5)
    ax.xaxis.grid(True, linestyle=(0, (3, 3)), color='#dbe3ee', linewidth=0.8)
    _ejes(fig, ax)
    vmax = max(si + no) or 1
    for barras, vals in ((b1, si), (b2, no)):
        for b, v in zip(barras, vals):
            ax.text(v + vmax * 0.015, b.get_y() + b.get_height() / 2, fnum(v),
                    va='center', fontsize=8, fontweight='bold', color='#556070')
    ax.margins(x=0.09)
    ax.legend(loc='lower right', fontsize=8, frameon=False)
    return _guardar(fig, clave)


def dibujar_graficos(slug, d):
    """Los 10 gráficos de un corte. Devuelve {n: (b64, clave)} en el orden
    del Dashboard."""
    g = {}
    g[1] = g_donut(f'{slug}-uso-suelo',
                   [('Con riego', d['uso_suelo']['riego'], '#3b82f6'),
                    ('Sin riego', d['uso_suelo']['sin_riego'], '#f59e0b')],
                   lambda v: f'{f2(v)} ha',
                   centro=f"{f2(d['uso_suelo']['catastral'])} ha")
    if d['pecuario']:
        g[2] = g_barras_h(f'{slug}-pecuario', d['pecuario'],
                          lambda i, n: PIE_COLORS[i % len(PIE_COLORS)])
    if d['destino']:
        g[3] = g_barras_v(f'{slug}-destino',
                          [(n, v) for n, v, _ in d['destino']],
                          [c for _, _, c in d['destino']])
    if d['instruccion']:
        g[4] = g_barras_h(f'{slug}-instruccion', d['instruccion'],
                          lambda i, n: COLORES_INSTRUCCION.get(
                              n, PIE_COLORS[i % len(PIE_COLORS)]))
    if d['hijos']['hombres'] + d['hijos']['mujeres'] > 0:
        g[5] = g_barras_v(f'{slug}-hijos',
                          [('Hombres', d['hijos']['hombres']),
                           ('Mujeres', d['hijos']['mujeres'])],
                          ['#3b82f6', '#ec4899'])
    if d['comunitaria']:
        g[6] = g_si_no(f'{slug}-comunitaria', d['comunitaria'])
    if d['metodo']:
        g[7] = g_donut(f'{slug}-metodo',
                       [(n, v, c) for n, v, c in d['metodo']],
                       lambda v: f'{v} %')
    if d['cultivos_frec']:
        g[8] = g_barras_v(f'{slug}-cultivos', d['cultivos_frec'],
                          lambda i, n: PIE_COLORS[i % len(PIE_COLORS)], rot=45)
    if d['tenencia']:
        g[9] = g_donut(f'{slug}-tenencia',
                       [(n, v, PIE_COLORS[i % len(PIE_COLORS)])
                        for i, (n, v) in enumerate(d['tenencia'])],
                       fnum, centro=fnum(sum(v for _, v in d['tenencia'])))
    if d['parroquias']:
        g[10] = g_barras_v(f'{slug}-parroquias', d['parroquias'],
                           lambda i, n: '#8b5cf6')
    return g


# ─── Los títulos y notas de cada gráfico ─────────────────────────────────────

def titulos_y_notas(d, nombre_corte):
    granja = d['pecuario_granja_excluida']
    return {
        1: ('Uso del Suelo: Con Riego vs Sin Riego (ha)',
            [f"Medición catastral de las {d['uso_suelo']['n_com']} comunidades "
             f"de {nombre_corte}: riego ajustado y resto del polígono, cada "
             'predio contado una sola vez (fuente única '
             'superficie_por_comunidad.json). La superficie declarada por los '
             'comuneros es otra medición del mismo territorio y se cita en la '
             'lectura del sector; las dos familias no se suman entre sí.']),
        2: ('Especies Pecuarias Principales (Cabezas)',
            [NOTA_UNIV_TODAS + ' Las cinco especies con más cabezas, tal '
             'como las registró el técnico.',
             (f'Se excluyen {fnum(granja)} aves de la granja avícola de '
              'Asociación Rosalía (registros de 10.000 aves por titular sobre '
              'el mismo predio, pendiente de terceros del proyecto); el '
              'tablero web no aplica esta exclusión.') if granja else None]),
        3: ('Destino de la Producción Agrícola',
            [NOTA_UNIV_TODAS + ' Registros de cultivo que declaran cada '
             'destino; un mismo cultivo puede declarar varios, así que las '
             'barras no suman el total de registros.']),
        4: ('Nivel de Instrucción',
            [NOTA_UNIV_PRI + f" {fnum(d['instruccion_con_dato'])} de "
             f"{fnum(d['n_pri'])} titulares entrevistados con el dato "
             'registrado.']),
        5: ('Hijos por Familia',
            [NOTA_UNIV_PRI + f" {fnum(d['hijos']['hombres'] + d['hijos']['mujeres'])} "
             f"hijos declarados por {fnum(d['hijos']['familias'])} familias · "
             f"promedio {d['hijos']['promedio']:,.1f} por familia".replace(
                 '.', ',')]),
        6: ('Represa y Capacitación',
            [NOTA_UNIV_PRI + ' Verde Sí, rojo No; solo se cuentan las fichas '
             'con respuesta.']),
        7: ('Método de Riego (promedio %)',
            ['Promedio simple del porcentaje declarado en cada ficha, sobre '
             'todas las fichas del corte —incluidas las que no declaran '
             'método—, redondeado a enteros: es la misma cifra del tablero '
             'web. Por el redondeo, la suma puede no dar 100.']),
        8: ('Cultivos Más Frecuentes',
            [NOTA_UNIV_TODAS + ' Número de registros de cultivo, no '
             'superficie: los doce más registrados de '
             f"{fnum(d['cultivos_registros'])} registros del corte."]),
        9: ('Tenencia del Predio',
            [NOTA_UNIV_PRI + ' El tablero web cuenta aquí todas las fichas '
             '(las adicionales heredan la tenencia del titular), por lo que '
             'sus cifras son mayores; en este informe manda la regla del '
             'proyecto: los datos de las personas salen solo de las fichas '
             'principales.']),
        10: ('Fichas por Parroquia',
             [NOTA_UNIV_TODAS]),
    }


# ─── Documento ───────────────────────────────────────────────────────────────

def figura(b64, clave, titulo, notas):
    h = [f'<h3>{titulo}</h3>',
         '<div class="evitar-corte" style="margin:8px 0">',
         f'<img src="data:image/png;base64,{b64}" alt="{titulo}" '
         'style="max-width:100%;border:1px solid #dbe3ee;border-radius:7px;'
         'background:#fff;padding:6px">',
         '</div>']
    m = [f'### {titulo}', '', f'![{titulo}]({_ARCHIVOS[clave]})', '']
    for nde in notas:
        if nde:
            h.append(f'<p class="sub" style="margin-top:2px">{nde}</p>')
            m += [f'*{nde}*', '']
    return '\n'.join(h), '\n'.join(m)


def lectura_corte(nombre, d, coms, caudal_ls, es_sistema=False,
                  caudal_totales=None):
    """Párrafos de lectura del corte, con el universo de cada cifra nombrado."""
    uso = d['uso_suelo']
    decl = sum(c['sup_declarada'] for c in coms)
    decl_riego = sum(c['sup_riego_decl'] for c in coms)
    top_cult = d['cultivos_frec'][0][0] if d['cultivos_frec'] else '—'
    top_esp = d['pecuario'][0][0] if d['pecuario'] else '—'
    ins_top = max(d['instruccion'], key=lambda kv: kv[1])[0] if d['instruccion'] else '—'
    presa = next((x for x in d['comunitaria'] if x[0] == 'Conoce la represa'),
                 None)
    if es_sistema and caudal_totales:
        frase_caudal = (f'El caudal del sistema es de '
                        f'{f2(caudal_totales["caudal_sistema_ls"])} l/s: '
                        f'{f2(caudal_totales["caudal_comunidades_ls"])} l/s '
                        'que reciben las comunidades (moda por comunidad, '
                        'nunca suma de fichas) más '
                        f'{f2(caudal_totales["caudal_individual_ls"])} l/s de '
                        'tomas individuales.')
    else:
        frase_caudal = (f'Sus comunidades reciben {f1(caudal_ls)} l/s (moda '
                        'por comunidad, nunca suma de fichas; los caudales '
                        'heredados se muestran en el informe por comunidad y '
                        'no se suman).')
    p1 = (f'{"El sistema" if es_sistema else nombre} agrupa '
          f'{fnum(len(coms))} comunidades con {fnum(d["n_todas"])} fichas '
          f'catastrales ({fnum(d["n_pri"])} principales y '
          f'{fnum(d["n_todas"] - d["n_pri"])} adicionales). Los comuneros '
          f'declaran {f2(decl)} hectáreas, de las cuales {f2(decl_riego)} '
          f'({pct(decl_riego, decl):,.1f} %) con riego; la medición catastral '
          f'de sus comunidades es de {f2(uso["catastral"])} hectáreas. '
          + frase_caudal).replace('.0 %', ' %')
    p2 = (f'En producción, el cultivo más registrado es {top_cult} y la '
          f'especie pecuaria con más cabezas, {top_esp}. Entre los titulares '
          f'entrevistados predomina la instrucción {ins_top.lower()}' +
          (f'; {fnum(presa[1])} conocen el proyecto de la represa y '
           f'{fnum(presa[2])} no' if presa else '') + '.')
    return [p1, p2]


def construir_documento(comunidades, sup, caudal, datos_por_corte, mapas):
    titulo = 'Informe por Sector de Investigación'
    subtitulo = ('Los gráficos del tablero del padrón, para el sistema y por '
                 'sector de investigación · Cada gráfico nombra su universo')
    corte_linea = f'Datos al {FECHA_CORTE}'

    H = [E.cabecera(titulo, subtitulo)]
    M = [f'# {titulo}', '', f'*{subtitulo}*', '', f'*{corte_linea}.*', '']

    total = sup['total']
    H.append(E.kpis([
        (f0(total['fichas']), 'fichas catastrales'),
        (f0(total['regantes']), 'fichas principales'),
        (f2(total['superficie_declarada_ha']) + ' ha', 'superficie declarada'),
        (f2(caudal['totales']['caudal_sistema_ls']) + ' l/s',
         'caudal del sistema'),
    ]))

    H.append('<h2>Presentación</h2>')
    M += ['## Presentación', '']
    intro = [
        'Este informe reproduce, para el sistema completo y para cada sector '
        'de investigación, los diez gráficos del tablero de la aplicación '
        'web del padrón, tal como se ven en pantalla con el filtro de sector '
        'puesto. Es el complemento gráfico del Informe por Comunidad: aquella '
        'entrega responde cada pregunta de la ficha comunidad por comunidad; '
        'esta muestra el retrato de cada sector de un vistazo.',
        'Dos universos conviven en los gráficos y cada uno nombra el suyo al '
        'pie: los datos de tierra y producción salen de TODAS las fichas '
        '(cada ficha es un predio); los datos de las personas —instrucción, '
        'hijos, represa y capacitación, tenencia— salen SOLO de las fichas '
        'principales, porque las adicionales pertenecen al mismo titular y '
        'duplicarían su respuesta.',
        'Dos mediciones de superficie conviven también, y no se suman entre '
        'sí: la DECLARADA por los comuneros en la entrevista (el universo de '
        'este material) y la CATASTRAL de los polígonos municipales, que es '
        'la que usa el gráfico de uso del suelo — la misma elección del '
        'tablero web, y así lo etiqueta cada gráfico.',
        f'Las cifras corresponden al padrón al {FECHA_CORTE}, la misma fecha '
        'de referencia de los demás documentos entregados al consorcio. El '
        'levantamiento de campo está cerrado.',
    ]
    for p in intro:
        H.append(f'<p>{p}</p>')
        M += [p, '']

    caudal_de = {c['key']: c for c in comunidades}

    def caudal_corte(coms):
        return sum(c['caudal_ls'] or 0 for c in coms
                   if not c['caudal_heredado_de'])

    cortes = [('Todo el sistema', 'sistema', None)] + \
             [(s, s.lower().replace(' ', '-'), s) for s in SECTORES]

    for nombre, slug, sector in cortes:
        d = datos_por_corte[nombre]
        coms = [c for c in comunidades
                if sector is None or c['sector_informe'] == sector]
        H.append(f'<h2>{nombre}' +
                 (f' — {len(coms)} comunidades</h2>' if sector else '</h2>'))
        M += [f'## {nombre}' + (f' — {len(coms)} comunidades' if sector else ''), '']
        if sector is None:
            # el capítulo del sistema cita las fuentes únicas tal cual, no la
            # suma de comunidades (difieren en centavos de redondeo y el
            # caudal del sistema incluye las tomas individuales)
            H.append(E.kpis([
                (f0(total['fichas']), 'fichas catastrales'),
                (f0(total['regantes']), 'fichas principales'),
                (f2(total['superficie_declarada_ha']) + ' ha',
                 'superficie declarada'),
                (f2(caudal['totales']['caudal_sistema_ls']) + ' l/s',
                 'caudal del sistema'),
            ]))
        else:
            H.append(E.kpis([
                (f0(d['n_todas']), 'fichas catastrales'),
                (f0(d['n_pri']), 'fichas principales'),
                (f2(sum(c['sup_declarada'] for c in coms)) + ' ha',
                 'superficie declarada'),
                (f1(caudal_corte(coms)) + ' l/s', 'caudal de sus comunidades'),
            ]))
        mapa_clave = 'general' if sector is None else sector
        if mapa_clave in mapas:
            pie_mapa = ('El área de estudio sobre imagen satelital (Esri '
                        'World Imagery): límites oficiales de comunas del '
                        'GADM Cayambe, recortados al sistema y coloreados '
                        'por sector de investigación; la numeración 1–50 es '
                        'la del listado oficial de organizaciones de riego.'
                        if sector is None else
                        f'Las comunas oficiales del {sector} (límites del '
                        'GADM Cayambe, asignadas por cruce espacial); la '
                        'numeración es la del listado oficial de '
                        'organizaciones del consorcio.')
            H.append('<div class="evitar-corte" style="margin:10px 0">'
                     f'<img src="data:image/jpeg;base64,{mapas[mapa_clave]}" '
                     f'alt="{pie_mapa}" style="width:100%;border:1px solid '
                     '#dbe3ee;border-radius:7px">'
                     f'<p class="sub" style="margin-top:4px">{pie_mapa}</p></div>')
            M += [f'![{pie_mapa}]({DIR_GRAF}/mapa-{slug}.jpg)', '']
        for p in lectura_corte(nombre, d, coms, caudal_corte(coms),
                               es_sistema=(sector is None),
                               caudal_totales=caudal['totales']):
            H.append(f'<p>{p}</p>')
            M += [p, '']
        tyn = titulos_y_notas(d, nombre if sector else 'todo el sistema')
        for i in sorted(d['_graficos']):
            b64 = d['_graficos'][i]
            tit, notas = tyn[i]
            h, m = figura(b64, f'{slug}-' + CLAVES_GRAFICO[i], tit, notas)
            H.append(h)
            M += [m]

    H.append(E.pie(corte_linea.lower()))
    M += ['---', '',
          f'*{E.PIE_INSTITUCION} · {corte_linea}. Documento generado por '
          'scripts/generar_informe_sociologo_sector.py a partir de los '
          'gráficos del tablero web; encargo del cliente, 31-ago-2026.*', '']

    html = E.documento(f'{titulo} — Padrón Guanguilquí–Porotog', '\n'.join(H))
    return html, '\n'.join(M)


CLAVES_GRAFICO = {1: 'uso-suelo', 2: 'pecuario', 3: 'destino',
                  4: 'instruccion', 5: 'hijos', 6: 'comunitaria',
                  7: 'metodo', 8: 'cultivos', 9: 'tenencia', 10: 'parroquias'}


# ─── Excel de matrices crudas ────────────────────────────────────────────────

def escribir_xlsx(datos_por_corte, ruta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    cab_font = Font(bold=True, color='FFFFFF')
    cab_fill = PatternFill('solid', fgColor='1E4D8C')

    def hoja(nombre, columnas, filas):
        ws = wb.create_sheet(nombre[:31])
        ws.append(columnas)
        for cel in ws[1]:
            cel.font = cab_font
            cel.fill = cab_fill
            cel.alignment = Alignment(vertical='center', wrap_text=True)
        for fila in filas:
            ws.append(fila)
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            ancho = max(len(str(c.value or '')) for c in col[:40])
            ws.column_dimensions[col[0].column_letter].width = \
                min(max(ancho + 2, 10), 42)

    cortes = list(datos_por_corte.items())

    hoja('Uso del suelo (catastral)',
         ['Corte', 'Comunidades', 'Con riego ajustado (ha)',
          'Sin riego catastral (ha)', 'Catastral total (ha)'],
         [[n, d['uso_suelo']['n_com'], round(d['uso_suelo']['riego'], 2),
           round(d['uso_suelo']['sin_riego'], 2),
           round(d['uso_suelo']['catastral'], 2)] for n, d in cortes])

    hoja('Pecuario top 5',
         ['Corte', 'Especie', 'Cabezas', 'Aves granja excluidas (corte)'],
         [[n, esp, v, d['pecuario_granja_excluida']]
          for n, d in cortes for esp, v in d['pecuario']])

    hoja('Destino produccion',
         ['Corte', 'Destino', 'Registros de cultivo'],
         [[n, dest, v] for n, d in cortes for dest, v, _ in d['destino']])

    hoja('Instruccion',
         ['Corte', 'Nivel', 'Titulares', 'Con dato', 'Principales'],
         [[n, nv, v, d['instruccion_con_dato'], d['n_pri']]
          for n, d in cortes for nv, v in d['instruccion']])

    hoja('Hijos',
         ['Corte', 'Hijos hombres', 'Hijas mujeres', 'Familias con dato',
          'Promedio por familia'],
         [[n, d['hijos']['hombres'], d['hijos']['mujeres'],
           d['hijos']['familias'], round(d['hijos']['promedio'], 2)]
          for n, d in cortes])

    hoja('Represa y capacitacion',
         ['Corte', 'Pregunta', 'Si', 'No'],
         [[n, preg, si, no] for n, d in cortes
          for preg, si, no in d['comunitaria']])

    hoja('Metodo de riego',
         ['Corte', 'Metodo', 'Promedio % (redondeado)'],
         [[n, met, v] for n, d in cortes for met, v, _ in d['metodo']])

    hoja('Cultivos frecuentes',
         ['Corte', 'Cultivo', 'Registros'],
         [[n, cu, v] for n, d in cortes for cu, v in d['cultivos_frec']])

    hoja('Tenencia (principales)',
         ['Corte', 'Tenencia', 'Fichas principales'],
         [[n, t, v] for n, d in cortes for t, v in d['tenencia']])

    hoja('Parroquias',
         ['Corte', 'Parroquia', 'Fichas'],
         [[n, pa, v] for n, d in cortes for pa, v in d['parroquias']])

    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    wb.save(ruta)


# ─── Autoverificación contra las fuentes únicas ──────────────────────────────

def verificar(datos_por_corte, comunidades, sup, caudal):
    """Los totales del documento contra las fuentes únicas. No corrige: avisa."""
    avisos = []
    tot = sup['total']
    d_sis = datos_por_corte['Todo el sistema']

    if d_sis['n_todas'] != tot['fichas']:
        avisos.append(f"fichas del sistema: {d_sis['n_todas']} ≠ fuente "
                      f"{tot['fichas']}")
    if d_sis['n_pri'] != tot['regantes']:
        avisos.append(f"principales: {d_sis['n_pri']} ≠ fuente {tot['regantes']}")

    suma_f = sum(datos_por_corte[s]['n_todas'] for s in SECTORES)
    if suma_f != d_sis['n_todas']:
        avisos.append(f'fichas por sector no suman el sistema: {suma_f} ≠ '
                      f"{d_sis['n_todas']}")

    decl = sum(c['sup_declarada'] for c in comunidades)
    if abs(decl - tot['superficie_declarada_ha']) > 0.5:
        avisos.append(f"declarada: {decl:,.2f} ≠ fuente "
                      f"{tot['superficie_declarada_ha']:,.2f}")

    cat_sis = d_sis['uso_suelo']['catastral']
    if abs(cat_sis - tot['superficie_catastral_ha']) > 0.5:
        avisos.append(f"catastral: {cat_sis:,.2f} ≠ fuente "
                      f"{tot['superficie_catastral_ha']:,.2f}")

    q = sum(c['caudal_ls'] or 0 for c in comunidades
            if not c['caudal_heredado_de'])
    q_ref = caudal['totales']['caudal_comunidades_ls']
    if abs(q - q_ref) > 0.1:
        avisos.append(f'caudal de comunidades: {q:,.2f} ≠ fuente {q_ref:,.2f}')

    # La discrepancia CONOCIDA del bloque `sectores` del JSON: cuenta a
    # ASOCIACIÓN ROSALÍA en el Sector 2 (lista duplicada en
    # generar_capas_sectores_comunidades.py) mientras el catálogo oficial, la
    # web y este informe la cuentan en el Sector 3. Se avisa siempre para que
    # nadie compare a ciegas contra ese bloque.
    js = sup.get('sectores', {})
    for s in SECTORES:
        f_doc = datos_por_corte[s]['n_todas']
        f_json = int(js.get(s, {}).get('fichas') or 0)
        if f_doc != f_json:
            print(f'ℹ {s}: el informe cuenta {f_doc} fichas y el bloque '
                  f'`sectores` del JSON dice {f_json} — diferencia esperada: '
                  'Asociación Rosalía (47 fichas) va en el Sector 3, como la '
                  'web; el JSON aún la cuenta en el Sector 2 (bug reportado).')

    for a in avisos:
        print(f'⚠ NO CUADRA — {a}')
    if not avisos:
        print(f"✔ Cuadre contra fuentes únicas: {tot['fichas']:,} fichas "
              f"({tot['regantes']:,} principales) · declarada "
              f"{tot['superficie_declarada_ha']:,.2f} ha · catastral "
              f"{tot['superficie_catastral_ha']:,.2f} ha · caudal "
              f"{caudal['totales']['caudal_sistema_ls']:,.2f} l/s")
        for s in SECTORES:
            ds = datos_por_corte[s]
            coms_s = [c for c in comunidades if c['sector_informe'] == s]
            print(f"   {s}: {ds['n_todas']:,} fichas · declarada "
                  f"{sum(c['sup_declarada'] for c in coms_s):,.2f} ha · "
                  f"catastral {ds['uso_suelo']['catastral']:,.2f} ha")
    return avisos


# ─── main ────────────────────────────────────────────────────────────────────

def main():
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    plt.rcParams.update({'font.family': ['Segoe UI', 'DejaVu Sans'],
                         'axes.edgecolor': '#dbe3ee',
                         'figure.facecolor': 'white'})

    comunidades, sup, caudal, corte_txt, fichas = agregar_todo()

    # El sector del INFORME es el del catálogo oficial (constants.ts), el
    # mismo que usa la web — con Asociación Rosalía en el Sector 3.
    for c in comunidades:
        c['sector_informe'] = c['sector']

    # Datos por ficha para los gráficos, con la misma limpieza del motor:
    # hijas pendientes fuera (hoy 0), clave canónica ya puesta en '_key'.
    todas = [p for p in fichas
             if not (es_hija(p) and
                     (p.get('estado_investigacion') or '') != 'completada')]
    pri = [p for p in fichas if not es_hija(p)]

    import json
    GEO = os.path.join(BASE, 'public', 'geo')
    with open(os.path.join(GEO, 'cultivos.json'), encoding='utf-8') as f:
        cultivos = json.load(f)
    with open(os.path.join(GEO, 'animales.json'), encoding='utf-8') as f:
        animales = json.load(f)

    sector_de_key = {c['key']: c['sector_informe'] for c in comunidades}
    id_a_sector = {p.get('id'): sector_de_key.get(p['_key'])
                   for p in todas}
    ids_todas = {p.get('id') for p in todas}
    cultivos = [c for c in cultivos if c.get('ficha_id') in ids_todas]
    animales = [a for a in animales if a.get('ficha_id') in ids_todas]

    from comunidades_canon import canonica
    sup_por_key = {canonica(c['comunidad']): c for c in sup['comunidades']}

    datos_por_corte = {}
    for nombre, sector in [('Todo el sistema', None)] + \
                          [(s, s) for s in SECTORES]:
        if sector is None:
            t, p_ = todas, pri
            cu, an = cultivos, animales
            coms_sup = [sup_por_key[c['key']] for c in comunidades
                        if c['key'] in sup_por_key]
        else:
            keys = {c['key'] for c in comunidades
                    if c['sector_informe'] == sector}
            t = [p for p in todas if p['_key'] in keys]
            p_ = [p for p in pri if p['_key'] in keys]
            ids = {p.get('id') for p in t}
            cu = [c for c in cultivos if c.get('ficha_id') in ids]
            an = [a for a in animales if a.get('ficha_id') in ids]
            coms_sup = [sup_por_key[k] for k in keys if k in sup_por_key]
        d = datos_graficos(t, p_, cu, an, coms_sup)
        if sector is None:
            # Sin filtro, la web usa la fila `total` del JSON, no la suma de
            # comunidades: las filas por comunidad no cierran exactas entre sí
            # (su «sin riego» suma 8,01 ha más que el total) y con la suma el
            # gráfico del sistema no coincidiría con la pantalla.
            d['uso_suelo'] = {
                'riego': sup['total']['riego_ajustado_ha'],
                'sin_riego': sup['total']['sin_riego_catastral_ha'],
                'catastral': sup['total']['superficie_catastral_ha'],
                'n_com': len(coms_sup),
            }
        slug = 'sistema' if sector is None else sector.lower().replace(' ', '-')
        graficos = dibujar_graficos(slug, d)
        # re-mapear al índice fijo de claves para el documento
        d['_graficos'] = graficos
        datos_por_corte[nombre] = d

    verificar(datos_por_corte, comunidades, sup, caudal)

    # Mapas satelitales: los mismos del informe por comunidad, generados en
    # memoria (pdf_ruta=None para no tocar el anexo PDF de aquel documento) y
    # escritos como archivo propio para el Markdown/Word.
    mapas = generar_mapas(comunidades, pdf_ruta=None)
    carpeta = os.path.join(BASE, 'docs', DIR_GRAF)
    os.makedirs(carpeta, exist_ok=True)
    for clave, slug in [('general', 'sistema'), ('Sector 1', 'sector-1'),
                        ('Sector 2', 'sector-2'), ('Sector 3', 'sector-3')]:
        with open(os.path.join(carpeta, f'mapa-{slug}.jpg'), 'wb') as f:
            f.write(base64.b64decode(mapas[clave]))

    html, md = construir_documento(comunidades, sup, caudal,
                                   datos_por_corte, mapas)
    with open(HTML_OUT, 'w', encoding='utf-8') as f:
        f.write(html)
    with open(MD_OUT, 'w', encoding='utf-8') as f:
        f.write(md)
    escribir_xlsx(datos_por_corte, XLSX_OUT)
    print(f'✔ {os.path.relpath(HTML_OUT, BASE)}')
    print(f'✔ {os.path.relpath(MD_OUT, BASE)}')
    print(f'✔ {os.path.relpath(XLSX_OUT, BASE)}')
    print('  Word: python scripts/md_a_docx.py docs/INFORME-SOCIOLOGO-por-sector.md')


if __name__ == '__main__':
    main()
