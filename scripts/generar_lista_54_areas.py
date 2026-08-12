# -*- coding: utf-8 -*-
"""Excel con las 54 fichas clasificadas por su posición respecto al canal,
para validarlas en el mapa antes de aplicar la corrección de áreas."""
import os, sys
from importlib import util as _util
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import sqlite3

BASE = r"C:\Users\HP\OneDrive\Escritorio\CAYAMBE CATASTRO RIEGO\padron-app"
sys.path.insert(0, os.path.join(BASE, 'scripts'))
_spec = _util.spec_from_file_location(
    'cor', os.path.join(BASE, 'scripts', 'corregir_areas_sin_riego.py'))
_c = _util.module_from_spec(_spec)
_spec.loader.exec_module(_c)

SALIDA = os.path.join(BASE, 'docs', 'REVISION-54-areas-sobre-el-canal.xlsx')
AZUL, ROJO, VERDE, AMBAR = 'FF1F4E79', 'FFFCE4E4', 'FFE2EFDA', 'FFFFF2CC'
FINA = Side(style='thin', color='FFBFBFBF')
BORDE = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)


def relleno(color):
    """Relleno sólido que se ve en cualquier visor.

    `PatternFill('solid', fgColor=...)` deja el bgColor en negro-transparente,
    y hay versiones de Excel/WPS que resuelven el sólido contra el fondo: la
    celda sale blanca aunque el color esté escrito en el archivo. Fijar los dos
    colores lo pinta en todos.
    """
    return PatternFill(fill_type='solid', start_color=color, end_color=color)

con = sqlite3.connect(_c.GPKG)
cur = con.cursor()
t = '"{}"'.format(_c.TABLA)
fallo, n_reg = _c.clasificar_por_canal(cur, t, 25.0, 15)
cur.execute(
    "SELECT COALESCE(clave_catastral,''), "
    "TRIM(COALESCE(apellidos,'')||' '||COALESCE(nombres,'')), "
    "COALESCE(comunidad,'(sin comunidad)'), COALESCE(cedula,''), "
    "COALESCE(area_total,0), COALESCE(coord_x_utm,0), COALESCE(coord_y_utm,0), "
    "SUBSTR(CAST(fecha_creacion AS TEXT),1,10) "
    "FROM {} WHERE {} ORDER BY 3, 5 DESC".format(t, _c.PATRON))
filas = cur.fetchall()
con.close()

wb = Workbook()
ws = wb.active
ws.title = 'Las 54'
tit = ['Clave catastral', 'Regante', 'Comunidad', 'Cédula', 'Área (ha)',
       'Cota (m)', 'Cota vecinas', 'Diferencia', 'VEREDICTO', '¿CORRECTO?',
       'Observación', 'X (UTM 17S)', 'Y (UTM 17S)']
anchos = [17, 34, 18, 13, 11, 10, 13, 11, 26, 14, 40, 13, 13]
for i, (t_, a) in enumerate(zip(tit, anchos), start=1):
    c = ws.cell(row=1, column=i, value=t_)
    c.font = Font(bold=True, color='FFFFFFFF', size=10)
    c.fill = relleno(AZUL)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDE
    ws.column_dimensions[get_column_letter(i)].width = a
ws.row_dimensions[1].height = 30

r = 2
n_sobre = n_bajo = 0
for clave, nom, com, ced, area, x, y, fecha in filas:
    v = fallo.get(clave.strip())
    if v:
        direccion, cota, ref, dif = v
        sobre = direccion == 'riego'
        ver = 'SOBRE el canal — NO riega' if sobre else 'BAJO el canal — SÍ riega'
        limite = abs(dif - 25.0) <= 10
    else:
        cota = ref = dif = None
        sobre, limite = False, True
        ver = 'sin cota — revisar'
    n_sobre += 1 if sobre else 0
    n_bajo += 0 if sobre else 1
    vals = [clave, nom, com, ced, round(area/10000.0, 2),
            round(cota) if cota else None, round(ref) if ref else None,
            round(dif) if dif is not None else None, ver, '',
            'Cerca del límite: conviene mirarla' if limite else '',
            round(x) if x else None, round(y) if y else None]
    for i, val in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.border = BORDE
        c.font = Font(size=10)
        if i == 9:
            c.fill = relleno(ROJO if sobre else VERDE)
        if i == 10:
            c.fill = relleno('FFF2F2F2')
        if i == 11 and limite:
            c.fill = relleno(AMBAR)
    r += 1

ws.auto_filter.ref = 'A1:M{}'.format(r - 1)
ws.freeze_panes = 'C2'

ws2 = wb.create_sheet('Cómo se clasificó')
ws2.column_dimensions['B'].width = 108
lineas = [
    ('Las 54 fichas con el área repetida tres veces', 'titulo'),
    ('', ''),
    ('EL PROBLEMA', 'h'),
    ('En estas fichas el área total, el área con riego y el área sin riego traen el '
     'mismo número. Eso hace que el predio se cuente dos veces —entero con riego y '
     'entero sin riego— y son 56,38 ha de más en la superficie del padrón.', ''),
    ('Todas son del mismo levantamiento, del 24 al 30 de julio de 2026.', ''),
    ('', ''),
    ('EL CRITERIO: LA POSICIÓN RESPECTO AL CANAL', 'h'),
    ('El canal riega por gravedad: un predio que queda por encima del canal no recibe '
     'agua, y uno que queda por debajo sí. Para clasificar cada ficha se comparó su '
     'cota con la de los 15 predios regantes más cercanos; si está más de 25 m por '
     'encima, quedó sobre el canal.', ''),
    ('Resultado: Pambamarca 22 sobre y 8 bajo; Chaupiestancia 1 sobre y 17 bajo; '
     'Pucará 2 sobre; Chinchinloma 1 bajo; sin comunidad 2 sobre y 1 bajo.', ''),
    ('', ''),
    ('QUÉ HAY QUE HACER CON ESTA HOJA', 'h'),
    ('1. Revisar en el mapa las filas con la observación «Cerca del límite»: están a '
     'pocos metros del corte de los 25 m y son las únicas dudosas.', ''),
    ('2. Marcar la columna ¿CORRECTO? con SÍ o NO en las que se revisen.', ''),
    ('3. Con la lista validada se aplica la corrección.', ''),
    ('', ''),
    ('QUÉ HACE LA CORRECCIÓN', 'h'),
    ('A las que quedan SOBRE el canal se les pone el área con riego en cero; a las que '
     'quedan BAJO, el área sin riego en cero. El área total del predio no se toca en '
     'ningún caso: el polígono del catastro la confirma.', ''),
]
rr = 2
for texto, tipo in lineas:
    c = ws2.cell(row=rr, column=2, value=texto)
    if tipo == 'titulo':
        c.font = Font(bold=True, size=15, color=AZUL)
    elif tipo == 'h':
        c.font = Font(bold=True, size=11, color=AZUL)
    else:
        c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    rr += 1

wb.save(SALIDA)
print("guardado: {}".format(SALIDA))
print("   {} fichas · {} sobre el canal · {} bajo el canal".format(
    len(filas), n_sobre, n_bajo))
