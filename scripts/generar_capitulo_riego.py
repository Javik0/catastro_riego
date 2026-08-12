# -*- coding: utf-8 -*-
"""
Capítulo del informe técnico: "El predio y el acceso al agua".

Cubre la sección 2 de la ficha de campo (Predio y riego): superficies, caudal,
frecuencia y turnos, grado de tecnificación, tarifas y reservorios.

CRITERIOS DE ANÁLISIS
---------------------
· Universo: las fichas PRINCIPALES. Cada una es un predio con su regante; las
  adicionales son otros predios del mismo titular y se cuentan aparte para no
  duplicar al entrevistado.
· Los porcentajes de método de riego vacíos son CEROS, no datos faltantes: el
  93,8 % de las fichas suma exactamente 100 % entre gravedad, aspersión y goteo.
· En dinero y turnos se usa la MEDIANA, no el promedio: unas pocas fichas con
  valores extremos desplazan el promedio y darían una cifra que no representa a
  nadie.
· ALPAKA declara tarifas de 672 y 308 USD "mensuales" en 491 fichas, cuando la
  mediana del sistema es 3 USD. No es la tarifa de riego sino otro concepto del
  fraccionamiento; se excluye del análisis económico y se reporta como anomalía.
· El caudal NO se suma ficha a ficha (ver docs/METODOLOGIA-CAUDAL.md): se toma
  de caudal_por_comunidad.json, que lo calcula una vez por comunidad.

SALIDAS
  docs/CAPITULO-predio-y-agua.html            capítulo imprimible
  build_entrega/Predio_y_Agua.xlsx            datos por comunidad y sector
"""

import json
import os
import sqlite3
import statistics as st
import sys
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica, nombre_publico, normalizar  # noqa: E402
import informe_estilo as E  # noqa: E402

GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
T = 'Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e'
BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
HTML = os.path.join(BASE, 'docs', 'CAPITULO-predio-y-agua.html')
XLSX = os.path.join(BASE, 'build_entrega', 'Predio_y_Agua.xlsx')
CAUDAL_JSON = os.path.join(BASE, 'public', 'geo', 'caudal_por_comunidad.json')

# Comunidad cuyas tarifas no son comparables (ver encabezado).
TARIFA_ANOMALA = 'ALPAKA'
MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
         'octubre noviembre diciembre').split()


def num(p, k):
    try:
        return float(p.get(k) or 0)
    except (TypeError, ValueError):
        return 0.0


def lleno(v):
    return v not in (None, '') and str(v).strip() != ''


def pct(a, b):
    return 100.0 * a / b if b else 0.0


def cargar():
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(f'SELECT * FROM "{T}" WHERE es_ficha_hija IS NOT 1')
    pri = [dict(r) for r in cur.fetchall()]
    cur.execute(f'SELECT MAX(fecha_creacion), MAX(fecha_completado) FROM "{T}"')
    f1, f2 = cur.fetchone()
    cur.execute(f'SELECT COUNT(*) FROM "{T}" WHERE es_ficha_hija = 1 AND '
                f'coalesce(estado_investigacion, "pendiente_produccion") != "completada"')
    pendientes = cur.fetchone()[0]
    con.close()

    from generar_capas_sectores_comunidades import COM_A_SECTOR
    vistos = defaultdict(Counter)
    for p in pri:
        crudo = p.get('comunidad') or ''
        p['_comk'] = canonica(crudo) or '(sin comunidad)'
        vistos[p['_comk']][nombre_publico(crudo) or '(sin comunidad)'] += 1
        p['_sec'] = (p.get('sector_investigacion') or '').strip()
    display = {}
    for k, c in vistos.items():
        validas = [(n_, v) for n_, v in c.most_common() if normalizar(n_) == k]
        display[k] = validas[0][0] if validas else k
    for p in pri:
        p['_com'] = display[p['_comk']]
        if p['_sec'] in ('', 'None'):
            p['_sec'] = COM_A_SECTOR.get(p['_comk'], '(sin sector)')

    corte = max(str(f1 or '')[:10], str(f2 or '')[:10])
    corte_txt = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                 if corte else 'la fecha de generación')
    return pri, corte_txt, pendientes


def metodo_predominante(p):
    m = {'Gravedad': num(p, 'metodo_gravedad_pct'),
         'Aspersión': num(p, 'metodo_aspersion_pct'),
         'Goteo': num(p, 'metodo_goteo_pct')}
    return max(m, key=m.get) if sum(m.values()) > 0 else None


def main():
    pri, corte_txt, pendientes = cargar()
    N = len(pri)

    # ── superficies ──
    a_total = sum(num(p, 'area_total') for p in pri)
    a_riego = sum(num(p, 'area_riego') for p in pri)
    ha_t, ha_r = a_total / 10000, a_riego / 10000
    areas = sorted(num(p, 'area_total') for p in pri if num(p, 'area_total') > 0)
    med_area = st.median(areas)
    tramos = [('Menos de 1.000 m²', 0, 1000), ('1.000 – 5.000 m²', 1000, 5000),
              ('5.000 m² – 1 ha', 5000, 10000), ('1 – 5 ha', 10000, 50000),
              ('Más de 5 ha', 50000, float('inf'))]
    dist_area = [(et, sum(1 for a in areas if lo <= a < hi)) for et, lo, hi in tramos]

    # ── caudal (fuente única) ──
    with open(CAUDAL_JSON, encoding='utf-8') as f:
        caudal = json.load(f)
    tot_c = caudal['totales']

    # ── frecuencia / turnos ──
    frec = Counter(str(p['frecuencia_riego']).strip() for p in pri
                   if lleno(p.get('frecuencia_riego')))
    dias = [num(p, 'dias_riego') for p in pri if lleno(p.get('dias_riego'))]
    horas = [num(p, 'horas_turno') for p in pri if lleno(p.get('horas_turno'))]
    dias_ok = [d for d in dias if 0 < d <= 7]
    horas_ok = [h for h in horas if 0 < h <= 24]

    # ── tecnificación ──
    pred = Counter(metodo_predominante(p) for p in pri if metodo_predominante(p))
    sup_met = {'Gravedad': 0.0, 'Aspersión': 0.0, 'Goteo': 0.0}
    for p in pri:
        a = num(p, 'area_riego')
        for campo, et in (('metodo_gravedad_pct', 'Gravedad'),
                          ('metodo_aspersion_pct', 'Aspersión'),
                          ('metodo_goteo_pct', 'Goteo')):
            sup_met[et] += a * num(p, campo) / 100
    sup_total_met = sum(sup_met.values())

    tecnificada = sup_met['Aspersión'] + sup_met['Goteo']

    # tecnificación por sector
    tec_sector = {}
    for sec in sorted({p['_sec'] for p in pri if not p['_sec'].startswith('(')}):
        ps = [p for p in pri if p['_sec'] == sec]
        s_tec = sum(num(p, 'area_riego') * (num(p, 'metodo_aspersion_pct')
                                            + num(p, 'metodo_goteo_pct')) / 100 for p in ps)
        s_tot = sum(num(p, 'area_riego') * (num(p, 'metodo_gravedad_pct')
                                            + num(p, 'metodo_aspersion_pct')
                                            + num(p, 'metodo_goteo_pct')) / 100 for p in ps)
        tec_sector[sec] = (s_tec / 10000, s_tot / 10000, pct(s_tec, s_tot))

    # ── tarifas (excluye la comunidad con valores no comparables) ──
    def tarifa(tipo, excluir_anomala=True):
        return [num(p, 'valor_tarifa') for p in pri
                if str(p.get('tipo_tarifa') or '').strip() == tipo
                and lleno(p.get('valor_tarifa'))
                and not (excluir_anomala and p['_comk'] == TARIFA_ANOMALA)]
    t_mes, t_anio = tarifa('fijo mensual'), tarifa('fijo anual')
    anomalas = [p for p in pri if p['_comk'] == TARIFA_ANOMALA and lleno(p.get('valor_tarifa'))]
    val_anom = Counter(num(p, 'valor_tarifa') for p in anomalas)

    reserv = Counter(str(p['tiene_reservorio']).strip() for p in pri
                     if lleno(p.get('tiene_reservorio')))
    n_res = sum(reserv.values())

    # ── documento ──
    B = []
    A = B.append
    A(E.cabecera('El predio y el acceso al agua',
                 'Superficies, caudal, turnos, tecnificación y tarifas · '
                 'Capítulo del informe técnico'))
    A(E.aviso_corte(corte_txt, N, pendientes))
    A(E.kpis([
        (f'{ha_r:,.0f} ha', 'bajo riego'),
        (f'{pct(a_riego, a_total):.1f}%', 'del área empadronada'),
        (f'{tot_c["caudal_sistema_ls"]:,.0f} l/s', 'caudal del sistema'),
        (f'{pct(tecnificada, sup_total_met):.1f}%', 'superficie tecnificada'),
    ]))

    A('<h2>1. Superficie empadronada y superficie bajo riego</h2>')
    A(f'<p>Los <b>{N:,} predios</b> registrados hasta el corte suman '
      f'<b>{ha_t:,.1f} hectáreas</b>, de las cuales <b>{ha_r:,.1f} ha '
      f'({pct(a_riego, a_total):.1f} %) cuentan con riego</b>. La quinta parte '
      # «secano» se retiró de toda la interfaz web por pedido del cliente
      # (12-ago-2026) y los informes usan el mismo término: «sin riego».
      'restante corresponde a áreas sin dotación: pastos sin riego, bosque o '
      'terreno no cultivable dentro del mismo predio.</p>')
    A(f'<p>El predio tiene una superficie <b>mediana de {med_area:,.0f} m²</b>. '
      'La distribución muestra una estructura de <b>minifundio</b>:</p>')
    A('<table class="evitar-corte"><tr><th>Tamaño del predio</th>'
      '<th class="n">Predios</th><th>Peso</th></tr>')
    for et, n in dist_area:
        A(f'<tr><td>{et}</td><td class="n">{n:,}</td><td>{E.barra(pct(n, len(areas)))}</td></tr>')
    A('</table>')
    A('<p>Esta estructura condiciona cualquier intervención: el sistema atiende a '
      'una mayoría de productores con parcelas pequeñas, para quienes el acceso al '
      'agua es determinante de la viabilidad productiva.</p>')

    A('<h2>2. Caudal del sistema</h2>')
    A(f'<p>El sistema entrega <b>{tot_c["caudal_sistema_ls"]:,.2f} l/s</b>, '
      f'resultado de sumar el caudal de las <b>{len(caudal["comunidades"])} '
      f'comunidades</b> ({tot_c["caudal_comunidades_ls"]:,.2f} l/s) y las '
      f'{tot_c["fichas_individuales"]} concesiones individuales '
      f'({tot_c["caudal_individual_ls"]:,.2f} l/s).</p>')
    A('<div class="nota"><b>Nota metodológica.</b> El caudal <b>no se suma ficha '
      'a ficha</b>. Los técnicos anotaron en cada ficha el caudal que recibe '
      '<i>su comunidad</i>, de modo que el mismo valor se repite en todas las '
      'fichas de esa comunidad; sumarlo daría un caudal físicamente imposible. '
      'Se contabiliza una sola vez por comunidad. El procedimiento completo está '
      'en el documento de metodología del caudal.</div>')
    # Solo comunidades con llave propia: las de caudal heredado repiten el valor
    # de otra y aparecerían como si aportaran un caudal que no existe.
    top_c = sorted((k, v) for k, v in caudal['comunidades'].items()
                   if 'caudal_heredado_de' not in v)
    top_c = sorted(top_c, key=lambda x: -x[1]['caudal_ls'])[:8]
    A('<table class="evitar-corte"><tr><th>Comunidad</th><th class="n">Caudal (l/s)</th>'
      '<th class="n">Fichas</th><th>Origen del dato</th></tr>')
    for com, d in top_c:
        A(f'<tr><td>{com}</td><td class="n">{d["caudal_ls"]:,.2f}</td>'
          f'<td class="n">{d["fichas"]:,}</td><td>{d["origen"].capitalize()}</td></tr>')
    A('</table>')
    heredadas = caudal.get('caudal_heredado', {})
    if heredadas:
        A(f'<p style="font-size:9pt;color:#667;margin-top:-8px">No se listan '
          f'{len(heredadas)} usuarios individuales cuyo caudal declarado coincide '
          'con el de su comunidad de origen: comparten la misma llave y su valor no '
          'se contabiliza por separado.</p>')

    A('<h2>3. Frecuencia y turnos de riego</h2>')
    A('<table class="evitar-corte"><tr><th>Frecuencia</th><th class="n">Predios</th>'
      '<th>Peso</th></tr>')
    for k, n in frec.most_common():
        A(f'<tr><td>{k}</td><td class="n">{n:,}</td>'
          f'<td>{E.barra(pct(n, sum(frec.values())))}</td></tr>')
    A('</table>')
    A(f'<p>El turno <b>semanal</b> es el régimen dominante '
      f'({pct(frec.get("Semanal", 0), sum(frec.values())):.1f} %). La mediana es de '
      f'<b>{st.median(dias_ok):.0f} días de riego</b> por turno y '
      f'<b>{st.median(horas_ok):.0f} horas</b> por jornada de riego.</p>')

    A('<h2>4. Tecnificación del riego</h2>')
    A('<p>Se midió la superficie regada por cada método, ponderando el área de '
      'cada predio por el porcentaje declarado. Es una medida más precisa que '
      'contar predios, porque una hectárea por aspersión pesa lo mismo tenga uno '
      'o diez propietarios.</p>')
    A('<table class="evitar-corte"><tr><th>Método</th><th class="n">Superficie (ha)</th>'
      '<th class="n">Predios donde predomina</th><th>Peso en superficie</th></tr>')
    for met in ('Aspersión', 'Gravedad', 'Goteo'):
        A(f'<tr><td>{met}</td><td class="n">{sup_met[met] / 10000:,.1f}</td>'
          f'<td class="n">{pred.get(met, 0):,}</td>'
          f'<td>{E.barra(pct(sup_met[met], sup_total_met))}</td></tr>')
    A('</table>')
    A(f'<div class="hallazgo"><b>Hallazgo.</b> El '
      f'<b>{pct(tecnificada, sup_total_met):.1f} % de la superficie regada ya usa '
      f'métodos tecnificados</b> (aspersión o goteo): {tecnificada / 10000:,.1f} ha '
      f'de {sup_total_met / 10000:,.1f} ha. La aspersión es el método dominante del '
      f'sistema, mientras el goteo apenas alcanza {sup_met["Goteo"] / 10000:,.1f} ha '
      f'({pct(sup_met["Goteo"], sup_total_met):.1f} %) y representa el mayor margen '
      'de mejora en eficiencia.</div>')
    A('<h3>Tecnificación por sector</h3>')
    A('<table class="evitar-corte"><tr><th>Sector</th>'
      '<th class="n">Superficie regada (ha)</th><th class="n">Tecnificada (ha)</th>'
      '<th>% tecnificado</th></tr>')
    for sec, (s_tec, s_tot, p) in sorted(tec_sector.items()):
        A(f'<tr><td>{sec}</td><td class="n">{s_tot:,.1f}</td>'
          f'<td class="n">{s_tec:,.1f}</td><td>{E.barra(p)}</td></tr>')
    A('</table>')

    A('<h2>5. Tarifas y reservorios</h2>')
    A(f'<p>La tarifa <b>fija mensual</b> es la modalidad más extendida '
      f'({len(t_mes):,} predios), con una <b>mediana de {st.median(t_mes):,.2f} USD</b>. '
      f'La modalidad <b>anual</b> ({len(t_anio):,} predios) tiene una mediana de '
      f'<b>{st.median(t_anio):,.2f} USD</b>.</p>')
    A('<div class="nota"><b>Por qué se usa la mediana.</b> Unos pocos registros con '
      'valores muy altos desplazan el promedio hasta cifras que no representan a '
      'ningún regante real. La mediana —el valor central— describe lo que paga '
      'efectivamente la mayoría.</div>')
    if anomalas:
        A(f'<div class="alerta"><b>Dato a verificar en campo.</b> '
          f'{len(anomalas)} fichas de {TARIFA_ANOMALA} registran tarifas de '
          + ' y '.join(f'{v:,.0f}' for v, _ in val_anom.most_common(2))
          + ' USD como <i>fijo mensual</i>, frente a una mediana de '
          f'{st.median(t_mes):,.2f} USD en el resto del sistema. Por su magnitud no '
          'corresponden a una tarifa mensual de riego sino, probablemente, a otro '
          'concepto del proceso de fraccionamiento. <b>Se excluyen de las cifras '
          'de este capítulo</b> hasta que se verifiquen con los usuarios.</div>')
    A('<h3>Reservorios</h3>')
    A('<table class="evitar-corte"><tr><th>Tipo de reservorio</th>'
      '<th class="n">Predios</th><th>Peso</th></tr>')
    for k, n in reserv.most_common():
        et = {'No': 'Sin reservorio'}.get(k, f'Reservorio {k.lower()}')
        A(f'<tr><td>{et}</td><td class="n">{n:,}</td><td>{E.barra(pct(n, n_res))}</td></tr>')
    A('</table>')
    A(f'<p>El <b>{pct(reserv.get("Comunitario", 0), n_res):.1f} %</b> de los predios '
      'se sirve de un <b>reservorio comunitario</b>, lo que confirma el carácter '
      'colectivo de la infraestructura de almacenamiento: la gestión del agua no '
      'es predio a predio sino comunitaria.</p>')

    A('<h2>6. Conclusiones</h2>')
    A('<ul>')
    A(f'<li>El padrón registra <b>{ha_t:,.0f} ha</b>, de las cuales '
      f'<b>{ha_r:,.0f} ha ({pct(a_riego, a_total):.0f} %) tienen riego</b>.</li>')
    A(f'<li>Predomina el <b>minifundio</b>: la mitad de los predios no supera los '
      f'{med_area:,.0f} m².</li>')
    A(f'<li>La <b>tecnificación alcanza el {pct(tecnificada, sup_total_met):.1f} %</b> '
      f'de la superficie regada, concentrada en aspersión; el goteo, con '
      f'{pct(sup_met["Goteo"], sup_total_met):.1f} %, es el margen de mejora.</li>')
    A(f'<li>El régimen es <b>semanal</b> para el '
      f'{pct(frec.get("Semanal", 0), sum(frec.values())):.0f} % y el almacenamiento '
      f'es <b>comunitario</b> para el {pct(reserv.get("Comunitario", 0), n_res):.0f} %.</li>')
    A(f'<li>La tarifa mediana es de <b>{st.median(t_mes):,.2f} USD mensuales</b>, '
      'una contribución baja que sostiene la operación del sistema.</li>')
    A('</ul>')
    A(E.pie(corte_txt))

    os.makedirs(os.path.dirname(HTML), exist_ok=True)
    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(E.documento('El predio y el acceso al agua — Padrón Guanguilquí–Porotog',
                            '\n'.join(B)))
    print(f'  capítulo: {os.path.relpath(HTML, BASE)}  (corte: {corte_txt})')

    # ── Excel ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    azul, blanco = PatternFill('solid', fgColor='1e4d8c'), Font(color='FFFFFF', bold=True)

    def hoja(nombre, cab, filas):
        ws = wb.create_sheet(nombre)
        ws.append(cab)
        for c in ws[1]:
            c.fill, c.font = azul, blanco
        for f_ in filas:
            ws.append(f_)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                12, min(42, max(len(str(c.value or '')) for c in col) + 2))

    hoja('Resumen', ['Indicador', 'Valor'], [
        ['Predios registrados', N], ['Superficie total (ha)', round(ha_t, 2)],
        ['Superficie con riego (ha)', round(ha_r, 2)],
        ['% con riego', round(pct(a_riego, a_total), 1)],
        ['Superficie mediana del predio (m2)', round(med_area)],
        ['Caudal del sistema (l/s)', tot_c['caudal_sistema_ls']],
        ['Superficie tecnificada (ha)', round(tecnificada / 10000, 1)],
        ['% tecnificado', round(pct(tecnificada, sup_total_met), 1)],
        ['Tarifa mediana mensual (USD)', round(st.median(t_mes), 2)],
        ['Tarifa mediana anual (USD)', round(st.median(t_anio), 2)],
    ])

    filas = []
    for com in sorted({p['_com'] for p in pri}):
        ps = [p for p in pri if p['_com'] == com]
        ar = sum(num(p, 'area_riego') for p in ps)
        at = sum(num(p, 'area_total') for p in ps)
        s_tec = sum(num(p, 'area_riego') * (num(p, 'metodo_aspersion_pct')
                                            + num(p, 'metodo_goteo_pct')) / 100 for p in ps)
        tar = [num(p, 'valor_tarifa') for p in ps
               if str(p.get('tipo_tarifa') or '').strip() == 'fijo mensual'
               and lleno(p.get('valor_tarifa')) and p['_comk'] != TARIFA_ANOMALA]
        filas.append([com, ps[0]['_sec'], len(ps), round(at / 10000, 2), round(ar / 10000, 2),
                      round(pct(ar, at), 1), round(s_tec / 10000, 2),
                      round(pct(s_tec, ar), 1) if ar else None,
                      round(st.median(tar), 2) if tar else None])
    hoja('Por comunidad', ['Comunidad', 'Sector', 'Predios', 'Área total (ha)',
                           'Área riego (ha)', '% riego', 'Tecnificada (ha)',
                           '% tecnificado', 'Tarifa mediana mensual'], filas)
    hoja('Métodos', ['Método', 'Superficie (ha)', 'Predios donde predomina', '% superficie'],
         [[m, round(sup_met[m] / 10000, 2), pred.get(m, 0), round(pct(sup_met[m], sup_total_met), 1)]
          for m in ('Aspersión', 'Gravedad', 'Goteo')])
    hoja('Frecuencia', ['Frecuencia', 'Predios'], [[k, v] for k, v in frec.most_common()])
    hoja('Reservorios', ['Tipo', 'Predios'], [[k, v] for k, v in reserv.most_common()])
    hoja('Tarifas anómalas', ['Comunidad', 'Valor declarado (USD)', 'Fichas'],
         [[TARIFA_ANOMALA, v, n] for v, n in val_anom.most_common()])

    del wb['Sheet']
    os.makedirs(os.path.dirname(XLSX), exist_ok=True)
    wb.save(XLSX)
    print(f'  excel   : {os.path.relpath(XLSX, BASE)}')
    print(f'\n  {ha_r:,.0f} ha bajo riego ({pct(a_riego, a_total):.1f}%) | '
          f'tecnificado {pct(tecnificada, sup_total_met):.1f}% | '
          f'tarifa mediana {st.median(t_mes):,.2f} USD/mes')


if __name__ == '__main__':
    main()
