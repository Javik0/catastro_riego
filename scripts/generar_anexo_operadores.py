# -*- coding: utf-8 -*-
"""
Anexo: operador del sistema por comunidad.

Pedido de Armando (2026-08-04): en el informe los operadores aparecían agrupados
por SECTOR, y no se podía saber a qué comunidad corresponde cada uno. Este anexo
lo lista al revés: una fila por comunidad, con su operador.

CÓMO SE DETERMINA EL OPERADOR DE UNA COMUNIDAD
----------------------------------------------
Cada regante declara a quién reconoce como operador de su sector. El nombre se
escribe a mano, así que la misma persona aparece con variantes ('ROBERTO
AIGAJE', 'ROBERTO HAIGAJE', 'LUIS ROBERTO AIGAJE'). Se toma la MODA por
comunidad —el nombre que más regantes repiten— y se indica cuántos lo
mencionan, para que se vea el respaldo de cada dato.

Las comunidades donde el operador más nombrado no llega a la mitad de las
menciones se marcan, porque ahí el reconocimiento está repartido y conviene
confirmarlo con la Junta.

SALIDAS
  docs/ANEXO-operadores-por-comunidad.html
  build_entrega/Operadores_por_Comunidad.xlsx
"""

import os
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica, nombre_publico, normalizar  # noqa: E402
import informe_estilo as E  # noqa: E402

GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
T = 'Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e'
BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
HTML = os.path.join(BASE, 'docs', 'ANEXO-operadores-por-comunidad.html')
XLSX = os.path.join(BASE, 'build_entrega', 'Operadores_por_Comunidad.xlsx')
MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
         'octubre noviembre diciembre').split()

# Numeración oficial del listado "SECTORES Y COMUNIDADES" de Armando.
ORDEN = {}
for n, com in enumerate([
    'LARCACHACA', 'LA LIBERTAD', 'SAN ANTONIO', 'SAN JOSE', 'MILAGRO', 'CHAMBITOLA',
    'LA CANDELARIA', 'CARRERA', 'COCHAPAMBA', 'JESUS GRAN PODER', 'SANTA BARBARA',
    'ASOCIACION POROTOG', 'COMUNA POROTOG', 'ASOCIACION 17 DE JUNIO',
    'SR. COLOMA MONTESERRIN BAJO', 'CORDILLERAS DE LOS ANDES', 'COMUNA IZACATA',
    'IZACATA', 'LOS ANDES IZACATA', 'LOMA GORDA', 'SAN JACINTO', 'MATIAS IMBAGO',
    'CUARTO LOTE', 'ASOC. SAN VICENTE BAJO', 'SANTA ROSA DE PACCHA',
    'ASOC. SAN VICENTE ALTO', 'PUCARA', 'ASOCIACION SAN PEDRO', 'PITANA ALTO',
    'ALPAKA', 'ASOC. PITANA BAJO', 'PROMEJ. PITANA BAJO', 'SANTA ROSA DE PINGULMI',
    'SANTA MARIANITA DE PINGULMI', 'PAMBAMARCA', 'OTONCITO', 'PAMBAMARQUITO',
    'SR. HERNAN TIMPE', 'HDA. SAN FRANSISCO', 'MONTESERRIN ALTO', 'CHAUPIESTANCIA',
    'PUEBLO DE OTON', 'CANGAHUA PUNGO', 'CHINCHINLOMA', 'ASOCIACION ROSALIA',
    'HDA. GUANGUILQUI', 'PUEBLO DE ASCAZUBI', 'EL MANZANO', 'JUNTA SAN LUIS',
], 1):
    ORDEN[com] = n


def limpiar(nombre):
    """Normaliza el nombre del operador para agrupar variantes de escritura."""
    t = unicodedata.normalize('NFD', (nombre or '').upper().strip())
    t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
    t = re.sub(r'[^A-Z ]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


def main():
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(f'SELECT comunidad, sector_investigacion, operador_sector, '
                f'nom_presidente FROM "{T}" WHERE es_ficha_hija IS NOT 1')
    filas = [dict(r) for r in cur.fetchall()]
    cur.execute(f'SELECT MAX(fecha_creacion), MAX(fecha_completado) FROM "{T}"')
    f1, f2 = cur.fetchone()
    con.close()
    corte = max(str(f1 or '')[:10], str(f2 or '')[:10])
    corte_txt = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                 if corte else '')

    from generar_capas_sectores_comunidades import COM_A_SECTOR

    # nombre a mostrar por comunidad (conserva acentos de campo)
    vistos = defaultdict(Counter)
    for f in filas:
        k = canonica(f.get('comunidad') or '')
        if k:
            vistos[k][nombre_publico(f.get('comunidad') or '')] += 1
    display = {}
    for k, c in vistos.items():
        val = [(n_, v) for n_, v in c.most_common() if normalizar(n_) == k]
        display[k] = val[0][0] if val else k

    por_com = defaultdict(Counter)
    total_com = Counter()
    for f in filas:
        k = canonica(f.get('comunidad') or '')
        if not k:
            continue
        total_com[k] += 1
        op = limpiar(f.get('operador_sector'))
        if op:
            por_com[k][op] += 1

    filas_out = []
    for k in sorted(total_com, key=lambda x: (ORDEN.get(x, 99), x)):
        ops = por_com.get(k)
        if not ops:
            filas_out.append((ORDEN.get(k), display.get(k, k),
                              COM_A_SECTOR.get(k, ''), '(sin dato)', 0, 0, 0, []))
            continue
        nom, n = ops.most_common(1)[0]
        menciones = sum(ops.values())
        otros = [f'{o} ({v})' for o, v in ops.most_common()[1:4]]
        filas_out.append((ORDEN.get(k), display.get(k, k), COM_A_SECTOR.get(k, ''),
                          nom.title(), n, menciones, 100.0 * n / menciones, otros))

    B = []
    A = B.append
    A(E.cabecera('Operadores del sistema por comunidad',
                 'Anexo del informe técnico · Padrón de Usuarios'))
    A(f'<div class="corte"><b>Datos con corte al {corte_txt}.</b> '
      'El operador de cada comunidad se establece a partir de lo que declaran sus '
      'propios regantes durante el empadronamiento. El levantamiento sigue en '
      'curso, de modo que este listado puede completarse en próximas '
      'actualizaciones.</div>')

    con_op = sum(1 for f in filas if limpiar(f.get('operador_sector')))
    A(E.kpis([
        (f'{len(filas_out)}', 'comunidades'),
        (f'{len({r[3] for r in filas_out if r[4]})}', 'operadores identificados'),
        (f'{con_op:,}', 'regantes que lo declararon'),
        (f'{100.0 * con_op / len(filas):.0f}%', 'de respuesta'),
    ]))

    A('<h2>Listado por comunidad</h2>')
    A('<p>Una fila por comunidad, en el orden del listado oficial del sistema. La '
      'columna <b>Respaldo</b> indica cuántos regantes de esa comunidad nombran a '
      'ese operador sobre el total que respondió, de modo que se vea la solidez de '
      'cada dato.</p>')
    A('<table><tr><th class="n">N°</th><th>Comunidad</th><th>Sector</th>'
      '<th>Operador</th><th class="n">Menciones</th><th>Respaldo</th></tr>')
    for n_of, com, sec, op, n, tot, p, _ in filas_out:
        clase = ' class="dest"' if p >= 50 else ''
        A(f'<tr><td class="n">{n_of or "—"}</td><td>{com}</td><td>{sec}</td>'
          f'<td{clase}>{op}</td><td class="n">{n or "—"}</td>'
          f'<td>{E.barra(p) if tot else "—"}</td></tr>')
    A('</table>')

    dudosas = [r for r in filas_out if r[5] and r[6] < 50]
    if dudosas:
        A('<h2>Comunidades con reconocimiento repartido</h2>')
        A('<p>En estas comunidades el operador más nombrado no alcanza la mitad de '
          'las menciones: los regantes reconocen a más de una persona. Conviene '
          'confirmarlo con la Junta de Agua antes de darlo por definitivo.</p>')
        A('<table><tr><th>Comunidad</th><th>Más nombrado</th>'
          '<th>Respaldo</th><th>Otros nombres mencionados</th></tr>')
        for n_of, com, sec, op, n, tot, p, otros in dudosas:
            A(f'<tr><td>{com}</td><td>{op}</td><td class="n">{p:.0f} %</td>'
              f'<td>{"; ".join(o.title() for o in otros) or "—"}</td></tr>')
        A('</table>')

    A('<div class="nota"><b>Sobre las variantes de escritura.</b> El nombre del '
      'operador se registra a mano, de modo que una misma persona puede aparecer '
      'escrita de varias formas. Para cada comunidad se toma el nombre que más '
      'regantes repiten; las variantes menores de ese mismo nombre no alteran el '
      'resultado.</div>')
    A(E.pie(corte_txt))

    os.makedirs(os.path.dirname(HTML), exist_ok=True)
    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(E.documento('Operadores por comunidad — Padrón Guanguilquí–Porotog',
                            '\n'.join(B)))
    print(f'  anexo : {os.path.relpath(HTML, BASE)}')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'Operadores por comunidad'
    ws.append(['N°', 'Comunidad', 'Sector', 'Operador', 'Menciones',
               'Total que respondió', '% respaldo', 'Otros nombres mencionados'])
    for c in ws[1]:
        c.fill = PatternFill('solid', fgColor='1e4d8c')
        c.font = Font(color='FFFFFF', bold=True)
    for n_of, com, sec, op, n, tot, p, otros in filas_out:
        ws.append([n_of, com, sec, op, n or None, tot or None,
                   round(p, 1) if tot else None, '; '.join(o.title() for o in otros)])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            10, min(46, max(len(str(c.value or '')) for c in col) + 2))
    os.makedirs(os.path.dirname(XLSX), exist_ok=True)
    wb.save(XLSX)
    # Sin esto, hay builds de Excel que heredan «sin relleno» del estilo
    # base y los colores no se pintan. Ver excel_compat.py.
    from excel_compat import aplicar_formatos
    aplicar_formatos(XLSX)
    print(f'  excel : {os.path.relpath(XLSX, BASE)}')
    print(f'\n  {len(filas_out)} comunidades | '
          f'{len({r[3] for r in filas_out if r[4]})} operadores | '
          f'{len(dudosas)} con reconocimiento repartido')


if __name__ == '__main__':
    main()
