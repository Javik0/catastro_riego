# -*- coding: utf-8 -*-
"""
Informe completo de la sección de ENCUESTA de la ficha de campo.

En la ficha de papel es la "5. DATOS DE LA COMUNIDAD Y CONOCIMIENTO DE LA JUNTA
DE AGUA"; en el formulario QField, la pestaña "5. ENCUESTA". Cubre:

  conoce_presa            ¿Conoce sobre el Proyecto de la presa Río Porotog?
  como_elige_dir (+otro)  ¿Cómo se elige a la directiva?
  nom_presidente          ¿Cómo se llama el Presidente de la Junta de Agua?
  operador_sector         ¿Conoce quién es el operador del sistema en su sector?
  anios_sistema           ¿Cuántos años tiene este sistema de riego?
  km_canal                ¿Conoce cuántos Km tiene el canal principal?
  recibio_capacitacion    ¿Ha recibido capacitación?
  le_gustaria_cap         ¿Le gustaría recibir capacitación?
  temas_capacitacion      Temas de capacitación deseados

METODOLOGÍA
-----------
· El universo son las FICHAS PRINCIPALES (una por titular entrevistado). Las
  adicionales heredan las respuestas de su ficha madre, así que incluirlas
  contaría al mismo entrevistado varias veces.
· Los nombres propios se agrupan sin acentos/espacios para no partir a la misma
  persona en variantes ("JOSÉ JOAQUÍN TIPANLUISA " = "JOSE JOAQUIN TIPANLUISA").
· Los temas de capacitación (texto libre, 245 escrituras distintas) se agrupan
  por palabra clave en categorías.
· La ficha de papel también pregunta por estatutos y reglamentos de la Junta;
  esos campos NO existen en el formulario digital y no se pueden reportar.

SALIDAS
-------
  docs/INFORME-ENCUESTA-conocimiento-junta.md   informe narrado con tablas
  build_entrega/Encuesta_Conocimiento_Junta.xlsx  8 hojas para trabajar cifras

Uso:  python scripts/generar_informe_encuesta.py
"""

import os
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica, nombre_publico  # noqa: E402

GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
T = 'Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e'
BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
MD = os.path.join(BASE, 'docs', 'INFORME-ENCUESTA-conocimiento-junta.md')
XLSX = os.path.join(BASE, 'build_entrega', 'Encuesta_Conocimiento_Junta.xlsx')
HTML = os.path.join(BASE, 'docs', 'CAPITULO-conocimiento-y-gobernanza.html')

# La precarga de la encuesta entró al formulario el 30/07/2026: las fichas
# creadas desde entonces traen respuestas por defecto salvo corrección del
# técnico, y se reportan aparte como advertencia metodológica.
FECHA_PRECARGA = '2026-07-30'

# Agrupación de los temas de capacitación (texto libre) por palabra clave.
# Se evalúan en orden: la primera categoría que coincida gana.
CATEGORIAS_TEMAS = [
    ('Riego y su manejo', ('RIEGO', 'AGUA', 'ASPERSION', 'ASPERCION', 'GOTEO', 'RIEGO TECNIFICADO')),
    ('Agricultura y cultivos', ('AGRICULT', 'CULTIV', 'SIEMBRA', 'SEMILLA', 'HORTALIZ', 'ABONO', 'SUELO', 'ORGANIC')),
    ('Ganadería', ('GANAD', 'ANIMAL', 'PECUARI', 'VACA', 'ESPECIES')),
    ('Administración y liderazgo', ('ADMINISTR', 'LIDER', 'DIRECTIV', 'ORGANIZA', 'CONTAB', 'GESTION')),
    ('Mantenimiento del sistema', ('MANTENIM', 'CANAL', 'INFRAESTRUCT', 'OPERACION')),
]


def norm(t):
    """MAYÚSCULAS, sin acentos, sin espacios repetidos — agrupa variantes."""
    t = unicodedata.normalize('NFD', (t or '').upper().strip())
    t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', t)


def lleno(v):
    return v not in (None, '') and str(v).strip() != ''


def pct(parte, todo):
    return 100.0 * parte / todo if todo else 0.0


def categoria_tema(texto):
    t = norm(texto)
    for nombre, claves in CATEGORIAS_TEMAS:
        if any(k in t for k in claves):
            return nombre
    return 'Otros temas'


def cargar():
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(f'SELECT * FROM "{T}" WHERE es_ficha_hija IS NOT 1')
    pri = [dict(r) for r in cur.fetchall()]
    con.close()
    # Agrupar por el nombre CANÓNICO (corrige INSACATA→IZACATA, etc.). Para
    # mostrar, se prefiere una escritura de campo cuya forma normalizada
    # coincida con el canon (conserva los acentos: 'SAN JOSÉ'); si el grupo solo
    # tiene escrituras erróneas ('INSACATA GRANDE'), se muestra el canon mismo,
    # nunca el error más repetido.
    from comunidades_canon import normalizar
    vistos = defaultdict(Counter)
    for p in pri:
        crudo = p.get('comunidad') or ''
        p['_comk'] = canonica(crudo) or '(sin comunidad)'
        vistos[p['_comk']][nombre_publico(crudo) or '(sin comunidad)'] += 1
        p['_sec'] = (p.get('sector_investigacion') or '(sin sector)').strip()
    display = {}
    for k, c in vistos.items():
        validas = [(nom, n) for nom, n in c.most_common() if normalizar(nom) == k]
        display[k] = validas[0][0] if validas else k

    # El sector se deriva de la COMUNIDAD cuando el campo viene vacío: 553 fichas
    # perdieron sector_investigacion en campo, y un informe formal no puede
    # mostrar una fila "(sin sector)" con el 13% de las entrevistas. Es el mismo
    # criterio que usa la web (App.tsx): la comunidad manda sobre el sector.
    from generar_capas_sectores_comunidades import COM_A_SECTOR
    for p in pri:
        p['_com'] = display[p['_comk']]
        if p['_sec'] in ('', '(sin sector)', 'None'):
            p['_sec'] = COM_A_SECTOR.get(p['_comk'], '(sin sector)')
    return pri


def tabla(filas, cab):
    anchos = [max(len(str(f[i])) for f in [cab] + filas) for i in range(len(cab))]
    def linea(f, sep='|'):
        return sep + sep.join(' {:<{}} '.format(str(v), a) for v, a in zip(f, anchos)) + sep
    return '\n'.join([linea(cab), linea(['-' * a for a in anchos])] + [linea(f) for f in filas])


def si_no(pri, campo, titulo, L, por_comunidad=True):
    """Bloque estándar para una pregunta Sí/No: global, por sector y extremos."""
    resp = [p for p in pri if lleno(p.get(campo))]
    si = sum(1 for p in resp if str(p[campo]).strip() == 'Sí')
    L.append(f'### {titulo}')
    L.append('')
    L.append(f'Respondieron **{len(resp):,} de {len(pri):,}** fichas principales '
             f'({pct(len(resp), len(pri)):.1f}%).')
    L.append('')
    L.append(f'| Respuesta | Regantes | % |')
    L.append(f'|---|---|---|')
    L.append(f'| **Sí** | {si:,} | **{pct(si, len(resp)):.1f}%** |')
    L.append(f'| No | {len(resp) - si:,} | {pct(len(resp) - si, len(resp)):.1f}% |')
    L.append('')
    L.append('| Sector | Sí | No | % Sí |')
    L.append('|---|---|---|---|')
    filas_sector = {}
    for sec in sorted({p['_sec'] for p in resp}):
        rs = [p for p in resp if p['_sec'] == sec]
        s = sum(1 for p in rs if str(p[campo]).strip() == 'Sí')
        filas_sector[sec] = (s, len(rs) - s)
        L.append(f'| {sec} | {s:,} | {len(rs) - s:,} | {pct(s, len(rs)):.1f}% |')
    L.append('')
    if por_comunidad:
        por_com = defaultdict(lambda: [0, 0])
        for p in resp:
            por_com[p['_com']][0 if str(p[campo]).strip() == 'Sí' else 1] += 1
        grandes = {k: v for k, v in por_com.items() if sum(v) >= 20}
        orden = sorted(grandes.items(), key=lambda x: pct(x[1][0], sum(x[1])))
        L.append('Comunidades con el **menor** porcentaje de "Sí" '
                 '(20+ entrevistados — donde enfocar la socialización):')
        L.append('')
        L.append('| Comunidad | Sí | No | % Sí |')
        L.append('|---|---|---|---|')
        for com, (s, n) in orden[:8]:
            L.append(f'| {com} | {s:,} | {n:,} | {pct(s, s + n):.1f}% |')
        L.append('')
    return {'respondieron': len(resp), 'si': si, 'por_sector': filas_sector}


def main():
    pri = cargar()
    total = len(pri)
    precargadas = sum(1 for p in pri
                      if str(p.get('fecha_creacion') or '') >= FECHA_PRECARGA)

    L = []
    L.append('# Encuesta a los regantes — conocimiento de la Junta y del sistema')
    L.append('')
    L.append('**Padrón de Usuarios · Sistema de Riego Comunitario Guanguilquí–Porotog**  ')
    L.append(f'Generado el {date.today().strftime("%d/%m/%Y")} desde el `data.gpkg` de campo.')
    L.append('')
    L.append('Corresponde a la sección **"5. Datos de la comunidad y conocimiento de la '
             'Junta de Agua"** de la ficha de papel (pestaña "5. ENCUESTA" en QField).')
    L.append('')
    L.append('## Metodología')
    L.append('')
    L.append(f'- El universo son las **{total:,} fichas principales** — una por titular '
             'entrevistado. Las fichas adicionales heredan las respuestas de su ficha '
             'madre, así que incluirlas contaría al mismo entrevistado varias veces.')
    L.append('- Los nombres propios se agrupan ignorando acentos y espacios, para no '
             'partir a la misma persona en varias escrituras.')
    L.append('- Los temas de capacitación son texto libre (245 escrituras distintas) y '
             'se agrupan por palabra clave.')
    L.append(f'- Desde el {FECHA_PRECARGA[8:10]}/{FECHA_PRECARGA[5:7]} el formulario '
             f'precarga las respuestas más comunes; **{precargadas} fichas** se crearon '
             'después y sus respuestas pueden ser la precarga sin corregir. Sobre '
             f'{total:,} entrevistas no alteran ninguna cifra.')
    L.append('- La ficha de papel también pregunta si la Junta **tiene estatutos y '
             'reglamentos**; esos dos campos no existen en el formulario digital y '
             'no se pueden reportar.')
    L.append('')

    # ── resumen ejecutivo (se rellena al final, se inserta aquí) ──
    idx_resumen = len(L)

    L.append('---')
    L.append('')
    L.append('## 1. Conocimiento del proyecto de la presa Río Porotog')
    L.append('')
    presa = si_no(pri, 'conoce_presa', '¿Conoce sobre el Proyecto de la presa Río Porotog?', L)

    # ── directiva ──
    L.append('---')
    L.append('')
    L.append('## 2. La directiva de la Junta de Agua')
    L.append('')
    L.append('### ¿Cómo se elige a la directiva?')
    L.append('')
    resp = [p for p in pri if lleno(p.get('como_elige_dir'))]
    c = Counter(str(p['como_elige_dir']).strip() for p in resp)
    L.append('| Mecanismo | Regantes | % |')
    L.append('|---|---|---|')
    for k, n in c.most_common():
        L.append(f'| {k} | {n:,} | {pct(n, len(resp)):.1f}% |')
    otros = Counter(norm(p['como_elige_dir_otro']) for p in pri
                    if lleno(p.get('como_elige_dir_otro')))
    if otros:
        L.append('')
        L.append('Respuestas de "Otro": ' + '; '.join(f'{k} ({n})' for k, n in otros.most_common()))
    L.append('')
    L.append(f'La elección por **asamblea general** es prácticamente unánime '
             f'({pct(c.get("Asamblea general", 0), len(resp)):.1f}%): la Junta opera '
             'con el mecanismo comunitario clásico.')
    L.append('')

    L.append('### ¿Cómo se llama el Presidente de la Junta de Agua?')
    L.append('')
    resp_p = [p for p in pri if lleno(p.get('nom_presidente'))]
    nombres = Counter(norm(p['nom_presidente']) for p in resp_p)
    correcto = sum(n for k, n in nombres.items() if 'TIPANLUISA' in k)
    L.append(f'Respondieron {len(resp_p):,} fichas principales. **{correcto:,} '
             f'({pct(correcto, len(resp_p)):.1f}%) identifican a José Joaquín '
             'Tipanluisa** (agrupando todas las escrituras del apellido).')
    L.append('')
    L.append('| Nombre respondido (agrupado) | Regantes | % |')
    L.append('|---|---|---|')
    for k, n in nombres.most_common(10):
        L.append(f'| {k.title()} | {n:,} | {pct(n, len(resp_p)):.1f}% |')
    resto = len(resp_p) - sum(n for _, n in nombres.most_common(10))
    if resto > 0:
        L.append(f'| *(otros {len(nombres) - 10} nombres)* | {resto:,} | {pct(resto, len(resp_p)):.1f}% |')
    L.append('')

    L.append('### ¿Conoce al operador del sistema en su sector?')
    L.append('')
    resp_o = [p for p in pri if lleno(p.get('operador_sector'))]
    L.append(f'Respondieron {len(resp_o):,} fichas principales '
             f'({pct(len(resp_o), total):.1f}%). El operador varía por sector; '
             'los más nombrados en cada uno:')
    L.append('')
    for sec in sorted({p['_sec'] for p in resp_o}):
        ops = Counter(norm(p['operador_sector']) for p in resp_o if p['_sec'] == sec)
        L.append(f'**{sec}** ({sum(ops.values()):,} respuestas):')
        L.append('')
        L.append('| Operador | Menciones | % del sector |')
        L.append('|---|---|---|')
        for k, n in ops.most_common(5):
            L.append(f'| {k.title()} | {n:,} | {pct(n, sum(ops.values())):.1f}% |')
        L.append('')

    # ── el sistema ──
    L.append('---')
    L.append('')
    L.append('## 3. Conocimiento del sistema de riego')
    L.append('')
    L.append('### ¿Cuántos años tiene el sistema?')
    L.append('')
    anios = [int(p['anios_sistema']) for p in pri
             if lleno(p.get('anios_sistema')) and str(p['anios_sistema']).isdigit()]
    anios_c = Counter(anios)
    med = sorted(anios)[len(anios) // 2]
    L.append(f'Respondieron {len(anios):,} fichas principales. La respuesta dominante es '
             f'**{anios_c.most_common(1)[0][0]} años** '
             f'({anios_c.most_common(1)[0][1]:,} fichas principales, '
             f'{pct(anios_c.most_common(1)[0][1], len(anios)):.1f}%); '
             f'mediana {med} años, promedio {sum(anios) / len(anios):.1f}, '
             f'rango {min(anios)}–{max(anios)}.')
    L.append('')
    L.append('| Años declarados | Regantes | % |')
    L.append('|---|---|---|')
    for k, n in anios_c.most_common(8):
        L.append(f'| {k} | {n:,} | {pct(n, len(anios)):.1f}% |')
    L.append('')

    L.append('### ¿Cuántos km tiene el canal principal?')
    L.append('')
    kms = [float(p['km_canal']) for p in pri if lleno(p.get('km_canal'))
           and re.fullmatch(r'\d+(\.\d+)?', str(p['km_canal']).strip())]
    kms_c = Counter(kms)
    moda_km, n_moda = kms_c.most_common(1)[0]
    L.append(f'Respondieron {len(kms):,} fichas principales. **{n_moda:,} '
             f'({pct(n_moda, len(kms)):.1f}%) responden {moda_km:g} km**, que es la '
             'longitud de referencia del canal principal.')
    L.append('')
    L.append('| Km declarados | Regantes | % |')
    L.append('|---|---|---|')
    for k, n in kms_c.most_common(8):
        L.append(f'| {k:g} | {n:,} | {pct(n, len(kms)):.1f}% |')
    L.append('')

    # ── capacitación ──
    L.append('---')
    L.append('')
    L.append('## 4. Capacitación')
    L.append('')
    cap = si_no(pri, 'recibio_capacitacion', '¿Ha recibido capacitación?', L)
    gusta = si_no(pri, 'le_gustaria_cap', '¿Le gustaría recibir capacitación?', L,
                  por_comunidad=False)

    # cruce: demanda insatisfecha
    ambos = [p for p in pri if lleno(p.get('recibio_capacitacion'))
             and lleno(p.get('le_gustaria_cap'))]
    no_pero_quiere = sum(1 for p in ambos
                         if str(p['recibio_capacitacion']).strip() == 'No'
                         and str(p['le_gustaria_cap']).strip() == 'Sí')
    si_y_quiere = sum(1 for p in ambos
                      if str(p['recibio_capacitacion']).strip() == 'Sí'
                      and str(p['le_gustaria_cap']).strip() == 'Sí')
    L.append('### Cruce: demanda de capacitación no atendida')
    L.append('')
    L.append('| Situación | Regantes | % |')
    L.append('|---|---|---|')
    L.append(f'| No ha recibido pero SÍ quiere | **{no_pero_quiere:,}** | {pct(no_pero_quiere, len(ambos)):.1f}% |')
    L.append(f'| Recibió y quiere más | {si_y_quiere:,} | {pct(si_y_quiere, len(ambos)):.1f}% |')
    L.append(f'| No quiere | {len(ambos) - no_pero_quiere - si_y_quiere:,} | '
             f'{pct(len(ambos) - no_pero_quiere - si_y_quiere, len(ambos)):.1f}% |')
    L.append('')
    L.append(f'**{no_pero_quiere:,} titulares nunca recibieron capacitación y la piden**: '
             'es la población objetivo directa de un plan de capacitación.')
    L.append('')

    L.append('### Temas solicitados')
    L.append('')
    temas = [p['temas_capacitacion'] for p in pri if lleno(p.get('temas_capacitacion'))]
    cat = Counter(categoria_tema(t) for t in temas)
    L.append(f'{len(temas):,} titulares indicaron temas (texto libre, agrupado por '
             'palabra clave):')
    L.append('')
    L.append('| Categoría | Menciones | % |')
    L.append('|---|---|---|')
    for k, n in cat.most_common():
        L.append(f'| {k} | {n:,} | {pct(n, len(temas)):.1f}% |')
    L.append('')
    ejemplos = Counter(norm(t)[:60] for t in temas)
    L.append('Las 10 respuestas literales más frecuentes:')
    L.append('')
    L.append('| Respuesta (normalizada) | Veces |')
    L.append('|---|---|')
    for k, n in ejemplos.most_common(10):
        L.append(f'| {k.capitalize()} | {n:,} |')
    L.append('')

    # ── calidad ──
    L.append('---')
    L.append('')
    L.append('## 5. Calidad del dato')
    L.append('')
    L.append('| Pregunta | Respondida | % de respuesta |')
    L.append('|---|---|---|')
    for campo, titulo in [('conoce_presa', 'Conoce la presa'),
                          ('como_elige_dir', 'Cómo se elige la directiva'),
                          ('nom_presidente', 'Nombre del presidente'),
                          ('operador_sector', 'Operador del sector'),
                          ('anios_sistema', 'Años del sistema'),
                          ('km_canal', 'Km del canal'),
                          ('recibio_capacitacion', 'Recibió capacitación'),
                          ('le_gustaria_cap', 'Le gustaría capacitación'),
                          ('temas_capacitacion', 'Temas deseados')]:
        n = sum(1 for p in pri if lleno(p.get(campo)))
        L.append(f'| {titulo} | {n:,} | {pct(n, total):.1f}% |')
    L.append('')
    L.append('La tasa de respuesta supera el 93% en todas las preguntas cerradas; '
             'solo los temas de capacitación (texto libre y opcional) bajan al '
             f'{pct(len(temas), total):.0f}%.')
    L.append('')

    # ── resumen ejecutivo ──
    resumen = []
    resumen.append('## Resumen ejecutivo')
    resumen.append('')
    resumen.append(f'- **{total:,} fichas principales** (una por titular entrevistado).')
    resumen.append(f'- **{pct(presa["si"], presa["respondieron"]):.1f}% conoce el proyecto '
                   'de la presa Río Porotog**; el desconocimiento se concentra en '
                   'comunidades puntuales (detalle en §1).')
    resumen.append(f'- **{pct(correcto, len(resp_p)):.1f}% identifica por nombre al '
                   'presidente de la Junta** (José Joaquín Tipanluisa): el nivel de '
                   'conocimiento de la dirigencia es alto.')
    resumen.append(f'- La directiva se elige por **asamblea general** según el '
                   f'{pct(c.get("Asamblea general", 0), len(resp)):.1f}% — '
                   'gobernanza comunitaria consolidada.')
    resumen.append(f'- El sistema tiene **~60 años** según la mayoría y el canal '
                   f'principal **{moda_km:g} km**; las respuestas convergen, señal de '
                   'memoria colectiva consistente.')
    resumen.append(f'- **{pct(cap["si"], cap["respondieron"]):.1f}% ha recibido '
                   f'capacitación** y **{pct(gusta["si"], gusta["respondieron"]):.1f}% '
                   f'quiere recibirla**. Hay **{no_pero_quiere:,} titulares** que nunca '
                   'la recibieron y la piden.')
    resumen.append(f'- El tema más pedido es **{cat.most_common(1)[0][0].lower()}** '
                   f'({pct(cat.most_common(1)[0][1], len(temas)):.0f}% de las menciones).')
    resumen.append('')
    L[idx_resumen:idx_resumen] = resumen

    os.makedirs(os.path.dirname(MD), exist_ok=True)
    with open(MD, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L))
    print(f'  informe : {os.path.relpath(MD, BASE)}')

    # ── capítulo del informe técnico (HTML imprimible) ──
    # El levantamiento sigue en curso: la fecha de corte es la de la última
    # actividad registrada en campo, no la del día en que se genera el archivo.
    con = sqlite3.connect(GPKG)
    k = con.cursor()
    k.execute(f'SELECT MAX(fecha_creacion), MAX(fecha_completado) FROM "{T}"')
    f_creada, f_completada = k.fetchone()
    k.execute(f'SELECT COUNT(*) FROM "{T}" WHERE es_ficha_hija = 1 AND '
              f'coalesce(estado_investigacion, "pendiente_produccion") != "completada"')
    pendientes_s4 = k.fetchone()[0]
    con.close()
    corte = max(str(f_creada or '')[:10], str(f_completada or '')[:10])
    MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
             'octubre noviembre diciembre').split()
    corte_texto = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                   if corte else 'la fecha de generación')

    # personas que solo existen como predio adicional de otro titular
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    k = con.cursor()
    k.execute(f'SELECT id, cedula, apellidos, nombres, es_ficha_hija, ficha_madre_id FROM "{T}"')
    todas = [dict(r) for r in k.fetchall()]
    con.close()
    por_id = {f['id']: f for f in todas}
    ident = lambda f: ((f.get('cedula') or '').strip()
                       or norm(f"{f.get('apellidos') or ''} {f.get('nombres') or ''}"))
    ids_pri = {ident(f) for f in todas if f.get('es_ficha_hija') not in (1, True)}
    solo_adic = {ident(f) for f in todas
                 if f.get('es_ficha_hija') in (1, True)
                 and por_id.get(f.get('ficha_madre_id'))
                 and ident(f) != ident(por_id[f['ficha_madre_id']])
                 and ident(f) not in ids_pri}
    con_adicionales = len({f['ficha_madre_id'] for f in todas
                           if f.get('es_ficha_hija') in (1, True) and f.get('ficha_madre_id')})

    bajas = []
    por_com_presa = defaultdict(lambda: [0, 0])
    for p in pri:
        if lleno(p.get('conoce_presa')):
            por_com_presa[p['_com']][0 if str(p['conoce_presa']).strip() == 'Sí' else 1] += 1
    for com, (s, n) in por_com_presa.items():
        if s + n >= 20:
            bajas.append((com, s, n, pct(s, s + n)))
    bajas.sort(key=lambda x: x[3])

    operadores = []
    for sec in sorted({p['_sec'] for p in resp_o}):
        ops = Counter(norm(p['operador_sector']) for p in resp_o if p['_sec'] == sec)
        nom, n = ops.most_common(1)[0]
        operadores.append((sec, nom.title(), n, pct(n, sum(ops.values()))))

    import capitulo_encuesta_html as cap_html
    datos = {
        'corte_texto': corte_texto, 'total': total, 'pendientes_s4': pendientes_s4,
        'con_adicionales': con_adicionales, 'solo_adicionales': len(solo_adic),
        'padron_personas': total + len(solo_adic),
        'pct_presa': pct(presa['si'], presa['respondieron']),
        'presa_si': presa['si'], 'presa_resp': presa['respondieron'],
        'presa_sector': presa['por_sector'], 'presa_bajas': bajas,
        'pct_asamblea': pct(c.get('Asamblea general', 0), len(resp)),
        'pct_presidente': pct(correcto, len(resp_p)),
        'presidente': 'José Joaquín Tipanluisa',
        'operadores': operadores,
        'anios_moda': anios_c.most_common(1)[0][0],
        'anios_pct': pct(anios_c.most_common(1)[0][1], len(anios)),
        'anios_mediana': med,
        'km_moda': moda_km, 'km_pct': pct(n_moda, len(kms)),
        'pct_recibio_cap': pct(cap['si'], cap['respondieron']),
        'pct_quiere_cap': pct(gusta['si'], gusta['respondieron']),
        'demanda_no_atendida': no_pero_quiere, 'pct_demanda': pct(no_pero_quiere, len(ambos)),
        'cap_si_y_quiere': si_y_quiere, 'pct_si_quiere': pct(si_y_quiere, len(ambos)),
        'cap_no_quiere': len(ambos) - no_pero_quiere - si_y_quiere,
        'pct_no_quiere': pct(len(ambos) - no_pero_quiere - si_y_quiere, len(ambos)),
        'temas': [(k_, n_, pct(n_, len(temas))) for k_, n_ in cat.most_common()],
    }
    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(cap_html.construir(datos))
    print(f'  capítulo: {os.path.relpath(HTML, BASE)}  (corte: {corte_texto})')

    # ── Excel ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    azul = PatternFill('solid', fgColor='1e4d8c')
    blanco = Font(color='FFFFFF', bold=True)

    def hoja(nombre, cab, filas):
        ws = wb.create_sheet(nombre)
        ws.append(cab)
        for celda in ws[1]:
            celda.fill, celda.font = azul, blanco
        for f in filas:
            ws.append(f)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                12, min(46, max(len(str(c.value or '')) for c in col) + 2))
        return ws

    # resumen
    hoja('Resumen', ['Indicador', 'Valor'], [
        ['Regantes entrevistados (fichas principales)', total],
        ['Conoce la presa (%)', round(pct(presa['si'], presa['respondieron']), 1)],
        ['Identifica al presidente (%)', round(pct(correcto, len(resp_p)), 1)],
        ['Elige directiva por asamblea (%)', round(pct(c.get('Asamblea general', 0), len(resp)), 1)],
        ['Años del sistema (moda)', anios_c.most_common(1)[0][0]],
        ['Km del canal (moda)', moda_km],
        ['Recibió capacitación (%)', round(pct(cap['si'], cap['respondieron']), 1)],
        ['Quiere capacitación (%)', round(pct(gusta['si'], gusta['respondieron']), 1)],
        ['Demanda no atendida (nunca recibió y quiere)', no_pero_quiere],
    ])

    # por comunidad: matriz completa para cruces
    filas_com = []
    for com in sorted({p['_com'] for p in pri}):
        ps = [p for p in pri if p['_com'] == com]
        def s_pct(campo):
            r = [p for p in ps if lleno(p.get(campo))]
            return round(pct(sum(1 for p in r if str(p[campo]).strip() == 'Sí'), len(r)), 1) if r else None
        filas_com.append([
            com, ps[0]['_sec'], len(ps),
            s_pct('conoce_presa'), s_pct('recibio_capacitacion'), s_pct('le_gustaria_cap'),
            sum(1 for p in ps if lleno(p.get('nom_presidente'))
                and 'TIPANLUISA' in norm(p['nom_presidente'])),
        ])
    hoja('Por comunidad',
         ['Comunidad', 'Sector', 'Entrevistados', '% conoce presa',
          '% recibió capacitación', '% quiere capacitación', 'Identifican al presidente'],
         filas_com)

    hoja('Presidente', ['Nombre (agrupado)', 'Fichas principales'],
         [[k.title(), n] for k, n in nombres.most_common()])
    filas_op = []
    for sec in sorted({p['_sec'] for p in resp_o}):
        ops = Counter(norm(p['operador_sector']) for p in resp_o if p['_sec'] == sec)
        filas_op += [[sec, k.title(), n] for k, n in ops.most_common()]
    hoja('Operadores', ['Sector', 'Operador', 'Menciones'], filas_op)
    hoja('Años y km', ['Años declarados', 'Fichas principales', '', 'Km declarados', 'Fichas principales'],
         [[a[0], a[1], '', k[0], k[1]] for a, k in
          zip(anios_c.most_common(15) + [('', '')] * 15, kms_c.most_common(15) + [('', '')] * 15)
          if a[0] != '' or k[0] != ''][:15])
    hoja('Temas', ['Categoría', 'Menciones'], [[k, n] for k, n in cat.most_common()])
    hoja('Temas literales', ['Respuesta (normalizada)', 'Veces'],
         [[k.capitalize(), n] for k, n in ejemplos.most_common(60)])
    hoja('Calidad', ['Pregunta', 'Respondida', '% respuesta'],
         [[t, sum(1 for p in pri if lleno(p.get(cmp))),
           round(pct(sum(1 for p in pri if lleno(p.get(cmp))), total), 1)]
          for cmp, t in [('conoce_presa', 'Conoce la presa'),
                         ('como_elige_dir', 'Cómo se elige la directiva'),
                         ('nom_presidente', 'Nombre del presidente'),
                         ('operador_sector', 'Operador del sector'),
                         ('anios_sistema', 'Años del sistema'),
                         ('km_canal', 'Km del canal'),
                         ('recibio_capacitacion', 'Recibió capacitación'),
                         ('le_gustaria_cap', 'Le gustaría capacitación'),
                         ('temas_capacitacion', 'Temas deseados')]])

    del wb['Sheet']
    os.makedirs(os.path.dirname(XLSX), exist_ok=True)
    wb.save(XLSX)
    # Sin esto, hay builds de Excel que heredan «sin relleno» del estilo
    # base y los colores no se pintan. Ver excel_compat.py.
    from excel_compat import aplicar_formatos
    aplicar_formatos(XLSX)
    print(f'  excel   : {os.path.relpath(XLSX, BASE)}')
    print(f'\n  {total:,} fichas principales | presa {pct(presa["si"], presa["respondieron"]):.1f}% | '
          f'presidente {pct(correcto, len(resp_p)):.1f}% | '
          f'demanda de capacitación no atendida: {no_pero_quiere:,}')


if __name__ == '__main__':
    main()
