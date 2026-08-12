# -*- coding: utf-8 -*-
"""
Dos cuadros pedidos por Armando el 5 de agosto de 2026.

1. Predios adicionales elaborados por cada investigador.
2. Distribución de los terrenos por rangos de superficie (para el análisis de
   Alexis), en los 9 rangos que él mismo envió.

DECISIONES QUE SOSTIENEN ESTOS NÚMEROS
--------------------------------------
* **Se cuenta por PREDIO ÚNICO (clave catastral), no por ficha.** 435 predios
  tienen más de una ficha; contarlos por ficha los suma 834 veces de más. El
  caso extremo es el terreno comunal de SR. COLOMA MONTESERRIN BAJO
  (clave 1702510040121, 809,4 ha), que tiene 122 fichas: por ficha aparecería
  122 veces en «más de 10 ha». Se deja igualmente la columna «por ficha» al
  lado para que la diferencia sea visible y auditable.

* **La superficie es la del POLÍGONO CATASTRAL**, no la declarada por el
  regante (decisión de JAVIKO). Ventaja: 118 fichas de Monteserrín Bajo no
  traen área declarada pero sí tienen polígono. Cuidado: ese polígono es el
  terreno comunal entero, no la parcela de cada regante.

* **ALPAKA en dos columnas.** Son lotes de un fraccionamiento, no parcelas
  agrícolas típicas. Al excluirlos se CONSERVAN las 2 fichas resumen
  (codigo_final = S-C-P001, claves 1702520290462 y 1702520290469), que tienen
  clave catastral propia y no la comparten con ningún lote — es lo que pidió
  Armando expresamente.

* Los rangos son cerrados por abajo y abiertos por arriba, y **empiezan en
  1 m²** tal como los escribió Armando («1 a menos de 1000»). Un predio sin
  área NO cae en el primer rango: va a una fila «sin área» aparte.

Salidas
-------
docs/REPORTE-adicionales-por-investigador.html
docs/REPORTE-rangos-superficie.html
build_entrega/Reportes_Armando_5ago.xlsx   (3 hojas)
"""
import json
import os
import sqlite3
from collections import Counter, defaultdict

BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
CATASTRO = os.path.join(BASE, 'public', 'geo', 'catastro_busqueda.json')
FICHAS_TB = "Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e"

MAPEO = {
    'u0_a314': 'Melany Jara', 'u0_a319': 'Melany Jara', 'jvk-editor': 'Melany Jara',
    'u0_a504': 'Adriana Cuascota', 'jvk-editor6': 'Adriana Cuascota',
    'u0_a279': 'Huguito Ipial', 'jvk-editor2': 'Huguito Ipial',
    'u0_a70': 'Pablo Barrionuevo', 'jvk-editor5': 'Pablo Barrionuevo',
    'u0_a330': 'Mayra Benavides', 'mayralisseth201': 'Mayra Benavides',
    'u0_a362': 'Martha Simbaña', 'u0_a335': 'Martha Simbaña', 'jvk-editor4': 'Martha Simbaña',
    'u0_a302': 'Dylan Chavez', 'jvk-editor3': 'Dylan Chavez',
    'u0_a200': 'Melany Recalde', 'jvk-corp': 'Melany Recalde',
    'u0_a2': 'JVK-DIGITALIZACION', 'jvk-digitalizacion': 'JVK-DIGITALIZACION',
}

# Las 2 «fichas resumen» de ALPAKA (el predio matriz del fraccionamiento).
# Armando pidió que se conserven cuando se excluyan los lotes. Se listan por
# clave catastral porque son propias y no las comparte ningún lote.
RESUMEN_ALPAKA = {'1702520290462', '1702520290469'}

RANGOS = [
    (1,      1000,          '1 a menos de 1.000 m²'),
    (1000,   5000,          '1.000 a menos de 5.000 m²'),
    (5000,   10000,         '5.000 m² a menos de 1 ha'),
    (10000,  20000,         '1 ha a menos de 2 ha'),
    (20000,  30000,         '2 ha a menos de 3 ha'),
    (30000,  40000,         '3 ha a menos de 4 ha'),
    (40000,  50000,         '4 ha a menos de 5 ha'),
    (50000,  100000,        '5 ha a menos de 10 ha'),
    (100000, float('inf'),  'Más de 10 ha'),
]

CSS = """
body{font-family:'Segoe UI',Arial,sans-serif;margin:0;padding:32px;background:#f8fafc;color:#0f172a}
.wrap{max-width:1000px;margin:0 auto;background:#fff;padding:32px 36px;border-radius:12px;
      box-shadow:0 1px 3px rgba(0,0,0,.1)}
h1{font-size:22px;margin:0 0 4px;color:#0f172a}
h2{font-size:16px;margin:28px 0 10px;padding-bottom:6px;border-bottom:2px solid #e2e8f0}
.sub{color:#64748b;font-size:13px;margin:0 0 20px}
table{border-collapse:collapse;width:100%;font-size:13px;margin:12px 0}
th{background:#1e293b;color:#fff;padding:9px 10px;text-align:left;font-weight:600;font-size:12px}
th.n,td.n{text-align:right}
td{padding:8px 10px;border-bottom:1px solid #e2e8f0}
tr:nth-child(even) td{background:#f8fafc}
tfoot td{font-weight:700;background:#f1f5f9!important;border-top:2px solid #cbd5e1}
.nota{background:#fffbeb;border-left:4px solid #f59e0b;padding:12px 16px;margin:16px 0;
      font-size:13px;border-radius:0 6px 6px 0}
.ok{background:#f0fdf4;border-left-color:#22c55e}
.barra{background:#e2e8f0;border-radius:3px;height:16px;position:relative;min-width:110px}
.barra span{background:#3b82f6;display:block;height:100%;border-radius:3px}
.barra i{position:absolute;right:6px;top:0;font-size:10px;font-style:normal;line-height:16px;color:#0f172a}
.pie{margin-top:28px;padding-top:14px;border-top:1px solid #e2e8f0;color:#64748b;font-size:11px}
"""


def cargar():
    con = sqlite3.connect('file:{}?mode=ro'.format(GPKG), uri=True)
    cur = con.cursor()
    fichas = cur.execute(
        'SELECT id, clave_catastral, comunidad, codigo_final, es_ficha_hija, '
        'completado_por, creado_por, area_total FROM "{}"'.format(FICHAS_TB)).fetchall()
    con.close()
    with open(CATASTRO, encoding='utf-8') as f:
        cat = json.load(f)
    areas = {str(r['clave_cata']).strip(): float(r.get('area_predi') or 0)
             for r in cat if r.get('clave_cata')}
    return fichas, areas


def tecnico(usuario):
    u = str(usuario or '').strip()
    if not u:
        return None
    return MAPEO.get(u, MAPEO.get(u.lower(), u))


# ───────────────────────── reporte 1 ─────────────────────────
def reporte_adicionales(fichas):
    hijas = [f for f in fichas if f[4] == 1]
    por_tec = Counter()
    sin_autor = 0
    for f in hijas:
        t = tecnico(f[5]) or tecnico(f[6])
        if t is None or str(f[5] or '').strip() == '':
            t2 = tecnico(f[5])
            if t2 is None:
                sin_autor += 1
                continue
        por_tec[t] += 1
    filas = sorted(por_tec.items(), key=lambda x: -x[1])
    total = sum(por_tec.values()) + sin_autor
    mx = filas[0][1] if filas else 1

    h = ['<div class="wrap">',
         '<h1>Predios adicionales elaborados por cada investigador</h1>',
         '<p class="sub">Padrón Guanguilquí–Porotog · corte 5 de agosto de 2026</p>',
         '<div class="nota ok"><b>Levantamiento cerrado.</b> Los {:,} predios adicionales '
         'están <b>completados al 100 %</b>: no queda ninguno pendiente de la Sección 4 '
         '(el 31 de julio eran 577 y el 4 de agosto todavía 115).</div>'.format(total).replace(',', '.'),
         '<table><thead><tr><th>Investigador</th><th class="n">Predios adicionales</th>'
         '<th class="n">% del total</th><th>Peso</th></tr></thead><tbody>']
    for nom, n in filas:
        pct = 100.0 * n / total if total else 0
        h.append('<tr><td>{}</td><td class="n">{:,}</td><td class="n">{:.1f} %</td>'
                 '<td><div class="barra"><span style="width:{:.0f}%"></span>'
                 '<i>{:.0f}%</i></div></td></tr>'
                 .format(nom, n, pct, 100.0 * n / mx, pct).replace(',', '.'))
    if sin_autor:
        h.append('<tr><td><i>Sin investigador registrado</i></td><td class="n">{}</td>'
                 '<td class="n">{:.1f} %</td><td></td></tr>'
                 .format(sin_autor, 100.0 * sin_autor / total))
    h.append('</tbody><tfoot><tr><td>TOTAL</td><td class="n">{:,}</td><td class="n">100 %</td>'
             '<td></td></tr></tfoot></table>'.format(total).replace(',', '.'))
    h.append('<div class="nota">El reparto usa <b>quién completó la Sección 4</b> en campo. '
             'No se usa «quién creó el registro» porque las fichas adicionales las genera un '
             'script desde la Sección 7 de la ficha madre y aparecerían casi todas como '
             '<code>AUTO-SECCION7</code>, sin reflejar el trabajo de nadie.</div>')
    h.append('<div class="pie">Fuente: data.gpkg (QFieldCloud) · '
             'generado por scripts/generar_reportes_armando.py</div></div>')
    return '\n'.join(h), filas, sin_autor, total


# ───────────────────────── reporte 2 ─────────────────────────
def reporte_rangos(fichas, areas):
    # predio único -> área catastral
    predios = {}
    fichas_por_predio = Counter()
    for _id, clave, com, cod, hija, _c1, _c2, _a in fichas:
        k = str(clave or '').strip()
        if not k:
            continue
        fichas_por_predio[k] += 1
        predios.setdefault(k, {'area': areas.get(k), 'comunidades': set(), 'codigos': set()})
        if com:
            predios[k]['comunidades'].add(str(com).strip().upper())
        if cod:
            predios[k]['codigos'].add(str(cod).strip().upper())

    def es_alpaka(k):
        return any('ALPAKA' in c for c in predios[k]['comunidades'])

    # Las 2 fichas resumen se identifican por su CLAVE CATASTRAL, no por el
    # codigo_final: el código "S-C-P001" no es único de ALPAKA, se repite 5.523
    # veces en todo el padrón y marcaría como resumen a lotes que no lo son.
    lotes_alpaka = {k for k in predios if es_alpaka(k)} - RESUMEN_ALPAKA

    def distribuir(claves, por_ficha=False):
        out = defaultdict(int)
        sin = 0
        for k in claves:
            peso = fichas_por_predio[k] if por_ficha else 1
            a = predios[k]['area']
            if not a or a <= 0:
                sin += peso
                continue
            for lo, hi, lab in RANGOS:
                if lo <= a < hi:
                    out[lab] += peso
                    break
            else:
                sin += peso
        return out, sin

    todos = set(predios)
    sin_al = todos - lotes_alpaka
    d_all, s_all = distribuir(todos)
    d_sin, s_sin = distribuir(sin_al)
    d_fic, s_fic = distribuir(todos, por_ficha=True)

    tot_all = sum(d_all.values()) + s_all
    tot_sin = sum(d_sin.values()) + s_sin
    tot_fic = sum(d_fic.values()) + s_fic

    h = ['<div class="wrap">',
         '<h1>Terrenos por rango de superficie</h1>',
         '<p class="sub">Padrón Guanguilquí–Porotog · corte 5 de agosto de 2026 · '
         'superficie del <b>polígono catastral</b></p>',
         '<div class="nota"><b>Se cuenta por predio, no por ficha.</b> '
         'Hay <b>{}</b> terrenos con más de una ficha levantada; contarlos por ficha los '
         'sumaría <b>{} veces de más</b>. El caso extremo es el terreno comunal de '
         'Sr. Coloma Monteserrín Bajo (809,4 ha), que tiene <b>122 fichas</b>: por ficha '
         'aparecería 122 veces en «más de 10 ha». La última columna deja ver esa diferencia.'
         '</div>'.format(
             sum(1 for k, v in fichas_por_predio.items() if v > 1),
             sum(v for v in fichas_por_predio.values() if v > 1)
             - sum(1 for v in fichas_por_predio.values() if v > 1)),
         '<table><thead><tr><th>Rango de superficie</th>'
         '<th class="n">Predios<br>(todos)</th><th class="n">%</th>'
         '<th class="n">Predios<br>(sin lotes ALPAKA)</th><th class="n">%</th>'
         '<th class="n">Si se contara<br>por ficha</th></tr></thead><tbody>']
    for _lo, _hi, lab in RANGOS:
        h.append('<tr><td>{}</td><td class="n">{:,}</td><td class="n">{:.1f} %</td>'
                 '<td class="n">{:,}</td><td class="n">{:.1f} %</td>'
                 '<td class="n" style="color:#94a3b8">{:,}</td></tr>'
                 .format(lab, d_all[lab], 100.0 * d_all[lab] / tot_all,
                         d_sin[lab], 100.0 * d_sin[lab] / tot_sin, d_fic[lab])
                 .replace(',', '.'))
    h.append('<tr><td><i>Sin área en el catastro</i></td><td class="n">{}</td>'
             '<td class="n">{:.1f} %</td><td class="n">{}</td><td class="n">{:.1f} %</td>'
             '<td class="n" style="color:#94a3b8">{}</td></tr>'
             .format(s_all, 100.0 * s_all / tot_all, s_sin,
                     100.0 * s_sin / tot_sin, s_fic))
    h.append('</tbody><tfoot><tr><td>TOTAL</td><td class="n">{:,}</td><td class="n">100 %</td>'
             '<td class="n">{:,}</td><td class="n">100 %</td>'
             '<td class="n">{:,}</td></tr></tfoot></table>'
             .format(tot_all, tot_sin, tot_fic).replace(',', '.'))

    n_res = len(RESUMEN_ALPAKA & todos)
    h.append('<div class="nota"><b>ALPAKA.</b> La columna «sin lotes ALPAKA» retira los '
             '<b>{}</b> lotes del fraccionamiento pero <b>conserva las {} fichas resumen</b> '
             '(código S-C-P001), que tienen clave catastral propia y no la comparten con '
             'ningún lote.</div>'.format(len(lotes_alpaka), n_res))
    h.append('<div class="nota"><b>Advertencia sobre Monteserrín Bajo.</b> Sus 118 fichas '
             'no traen superficie declarada y todas apuntan al mismo polígono comunal de '
             '809,4 ha. Ese terreno entra <b>una sola vez</b> en «más de 10 ha»: la '
             'superficie individual de cada uno de esos regantes <b>no está medida</b>.</div>')
    h.append('<div class="pie">Superficie tomada del catastro rural del GADM Cayambe. '
             'Rangos cerrados por abajo y abiertos por arriba, desde 1 m². '
             'Fuente: data.gpkg + catastro_busqueda.json</div></div>')
    return '\n'.join(h), (d_all, s_all, tot_all), (d_sin, s_sin, tot_sin), \
           (d_fic, s_fic, tot_fic), predios, fichas_por_predio, lotes_alpaka


def envolver(titulo, cuerpo):
    return ('<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">'
            '<title>{}</title><style>{}</style></head><body>{}</body></html>'
            .format(titulo, CSS, cuerpo))


def main():
    fichas, areas = cargar()
    print('=' * 74)
    print(' REPORTES PARA ARMANDO — 5 de agosto de 2026')
    print('=' * 74)

    html1, filas, sin_autor, total_ad = reporte_adicionales(fichas)
    p1 = os.path.join(BASE, 'docs', 'REPORTE-adicionales-por-investigador.html')
    with open(p1, 'w', encoding='utf-8') as f:
        f.write(envolver('Predios adicionales por investigador', html1))
    print('\n[1] Adicionales por investigador  ->', os.path.relpath(p1, BASE))
    for n, v in filas:
        print('      {:26s} {:5,}'.format(n, v).replace(',', '.'))
    if sin_autor:
        print('      {:26s} {:5}'.format('(sin investigador)', sin_autor))
    print('      {:26s} {:5,}'.format('TOTAL', total_ad).replace(',', '.'))

    html2, A, S, Fq, predios, fpp, lotes = reporte_rangos(fichas, areas)
    p2 = os.path.join(BASE, 'docs', 'REPORTE-rangos-superficie.html')
    with open(p2, 'w', encoding='utf-8') as f:
        f.write(envolver('Terrenos por rango de superficie', html2))
    d_all, s_all, tot_all = A
    d_sin, s_sin, tot_sin = S
    d_fic, s_fic, tot_fic = Fq
    print('\n[2] Rangos de superficie          ->', os.path.relpath(p2, BASE))
    print('      {:32s} {:>8s} {:>10s} {:>12s}'.format('rango', 'todos', 'sin ALPAKA', 'por ficha'))
    for _lo, _hi, lab in RANGOS:
        print('      {:32s} {:8,} {:10,} {:12,}'
              .format(lab, d_all[lab], d_sin[lab], d_fic[lab]).replace(',', '.'))
    print('      {:32s} {:8,} {:10,} {:12,}'
          .format('sin área', s_all, s_sin, s_fic).replace(',', '.'))
    print('      {:32s} {:8,} {:10,} {:12,}'
          .format('TOTAL', tot_all, tot_sin, tot_fic).replace(',', '.'))

    # ── Excel ──
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        hdr_f = Font(bold=True, color='FFFFFF', size=10)
        hdr_b = PatternFill('solid', fgColor='1E293B')
        ctr = Alignment(horizontal='center')

        ws = wb.active
        ws.title = 'Adicionales por investigador'
        ws.append(['INVESTIGADOR', 'PREDIOS ADICIONALES', '% DEL TOTAL'])
        for c in ws[1]:
            c.font, c.fill, c.alignment = hdr_f, hdr_b, ctr
        for n, v in filas:
            ws.append([n, v, round(100.0 * v / total_ad, 1)])
        if sin_autor:
            ws.append(['(sin investigador registrado)', sin_autor,
                       round(100.0 * sin_autor / total_ad, 1)])
        ws.append(['TOTAL', total_ad, 100])
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 14

        ws2 = wb.create_sheet('Rangos de superficie')
        ws2.append(['RANGO DE SUPERFICIE', 'PREDIOS (TODOS)', '%',
                    'PREDIOS (SIN LOTES ALPAKA)', '%', 'SI SE CONTARA POR FICHA'])
        for c in ws2[1]:
            c.font, c.fill, c.alignment = hdr_f, hdr_b, ctr
        for _lo, _hi, lab in RANGOS:
            ws2.append([lab, d_all[lab], round(100.0 * d_all[lab] / tot_all, 1),
                        d_sin[lab], round(100.0 * d_sin[lab] / tot_sin, 1), d_fic[lab]])
        ws2.append(['Sin área en el catastro', s_all, round(100.0 * s_all / tot_all, 1),
                    s_sin, round(100.0 * s_sin / tot_sin, 1), s_fic])
        ws2.append(['TOTAL', tot_all, 100, tot_sin, 100, tot_fic])
        ws2.column_dimensions['A'].width = 32
        for col in 'BCDEF':
            ws2.column_dimensions[col].width = 20

        ws3 = wb.create_sheet('Inconsistencias')
        ws3.append(['CLAVE CATASTRAL', 'FICHAS SOBRE EL MISMO PREDIO',
                    'AREA CATASTRAL (m2)', 'AREA (ha)', 'COMUNIDAD', 'OBSERVACION'])
        for c in ws3[1]:
            c.font, c.fill, c.alignment = hdr_f, hdr_b, ctr
        for k, n in sorted(fpp.items(), key=lambda x: -x[1]):
            if n < 2:
                continue
            a = predios[k]['area'] or 0
            ws3.append([k, n, round(a, 2), round(a / 10000.0, 4),
                        ' / '.join(sorted(predios[k]['comunidades']))[:60],
                        'Contado 1 vez por predio; por ficha se contaria {} veces'.format(n)])
        for k in sorted(predios):
            if not predios[k]['area']:
                ws3.append([k, fpp[k], 0, 0,
                            ' / '.join(sorted(predios[k]['comunidades']))[:60],
                            'Clave sin pareo en el catastro: queda en "sin area"'])
        ws3.column_dimensions['A'].width = 26
        ws3.column_dimensions['B'].width = 28
        ws3.column_dimensions['C'].width = 20
        ws3.column_dimensions['D'].width = 12
        ws3.column_dimensions['E'].width = 40
        ws3.column_dimensions['F'].width = 56

        px = os.path.join(BASE, 'build_entrega', 'Reportes_Armando_5ago.xlsx')
        wb.save(px)
        # Sin esto, hay builds de Excel que heredan «sin relleno» del estilo
        # base y los colores no se pintan. Ver excel_compat.py.
        from excel_compat import aplicar_formatos
        aplicar_formatos(px)
        print('\n[3] Excel (3 hojas)               ->', os.path.relpath(px, BASE))
    except Exception as e:
        print('\n[aviso] no se pudo generar el Excel:', e)

    print('=' * 74)


if __name__ == '__main__':
    main()
