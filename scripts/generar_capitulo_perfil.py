# -*- coding: utf-8 -*-
"""
Capítulo del informe técnico: "Perfil del regante".

Cubre la sección 1 de la ficha (Datos del propietario): escolaridad, tenencia
de la tierra, composición familiar y distribución territorial.

SOBRE LA ESTIMACIÓN DE GÉNERO
-----------------------------
La ficha no pregunta el sexo del titular, pero la participación de mujeres en la
titularidad del agua es un indicador estándar en proyectos de riego. Se estima a
partir del primer nombre con una lista de nombres inequívocos; lo que no encaja
queda como "no determinado" en lugar de forzar una clasificación. El capítulo
declara que es una estimación y no un dato registrado.

SALIDAS
  docs/CAPITULO-perfil-del-regante.html
  build_entrega/Perfil_del_Regante.xlsx
"""

import os
import sqlite3
import statistics as st
import sys
import unicodedata
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica, nombre_publico, normalizar  # noqa: E402
import informe_estilo as E  # noqa: E402

GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
T = 'Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e'
BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
HTML = os.path.join(BASE, 'docs', 'CAPITULO-perfil-del-regante.html')
XLSX = os.path.join(BASE, 'build_entrega', 'Perfil_del_Regante.xlsx')
MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
         'octubre noviembre diciembre').split()

# Nombres inequívocos en la zona. Lo que no esté aquí ni siga una terminación
# clara se deja SIN DETERMINAR: es preferible a inventar una clasificación.
FEMENINOS = {
    'MARIA', 'ROSA', 'BLANCA', 'LUZ', 'ZOILA', 'ANA', 'CARMEN', 'MARTHA', 'LAURA',
    'ELENA', 'GLORIA', 'OLGA', 'JUANA', 'INES', 'ELSA', 'NELLY', 'TERESA', 'SONIA',
    'PATRICIA', 'MERCEDES', 'ESTHER', 'DOLORES', 'CONSUELO', 'BEATRIZ', 'SUSANA',
    'VERONICA', 'MONICA', 'SANDRA', 'GLADYS', 'NARCISA', 'PIEDAD', 'AMPARO',
    'TRANSITO', 'PURIFICACION', 'ASUNCION', 'CONCEPCION', 'LUCIA', 'ISABEL',
    'MAGDALENA', 'ALEGRIA', 'JOSEFA', 'PAULINA', 'SILVIA', 'YOLANDA', 'ANITA',
    'ERLINDA', 'MERCY', 'JESSICA', 'MAYRA', 'ADRIANA', 'GABRIELA', 'FANNY',
}
MASCULINOS = {
    'SEGUNDO', 'JOSE', 'LUIS', 'JUAN', 'MANUEL', 'CARLOS', 'JORGE', 'CESAR',
    'MIGUEL', 'PEDRO', 'EDWIN', 'VICTOR', 'JAIME', 'FRANCISCO', 'ANGEL', 'WILMER',
    'VICENTE', 'CRISTIAN', 'RAFAEL', 'MARCO', 'HUGO', 'PABLO', 'FERNANDO',
    'RICARDO', 'ALFONSO', 'HECTOR', 'RAMON', 'JULIO', 'ANTONIO', 'DANIEL',
    'GERMAN', 'ALBERTO', 'GUILLERMO', 'HERNAN', 'KLEVER', 'NELSON', 'OSWALDO',
    'PATRICIO', 'RODRIGO', 'SANTIAGO', 'WASHINGTON', 'BYRON', 'DIEGO', 'EDISON',
    'ENRIQUE', 'FABIAN', 'GALO', 'IVAN', 'JAVIER', 'LEONIDAS', 'MARCELO',
    'MAURICIO', 'RAUL', 'ROBERTO', 'RUBEN', 'TELMO', 'VIRGILIO', 'WILSON',
}


def sin_tildes(s):
    s = unicodedata.normalize('NFD', (s or '').upper().strip())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')


def genero(nombres):
    """Estimación conservadora: nombre conocido, o terminación inequívoca."""
    partes = sin_tildes(nombres).split()
    if not partes:
        return None
    n = partes[0]
    if n in FEMENINOS:
        return 'Mujer'
    if n in MASCULINOS:
        return 'Hombre'
    # respaldo por terminación, solo para nombres suficientemente largos
    if len(n) >= 4:
        if n.endswith('A') and not n.endswith(('IA', 'ISTA')):
            return 'Mujer'
        if n.endswith(('O', 'OS', 'ER', 'IN')):
            return 'Hombre'
    return None


def num(p, k):
    try:
        return int(float(p.get(k) or 0))
    except (TypeError, ValueError):
        return 0


def lleno(v):
    return v not in (None, '') and str(v).strip() != ''


def pct(a, b):
    return 100.0 * a / b if b else 0.0


def cargar():
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(f'SELECT * FROM "{T}" WHERE es_ficha_hija IS NOT 1')
    pri = [dict(r) for r in cur.fetchall()]
    cur.execute(f'SELECT MAX(fecha_creacion), MAX(fecha_completado) FROM "{T}"')
    f1, f2 = cur.fetchone()
    cur.execute(f'SELECT COUNT(*) FROM "{T}" WHERE es_ficha_hija = 1 AND '
                f'coalesce(estado_investigacion, "pendiente_produccion") != "completada"')
    pendientes = cur.fetchone()[0]
    con.close()

    from generar_capas_sectores_comunidades import COM_A_SECTOR
    vistos = defaultdict(Counter)
    for p in pri:
        crudo = p.get('comunidad') or ''
        p['_comk'] = canonica(crudo) or '(sin comunidad)'
        vistos[p['_comk']][nombre_publico(crudo) or '(sin comunidad)'] += 1
        p['_sec'] = (p.get('sector_investigacion') or '').strip()
    display = {}
    for k, c in vistos.items():
        val = [(n_, v) for n_, v in c.most_common() if normalizar(n_) == k]
        display[k] = val[0][0] if val else k
    for p in pri:
        p['_com'] = display[p['_comk']]
        if p['_sec'] in ('', 'None'):
            p['_sec'] = COM_A_SECTOR.get(p['_comk'], '(sin sector)')

    corte = max(str(f1 or '')[:10], str(f2 or '')[:10])
    corte_txt = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                 if corte else 'la fecha de generación')
    return pri, corte_txt, pendientes


def main():
    pri, corte_txt, pendientes = cargar()
    N = len(pri)

    instr = Counter(str(p['nivel_instruccion']).strip() for p in pri
                    if lleno(p.get('nivel_instruccion')))
    n_ins = sum(instr.values())
    ten = Counter(str(p['tenencia_predio']).strip() for p in pri
                  if lleno(p.get('tenencia_predio')))
    n_ten = sum(ten.values())
    parr = Counter(sin_tildes(p['parroquia']).replace('OTON', 'OTÓN')
                   .replace('ASCAZUBI', 'ASCÁZUBI').title()
                   for p in pri if lleno(p.get('parroquia')))

    gen = Counter(genero(p.get('nombres')) for p in pri)
    n_gen = gen['Mujer'] + gen['Hombre']

    con_hijos = [p for p in pri
                 if lleno(p.get('hijos_hombres')) or lleno(p.get('hijos_mujeres'))]
    hijos = [num(p, 'hijos_hombres') + num(p, 'hijos_mujeres') for p in con_hijos]
    h_var = sum(num(p, 'hijos_hombres') for p in con_hijos)
    h_muj = sum(num(p, 'hijos_mujeres') for p in con_hijos)
    tel = sum(1 for p in pri if lleno(p.get('telefono_celular')))

    # tenencia por comunidad (para focalizar regularización)
    sin_titulo = []
    for com in {p['_com'] for p in pri}:
        ps = [p for p in pri if p['_com'] == com and lleno(p.get('tenencia_predio'))]
        if len(ps) < 20:
            continue
        st_ = sum(1 for p in ps if 'Posesión' in str(p['tenencia_predio']))
        sin_titulo.append((com, st_, len(ps), pct(st_, len(ps))))
    sin_titulo.sort(key=lambda x: -x[3])

    B = []
    A = B.append
    A(E.cabecera('Perfil del regante',
                 'Escolaridad, tenencia de la tierra y composición familiar · '
                 'Capítulo del informe técnico'))
    A(E.aviso_corte(corte_txt, N, pendientes))
    A(E.kpis([
        (f'{N:,}', 'fichas principales'),
        (f'{pct(instr["Primaria"] + instr["Ninguno"] + instr["Alfabetizado"], n_ins):.0f}%',
         'con instrucción básica o menos'),
        (f'{pct(ten["Posesión sin Título"], n_ten):.1f}%', 'sin título de propiedad'),
        (f'{st.mean(hijos):.1f}', 'hijos por familia'),
    ]))

    A('<h2>1. Quién es el usuario del sistema</h2>')
    A(f'<p>El padrón registra <b>{N:,} fichas principales</b> hasta la fecha de corte. Este '
      'capítulo describe sus características sociales, que condicionan tanto la '
      'forma de comunicarse con ellos como el tipo de acompañamiento técnico que '
      'requiere cualquier intervención.</p>')

    A('<h2>2. Nivel de instrucción</h2>')
    A('<table class="evitar-corte"><tr><th>Nivel alcanzado</th>'
      '<th class="n">Fichas principales</th><th>Peso</th></tr>')
    for k in ('Ninguno', 'Alfabetizado', 'Primaria', 'Secundaria', 'Superior'):
        if instr.get(k):
            A(f'<tr><td>{k}</td><td class="n">{instr[k]:,}</td>'
              f'<td>{E.barra(pct(instr[k], n_ins))}</td></tr>')
    A('</table>')
    basica = instr['Ninguno'] + instr['Alfabetizado'] + instr['Primaria']
    A(f'<div class="hallazgo"><b>Hallazgo.</b> El '
      f'<b>{pct(basica, n_ins):.1f} % de los titulares alcanzó como máximo la '
      f'primaria</b>, y un <b>{pct(instr["Ninguno"], n_ins):.1f} % no cursó ningún '
      f'nivel formal</b>. Solo el {pct(instr["Superior"], n_ins):.1f} % tiene '
      'estudios superiores. Toda comunicación técnica, reglamento o material de '
      'capacitación debe diseñarse en lenguaje sencillo y con apoyo visual: un '
      'documento escrito en registro administrativo no llega a la mayoría de los '
      'usuarios.</div>')

    A('<h2>3. Titularidad y participación de la mujer</h2>')
    A(f'<p>De los {n_gen:,} titulares cuyo nombre permite estimar el género, '
      f'<b>{gen["Mujer"]:,} son mujeres ({pct(gen["Mujer"], n_gen):.1f} %)</b> y '
      f'{gen["Hombre"]:,} hombres ({pct(gen["Hombre"], n_gen):.1f} %).</p>')
    A('<table class="evitar-corte"><tr><th>Titular</th><th class="n">Fichas principales</th>'
      '<th>Peso</th></tr>')
    for k in ('Hombre', 'Mujer'):
        A(f'<tr><td>{k}</td><td class="n">{gen[k]:,}</td>'
          f'<td>{E.barra(pct(gen[k], n_gen))}</td></tr>')
    A('</table>')
    A('<div class="nota"><b>Cómo se obtuvo este dato.</b> La ficha no pregunta el '
      'sexo del titular. La cifra es una <b>estimación</b> a partir del primer '
      'nombre, usando una lista de nombres inequívocos de la zona; los casos '
      f'ambiguos ({gen[None]:,}) quedan sin clasificar en lugar de asignarse por '
      'aproximación. Debe leerse como orden de magnitud, no como registro censal. '
      '<b>Se recomienda incorporar el campo de sexo al formulario</b> para futuras '
      'actualizaciones del padrón.</div>')

    A('<h2>4. Tenencia de la tierra</h2>')
    A('<table class="evitar-corte"><tr><th>Forma de tenencia</th>'
      '<th class="n">Fichas principales</th><th>Peso</th></tr>')
    for k, n in ten.most_common():
        A(f'<tr><td>{k}</td><td class="n">{n:,}</td>'
          f'<td>{E.barra(pct(n, n_ten))}</td></tr>')
    A('</table>')
    sp = ten['Posesión sin Título'] + ten['Herencia sin Legalizar']
    A(f'<div class="hallazgo"><b>Hallazgo.</b> <b>{sp:,} titulares '
      f'({pct(sp, n_ten):.1f} %) ocupan su predio sin título de propiedad</b>, sea '
      'por posesión o por herencia no legalizada. Es casi un tercio del padrón. '
      'La inseguridad jurídica limita el acceso a crédito y a programas públicos, '
      'y condiciona cualquier inversión predial que se quiera promover desde el '
      'sistema de riego.</div>')
    A('<h3>Comunidades con mayor proporción sin título</h3>')
    A('<p>Comunidades de veinte o más fichas principales, ordenadas por peso de la posesión '
      'sin título. Son el foco natural de un programa de regularización:</p>')
    A('<table class="evitar-corte"><tr><th>Comunidad</th><th class="n">Sin título</th>'
      '<th class="n">Total</th><th>Proporción</th></tr>')
    for com, s, t, p in sin_titulo[:8]:
        A(f'<tr><td>{com}</td><td class="n">{s:,}</td><td class="n">{t:,}</td>'
          f'<td>{E.barra(p)}</td></tr>')
    A('</table>')

    A('<h2>5. Composición familiar</h2>')
    A(f'<p>Los {len(con_hijos):,} titulares que informaron sobre su descendencia '
      f'declaran <b>{sum(hijos):,} hijos</b> en total: {h_var:,} varones y '
      f'{h_muj:,} mujeres. El promedio es de <b>{st.mean(hijos):.2f} hijos por '
      f'familia</b> (mediana {st.median(hijos):.0f}).</p>')
    dist = Counter(min(h, 6) for h in hijos)
    A('<table class="evitar-corte"><tr><th>Hijos por familia</th>'
      '<th class="n">Familias</th><th>Peso</th></tr>')
    for k in sorted(dist):
        et = f'{k}' if k < 6 else '6 o más'
        A(f'<tr><td>{et}</td><td class="n">{dist[k]:,}</td>'
          f'<td>{E.barra(pct(dist[k], len(hijos)))}</td></tr>')
    A('</table>')
    A(f'<p>Tomando el promedio declarado, los {N:,} predios empadronados '
      f'representan una población aproximada de <b>{N * (st.mean(hijos) + 2):,.0f} '
      'personas</b> vinculadas al sistema de riego, contando titular, pareja e '
      'hijos. Es una estimación de orden de magnitud sobre el alcance social del '
      'sistema.</p>')

    A('<h2>6. Distribución territorial</h2>')
    A('<table class="evitar-corte"><tr><th>Parroquia</th><th class="n">Fichas principales</th>'
      '<th>Peso</th></tr>')
    for k, n in parr.most_common():
        A(f'<tr><td>{k}</td><td class="n">{n:,}</td>'
          f'<td>{E.barra(pct(n, sum(parr.values())))}</td></tr>')
    A('</table>')
    A(f'<p>El sistema es esencialmente un servicio de la parroquia '
      f'<b>{parr.most_common(1)[0][0]}</b>, que concentra el '
      f'{pct(parr.most_common(1)[0][1], sum(parr.values())):.1f} % de los usuarios.</p>')
    A(f'<p>Se dispone de <b>teléfono de contacto para {tel:,} titulares</b> '
      f'({pct(tel, N):.1f} %), lo que permite una convocatoria directa para '
      'asambleas, capacitaciones o socialización del proyecto.</p>')

    A('<h2>7. Conclusiones</h2>')
    A('<ul>')
    A(f'<li><b>{pct(basica, n_ins):.0f} % de los titulares no superó la primaria</b>: '
      'la comunicación institucional debe ser oral, visual y en lenguaje llano.</li>')
    A(f'<li><b>Casi un tercio ({pct(sp, n_ten):.0f} %) carece de título de '
      'propiedad</b>, lo que limita su acceso a crédito e inversión.</li>')
    A(f'<li>Se estima que <b>{pct(gen["Mujer"], n_gen):.0f} % de los titulares son '
      'mujeres</b>; conviene registrar el sexo de forma explícita en el padrón.</li>')
    A(f'<li>Las familias tienen <b>{st.mean(hijos):.1f} hijos en promedio</b>, con '
      f'una población vinculada estimada en <b>{N * (st.mean(hijos) + 2):,.0f} '
      'personas</b>.</li>')
    A(f'<li>El <b>{pct(tel, N):.0f} % tiene teléfono registrado</b>: existe un canal '
      'directo de comunicación con la mayoría del padrón.</li>')
    A('</ul>')
    A(E.pie(corte_txt))

    os.makedirs(os.path.dirname(HTML), exist_ok=True)
    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(E.documento('Perfil del regante — Padrón Guanguilquí–Porotog',
                            '\n'.join(B)))
    print(f'  capítulo: {os.path.relpath(HTML, BASE)}  (corte: {corte_txt})')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    azul, blanco = PatternFill('solid', fgColor='1e4d8c'), Font(color='FFFFFF', bold=True)

    def hoja(nombre, cab, filas):
        ws = wb.create_sheet(nombre)
        ws.append(cab)
        for c in ws[1]:
            c.fill, c.font = azul, blanco
        for f_ in filas:
            ws.append(f_)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                12, min(40, max(len(str(c.value or '')) for c in col) + 2))

    hoja('Resumen', ['Indicador', 'Valor'], [
        ['Fichas principales', N],
        ['% instrucción básica o menos', round(pct(basica, n_ins), 1)],
        ['% sin instrucción formal', round(pct(instr['Ninguno'], n_ins), 1)],
        ['% sin título de propiedad', round(pct(sp, n_ten), 1)],
        ['% titulares mujeres (estimado)', round(pct(gen['Mujer'], n_gen), 1)],
        ['Hijos por familia (promedio)', round(st.mean(hijos), 2)],
        ['Población vinculada (estimada)', round(N * (st.mean(hijos) + 2))],
        ['% con teléfono registrado', round(pct(tel, N), 1)],
    ])
    hoja('Instrucción', ['Nivel', 'Fichas principales', '%'],
         [[k, v, round(pct(v, n_ins), 1)] for k, v in instr.most_common()])
    hoja('Tenencia', ['Forma de tenencia', 'Fichas principales', '%'],
         [[k, v, round(pct(v, n_ten), 1)] for k, v in ten.most_common()])
    hoja('Sin título por comunidad',
         ['Comunidad', 'Sin título', 'Total fichas principales', '% sin título'],
         [[c, s, t, round(p, 1)] for c, s, t, p in sin_titulo])
    hoja('Parroquias', ['Parroquia', 'Fichas principales'], [[k, v] for k, v in parr.most_common()])

    filas = []
    for com in sorted({p['_com'] for p in pri}):
        ps = [p for p in pri if p['_com'] == com]
        i_ = [p for p in ps if lleno(p.get('nivel_instruccion'))]
        t_ = [p for p in ps if lleno(p.get('tenencia_predio'))]
        g_ = [genero(p.get('nombres')) for p in ps]
        filas.append([
            com, ps[0]['_sec'], len(ps),
            round(pct(sum(1 for p in i_ if str(p['nivel_instruccion']).strip()
                          in ('Ninguno', 'Alfabetizado', 'Primaria')), len(i_)), 1) if i_ else None,
            round(pct(sum(1 for p in t_ if 'Posesión' in str(p['tenencia_predio'])), len(t_)), 1) if t_ else None,
            round(pct(g_.count('Mujer'), g_.count('Mujer') + g_.count('Hombre')), 1)
            if (g_.count('Mujer') + g_.count('Hombre')) else None,
        ])
    hoja('Por comunidad', ['Comunidad', 'Sector', 'Fichas principales', '% instrucción básica',
                           '% sin título', '% mujeres (est.)'], filas)

    del wb['Sheet']
    os.makedirs(os.path.dirname(XLSX), exist_ok=True)
    wb.save(XLSX)
    # Sin esto, hay builds de Excel que heredan «sin relleno» del estilo
    # base y los colores no se pintan. Ver excel_compat.py.
    from excel_compat import aplicar_formatos
    aplicar_formatos(XLSX)
    print(f'  excel   : {os.path.relpath(XLSX, BASE)}')
    print(f'\n  {N:,} fichas principales | {pct(basica, n_ins):.0f}% instrucción básica | '
          f'{pct(sp, n_ten):.0f}% sin título | {pct(gen["Mujer"], n_gen):.0f}% mujeres (est.)')


if __name__ == '__main__':
    main()
