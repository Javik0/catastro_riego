# -*- coding: utf-8 -*-
"""
Informe por comunidad — material del sociólogo del proyecto.

QUÉ ES
------
UN solo documento, largo y detallado, organizado por sector de investigación
(Sector 1, 2, 3). Cada sector abre con una introducción narrativa y luego cada
«pregunta» de la ficha catastral se vuelve un cuadro comparativo con una fila
por comunidad. Es la generalización por comunidad de lo que los 6 capítulos del
proyecto ya hacen a nivel de sistema.

Las especificaciones las dio Armando (el cliente) por WhatsApp el 24-ago-2026;
las decisiones de contenido de abajo son suyas y de JAVIKO y NO se renegocian
aquí:

 1. Universo DECLARADO («lo declarado por los comuneros») — es el material del
    sociólogo (regla 12 del proyecto). En superficie se agrega al pie la
    referencia catastral de la comunidad con la nota de que son dos mediciones
    distintas que no se suman.
 2. Caudal por la MODA, nunca sumando fichas (regla 3). Fuente única:
    public/geo/caudal_por_comunidad.json. Turnos (horas_turno) y días de riego
    (dias_riego) sí se agregan de las fichas.
 3. Cultivos: los 5 relevantes por comunidad por superficie declarada; el
    resto agrupado como «otros».
 4. Superficie de uso pecuario: derivada por comunidad usando el cultivo
    «Pasto mejorado» cruzado con los datos pecuarios (instrucción textual de
    Armando: «se puede obtener el dato por comunidad usando el pasto mejorado
    con lo pecuario»).
 5. «Otros usos + superficie» NO va — Armando: «no se ha hecho relevante en
    este proyecto».
 6. El dato de P001 (ALPAKA) queda pendiente «a lo que las comunidades
    definan» — nota estándar donde aplique, sin inventar el dato.
 7. SIN codificación S01-C01-R01-F01 (JAVIKO la dejó pendiente el 24-ago).
 8. SIN listados nominales de regantes: solo agregados por comunidad. El
    respaldo por ficha existe en el Excel Catastral Premium.
 9. Riego tecnificado = aspersión + goteo; inundación = gravedad (campos
    metodo_*_pct de la ficha).

REGLAS DURAS DEL PROYECTO QUE ESTE SCRIPT RESPETA
-------------------------------------------------
· El nombre de comunidad se canoniza SOLO con comunidades_canon.py (regla 4).
· Personas ≠ predios (regla 6): perfil (instrucción, hijos, presa,
  capacitación) SOLO de fichas principales; superficie y producción de TODAS
  (principales + adicionales completadas). Nunca se suman los dos universos.
· «Sin riego», nunca «secano» (regla 9).
· Las fichas hijas PENDIENTES se excluyen de producción (duplican datos
  heredados de la madre). Hoy hay 0 pendientes; se filtran igual.
· NO se toca el data.gpkg: todo se lee de los GeoJSON de public/geo/.
· ALPAKA es un fraccionamiento digitalizado desde planos de loteo, no una
  encuesta de campo: sus tarifas (672/308 USD «mensuales») se excluyen del
  análisis económico como en generar_capitulo_riego.py, y sus filas de perfil
  llevan nota. La granja avícola de Asociación Rosalía (registros de 10.000
  aves por titular sobre el mismo predio) se excluye del inventario pecuario,
  como en los pendientes de terceros del proyecto.
· El caudal de AVELLANEDA, SR. HERNÁN TIMPE y HDA. SAN FRANSISCO es HEREDADO
  (misma llave que su comunidad de origen): se muestra pero no se suma al
  total del sector, igual que hace caudal_por_comunidad.json.

SALIDAS
-------
  docs/INFORME-SOCIOLOGO-por-comunidad.html    documento imprimible (estilo casa)
  docs/INFORME-SOCIOLOGO-por-comunidad.md      fuente Markdown (para md_a_docx.py)
  build_entrega/Informe_Sociologo_Matrices.xlsx  matrices crudas, una hoja por bloque

  Con --sector N genera SOLO ese sector como muestra (archivos con sufijo
  -MUESTRA-sectorN) sin tocar los definitivos.

USO
---
  python -X utf8 scripts/generar_informe_sociologo.py --sector 1   # muestra
  python -X utf8 scripts/generar_informe_sociologo.py              # completo

Se corre con el Python del PATH (C:\\Python314): nada de aquí lee el data.gpkg.
"""

import argparse
import json
import os
import re
import statistics as st
import sys
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica  # noqa: E402
import informe_estilo as E  # noqa: E402

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
GEO = os.path.join(BASE, 'public', 'geo')
CONSTANTS_TS = os.path.join(BASE, 'src', 'lib', 'constants.ts')
HTML_OUT = os.path.join(BASE, 'docs', 'INFORME-SOCIOLOGO-por-comunidad.html')
MD_OUT = os.path.join(BASE, 'docs', 'INFORME-SOCIOLOGO-por-comunidad.md')
XLSX_OUT = os.path.join(BASE, 'build_entrega', 'Informe_Sociologo_Matrices.xlsx')

MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
         'octubre noviembre diciembre').split()
# Carpeta (relativa a docs/) donde quedan los mapas en disco, para que el
# Markdown y el Word puedan apuntarlos. El HTML los lleva embebidos.
DIR_MAPAS = 'mapas-sociologo'

# Fecha que cita el documento. Es EDITORIAL, no calculada: el 19 de agosto de
# 2026 es cuando se publicó y entregó el paquete al consorcio, y por eso todos
# los documentos deben citar la misma referencia (decisión de JAVIKO,
# 31-ago-2026). Verificado que las cifras no se han movido desde entonces: la
# depuración posterior (corrección de Coyago, 30-ago) no alteró ningún
# agregado. Si una sincronización futura cambia las cifras, hay que mover esta
# fecha o el documento estaría citando datos que no son los de ese día.
FECHA_CORTE = '19 de agosto de 2026'

NOTA_P001 = ('ALPAKA: el dato de las fichas P001 (representación del total del '
             'fraccionamiento) queda pendiente a lo que las comunidades definan.')
NOTA_ALPAKA_LOTEO = ('ALPAKA es un fraccionamiento digitalizado desde planos de '
                     'loteo, no una encuesta de campo: sus respuestas provienen '
                     'de ese registro.')
NOTA_UNIV_TODAS = ('Universo: lo declarado en TODAS las fichas (principales + '
                   'adicionales completadas). Cada ficha es un predio.')
NOTA_UNIV_PRI = ('Universo: fichas PRINCIPALES (una por titular entrevistado; '
                 'las adicionales heredan datos de su ficha principal y '
                 'duplicarían la respuesta).')


def num(d, k):
    try:
        return float(d.get(k) or 0)
    except (TypeError, ValueError):
        return 0.0


def lleno(v):
    return v not in (None, '') and str(v).strip() != ''


def pct(a, b):
    return 100.0 * a / b if b else 0.0


# Bajo este número de respuestas un porcentaje no describe una tendencia sino
# un caso individual: con una sola ficha, «0 %» y «100 %» son la misma
# respuesta de una persona dibujada como si fuera una estadística. Siete
# comunidades del padrón están en esa situación —las haciendas y los titulares
# individuales, que tienen una o dos fichas—. Detectado por JAVIKO el
# 31-ago-2026 al ver a Avellaneda con 0,0 % y barra vacía.
MIN_RESPUESTAS = 5
NOTA_POCAS = ('En las comunidades con menos de 5 respuestas se muestra el '
              'conteo en lugar del porcentaje: con tan pocas fichas el '
              'porcentaje describe un caso individual, no una tendencia.')


def pct_o_conteo(a, b):
    """Porcentaje (float → se dibuja con barra) o «a de b» (str → texto) si la
    base es demasiado pequeña para que el porcentaje signifique algo."""
    if b < MIN_RESPUESTAS:
        return f'{f0(a)} de {f0(b)}'
    return pct(a, b)


def f0(n):
    return f'{n:,.0f}'


def f1(n):
    return f'{n:,.1f}'


def f2(n):
    return f'{n:,.2f}'


# ─── Catálogo oficial de comunidades ─────────────────────────────────────────
# Se lee de src/lib/constants.ts (CATALOGO_COMUNIDADES) para no duplicar la
# fuente: el número y el nombre oficial son los del listado del consorcio.

def cargar_catalogo():
    with open(CONSTANTS_TS, encoding='utf-8') as f:
        ts = f.read()
    patron = re.compile(
        r"\{\s*n:\s*(\d+),\s*sector:\s*'([^']+)',\s*oficial:\s*'([^']+)',"
        r"\s*datos:\s*'([^']+)'(,\s*oculta:\s*true)?\s*\}")
    cat = []
    for m in patron.finditer(ts):
        if m.group(5):  # oculta: no se investiga
            continue
        cat.append({'n': int(m.group(1)), 'sector': m.group(2),
                    'oficial': m.group(3), 'datos': m.group(4),
                    'key': canonica(m.group(4))})
    if len(cat) != 50:
        print(f'⚠ Catálogo: se esperaban 50 comunidades visibles, salieron {len(cat)}')
    return cat


# ─── Carga de datos (solo GeoJSON publicados; el data.gpkg no se toca) ───────

def cargar_datos():
    def leer(nombre):
        with open(os.path.join(GEO, nombre), encoding='utf-8') as f:
            return json.load(f)

    fichas = [x['properties'] for x in leer('fichas_predios.geojson')['features']]
    cultivos = leer('cultivos.json')
    animales = leer('animales.json')
    sup = leer('superficie_por_comunidad.json')
    caudal = leer('caudal_por_comunidad.json')

    # hijas PENDIENTES fuera de producción (duplican lo heredado de la madre)
    todas = [p for p in fichas
             if not (p.get('es_ficha_hija') == 1 and
                     (p.get('estado_investigacion') or '') != 'completada')]
    pri = [p for p in fichas if p.get('es_ficha_hija') != 1]

    for p in fichas:
        p['_key'] = canonica(p.get('comunidad') or '') or '(sin comunidad)'

    corte = max(max(str(p.get('fecha_creacion') or '')[:10] for p in fichas),
                max(str(p.get('fecha_completado') or '')[:10] for p in fichas))
    corte_txt = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                 if corte else 'la fecha de generación')

    return fichas, todas, pri, cultivos, animales, sup, caudal, corte_txt


# ─── Agregación por comunidad ────────────────────────────────────────────────

GRUPOS_PECUARIOS = {
    'Vacas en producción': 'Bovinos', 'Vacas secas': 'Bovinos',
    'Vaconas': 'Bovinos', 'Toros': 'Bovinos', 'Toretes': 'Bovinos',
    'Terneros': 'Bovinos', 'Terneras': 'Bovinos',
    'Ovejas / Cabras': 'Ovinos y caprinos',
    'Porcino (Chanchos)': 'Porcinos',
    'Cuyes / Conejos': 'Cuyes y conejos',
    'Gallinas de campo': 'Aves', 'Gallinas ponedoras': 'Aves',
    'Pollos de engorde': 'Aves',
    'Equinos': 'Otras especies', 'Otros': 'Otras especies',
}

# Umbral de la granja avícola de Asociación Rosalía: registros de 10.000 aves
# por titular sobre el mismo predio, excluidos del informe (pendiente de
# terceros del proyecto). Ningún registro legítimo del padrón se le acerca.
GRANJA_AVICOLA_MIN = 10000


def clasificar_tema_cap(texto):
    """Los temas de capacitación son texto libre de los técnicos; se agrupan
    por palabra clave. RIEGO va primero porque domina la intención («manejo de
    sistemas de riego para la producción de nuevos cultivos» es riego)."""
    t = canonica(texto)  # MAYÚSCULAS sin acentos
    if not t:
        return None
    if 'RIEGO' in t or 'AGUA' in t:
        return 'Manejo de sistemas de riego'
    if 'AGROECOLOG' in t or 'ORGANIC' in t or 'ABONO' in t:
        return 'Agroecología y abonos'
    if 'PECUARI' in t or 'GANAD' in t or 'ANIMAL' in t or 'ESPECIES MENORES' in t:
        return 'Producción pecuaria'
    if 'CULTIV' in t or 'AGRICOL' in t or 'SIEMBRA' in t or 'SEMILLA' in t or 'PRODUCC' in t:
        return 'Producción agrícola'
    return 'Otros temas'


def agregar_comunidad(key, todas, pri, cult_por_ficha, anim_por_ficha,
                      sup_com, caudal_com, heredado_de):
    """Todas las cifras de una comunidad, en crudo. Cada bloque nombra su
    universo (regla 6): 'todas' para tierra y producción, 'pri' para personas
    y para lo que describe la entrega de agua al entrevistado."""
    a = {'key': key}

    # ── padrón y superficie: fuente única superficie_por_comunidad.json ──
    s = sup_com or {}
    a['fichas'] = int(s.get('fichas') or 0)
    a['regantes'] = int(s.get('regantes') or 0)
    a['adicionales'] = a['fichas'] - a['regantes']
    a['predios_catastrales'] = int(s.get('predios_catastrales') or 0)
    a['sup_declarada'] = s.get('superficie_declarada_ha') or 0.0
    a['sup_riego_decl'] = s.get('riego_declarado_ha') or 0.0
    a['sup_sin_riego_decl'] = s.get('sin_riego_declarado_ha') or 0.0
    a['sup_catastral'] = s.get('superficie_catastral_ha') or 0.0

    # ── tecnificación (pri): promedio de los % declarados por ficha ──
    con_met = [p for p in pri if num(p, 'metodo_gravedad_pct') +
               num(p, 'metodo_aspersion_pct') + num(p, 'metodo_goteo_pct') > 0]
    a['met_n'] = len(con_met)
    if con_met:
        a['met_gravedad'] = st.mean(num(p, 'metodo_gravedad_pct') for p in con_met)
        a['met_aspersion'] = st.mean(num(p, 'metodo_aspersion_pct') for p in con_met)
        a['met_goteo'] = st.mean(num(p, 'metodo_goteo_pct') for p in con_met)
    else:
        a['met_gravedad'] = a['met_aspersion'] = a['met_goteo'] = 0.0
    a['met_tecnificado'] = a['met_aspersion'] + a['met_goteo']

    # ── cultivos (todas): top 5 por superficie + «otros» ──
    ha_por_tipo = defaultdict(float)
    for c in cult_por_ficha:
        tipo = (c.get('tipo_cultivo') or '').strip() or '(sin tipo)'
        ha_por_tipo[tipo] += (c.get('superficie_m2') or 0) / 10000.0
    orden = sorted(ha_por_tipo.items(), key=lambda kv: -kv[1])
    a['cultivos_top'] = orden[:5]
    a['cultivos_otros_ha'] = sum(ha for _, ha in orden[5:])
    a['cultivos_total_ha'] = sum(ha_por_tipo.values())
    a['pasto_mejorado_ha'] = ha_por_tipo.get('Pasto mejorado', 0.0)

    # ── pecuario (todas): cabezas por grupo, granja avícola excluida ──
    cab = defaultdict(int)
    a['aves_granja_excluidas'] = 0
    for an in anim_por_ficha:
        n = int(an.get('cantidad') or 0)
        if n <= 0:
            continue
        grupo = GRUPOS_PECUARIOS.get((an.get('especie') or '').strip(),
                                     'Otras especies')
        if grupo == 'Aves' and n >= GRANJA_AVICOLA_MIN:
            a['aves_granja_excluidas'] += n
            continue
        cab[grupo] += n
    a['pecuario'] = dict(cab)
    a['bovinos'] = cab.get('Bovinos', 0)
    a['carga_bovina'] = (a['bovinos'] / a['pasto_mejorado_ha']
                         if a['pasto_mejorado_ha'] > 0 else None)

    # ── tenencia (pri con respuesta) ──
    ten = [str(p.get('tenencia_predio')).strip() for p in pri
           if lleno(p.get('tenencia_predio'))]
    a['ten_n'] = len(ten)
    a['ten_escritura'] = sum(1 for t in ten if 'Escritura' in t)
    a['ten_posesion'] = sum(1 for t in ten if 'Posesión' in t)
    a['ten_otras'] = len(ten) - a['ten_escritura'] - a['ten_posesion']

    # ── el agua (caudal de la fuente única; turnos/días de las fichas) ──
    a['caudal_ls'] = (caudal_com or {}).get('caudal_ls')
    a['caudal_origen'] = (caudal_com or {}).get('origen', '')
    a['caudal_heredado_de'] = heredado_de
    ht = [num(p, 'horas_turno') for p in pri if lleno(p.get('horas_turno'))]
    dr = [num(p, 'dias_riego') for p in pri if lleno(p.get('dias_riego'))]
    a['horas_turno_med'] = st.median(ht) if ht else None
    a['dias_riego_med'] = st.median(dr) if dr else None

    frec = Counter(str(p.get('frecuencia_riego')).strip() for p in pri
                   if lleno(p.get('frecuencia_riego')))
    a['frec_n'] = sum(frec.values())
    a['frec'] = dict(frec)
    a['frec_predominante'] = frec.most_common(1)[0][0] if frec else None

    # ── tarifas (pri con valor; ALPAKA excluida como en el capítulo de riego) ──
    # La mediana se calcula POR MODALIDAD: mezclar USD/mes con USD/año en una
    # sola mediana daría una cifra sin unidad.
    if key == 'ALPAKA':
        a['tarifa_mensual_med'] = a['tarifa_anual_med'] = None
        a['tarifa_mensual_n'] = a['tarifa_anual_n'] = 0
    else:
        tar = [(num(p, 'valor_tarifa'), str(p.get('tipo_tarifa') or '').strip())
               for p in pri if lleno(p.get('valor_tarifa'))]
        mens = [v for v, t in tar if 'mensual' in t]
        anua = [v for v, t in tar if 'anual' in t]
        a['tarifa_mensual_n'] = len(mens)
        a['tarifa_anual_n'] = len(anua)
        a['tarifa_mensual_med'] = st.median(mens) if mens else None
        a['tarifa_anual_med'] = st.median(anua) if anua else None

    res = Counter(str(p.get('tiene_reservorio')).strip() for p in pri
                  if lleno(p.get('tiene_reservorio')))
    a['res_n'] = sum(res.values())
    a['res_comunitario'] = res.get('Comunitario', 0)
    a['res_privado'] = res.get('Privado', 0)
    a['res_no'] = res.get('No', 0)

    # ── las personas (pri; regla 6) ──
    ins = Counter(str(p.get('nivel_instruccion')).strip() for p in pri
                  if lleno(p.get('nivel_instruccion')))
    a['ins_n'] = sum(ins.values())
    a['instruccion'] = dict(ins)

    hijos = [num(p, 'hijos_hombres') + num(p, 'hijos_mujeres') for p in pri
             if lleno(p.get('hijos_hombres')) or lleno(p.get('hijos_mujeres'))]
    a['hijos_n'] = len(hijos)
    a['hijos_prom'] = st.mean(hijos) if hijos else None

    a['n_pri'] = len(pri)
    # Servicios de la VIVIENDA: el denominador son las fichas que declaran
    # material de construcción, no todas. Regla 2 del cliente (9-ago-2026):
    # «sin material de construcción no hay vivienda: agua y luz vacías son la
    # respuesta correcta, no una omisión». Con el denominador completo el agua
    # aparentaría 64,9 % cuando entre viviendas es 96,3 %.
    con_viv = [p for p in pri if lleno(p.get('material_construccion'))]
    a['con_vivienda'] = len(con_viv)
    a['agua_consumo'] = sum(1 for p in con_viv if lleno(p.get('agua_consumo')))
    a['energia'] = sum(1 for p in con_viv if lleno(p.get('energia_electrica')))
    # El teléfono es dato de contacto de la persona, no de la casa: su
    # denominador sí son todas las fichas principales.
    a['telefono'] = sum(1 for p in pri if lleno(p.get('telefono_celular')) or
                        lleno(p.get('telefono_casa')))

    # Material de la vivienda y altitud: venían del capítulo «Servicios
    # básicos y hábitat», integrado aquí el 31-ago-2026 para que todo lo que
    # se mira por comunidad viva en un solo documento.
    mat = Counter(str(p.get('material_construccion')).strip() for p in con_viv)
    a['materiales'] = dict(mat)
    a['mat_top'] = mat.most_common(2)
    cotas = [num(p, 'cota_msnm') for p in pri if lleno(p.get('cota_msnm'))]
    a['cota_n'] = len(cotas)
    a['cota_min'] = min(cotas) if cotas else None
    a['cota_med'] = st.median(cotas) if cotas else None
    a['cota_max'] = max(cotas) if cotas else None

    presa = Counter(str(p.get('conoce_presa')).strip() for p in pri
                    if lleno(p.get('conoce_presa')))
    a['presa_n'] = sum(presa.values())
    a['presa_si'] = presa.get('Sí', 0)

    cap = Counter(str(p.get('recibio_capacitacion')).strip() for p in pri
                  if lleno(p.get('recibio_capacitacion')))
    a['cap_n'] = sum(cap.values())
    a['cap_si'] = cap.get('Sí', 0)
    des = Counter(str(p.get('le_gustaria_cap')).strip() for p in pri
                  if lleno(p.get('le_gustaria_cap')))
    a['des_n'] = sum(des.values())
    a['des_si'] = des.get('Sí', 0)
    temas = Counter()
    for p in pri:
        t = clasificar_tema_cap(p.get('temas_capacitacion') or '')
        if t:
            temas[t] += 1
    a['temas_top'] = temas.most_common(3)

    return a


def agregar_todo():
    catalogo = cargar_catalogo()
    fichas, todas, pri, cultivos, animales, sup, caudal, corte_txt = cargar_datos()

    sup_por_key = {canonica(c['comunidad']): c for c in sup['comunidades']}
    caudal_por_key = {canonica(k): v for k, v in caudal['comunidades'].items()}
    heredado = {canonica(k): v for k, v in caudal.get('caudal_heredado', {}).items()}

    todas_por_key = defaultdict(list)
    for p in todas:
        todas_por_key[p['_key']].append(p)
    pri_por_key = defaultdict(list)
    for p in pri:
        pri_por_key[p['_key']].append(p)

    id_a_key = {p.get('id'): p['_key'] for p in todas}
    cult_por_key = defaultdict(list)
    for c in cultivos:
        k = id_a_key.get(c.get('ficha_id'))
        if k:
            cult_por_key[k].append(c)
    anim_por_key = defaultdict(list)
    for an in animales:
        k = id_a_key.get(an.get('ficha_id'))
        if k:
            anim_por_key[k].append(an)

    comunidades = []
    for c in catalogo:
        k = c['key']
        a = agregar_comunidad(k, todas_por_key.get(k, []), pri_por_key.get(k, []),
                              cult_por_key.get(k, []), anim_por_key.get(k, []),
                              sup_por_key.get(k), caudal_por_key.get(k),
                              heredado.get(k))
        a.update(n=c['n'], oficial=c['oficial'], datos=c['datos'],
                 sector=c['sector'])
        comunidades.append(a)

    return comunidades, sup, caudal, corte_txt, fichas


# ─── Autoverificación contra las fuentes únicas ──────────────────────────────

def verificar(comunidades, sup, caudal, fichas):
    """Si algún total del documento no cuadra con la fuente única, se avisa en
    consola ANTES de que el documento salga. No corrige nada: avisa."""
    avisos = []
    tot = sup['total']

    d = sum(c['sup_declarada'] for c in comunidades)
    if abs(d - tot['superficie_declarada_ha']) > 0.5:
        avisos.append(f"superficie declarada: cuadros {d:,.2f} ≠ fuente "
                      f"{tot['superficie_declarada_ha']:,.2f}")

    nf = sum(c['fichas'] for c in comunidades)
    if nf != tot['fichas'] - sum(1 for p in fichas if p['_key'] == '(sin comunidad)'):
        avisos.append(f"fichas: cuadros {nf} ≠ fuente {tot['fichas']} "
                      f"(menos las sin comunidad)")

    q_sect = sum(c['caudal_ls'] or 0 for c in comunidades
                 if not c['caudal_heredado_de'])
    q_ref = caudal['totales']['caudal_comunidades_ls']
    if abs(q_sect - q_ref) > 0.1:
        avisos.append(f"caudal de comunidades: cuadros {q_sect:,.2f} ≠ fuente "
                      f"{q_ref:,.2f}")

    sin_caudal = [c['oficial'] for c in comunidades if c['caudal_ls'] is None]
    if sin_caudal:
        avisos.append(f"comunidades sin caudal en la fuente: {sin_caudal}")

    sin_sup = [c['oficial'] for c in comunidades if not c['fichas']]
    if sin_sup:
        avisos.append(f"comunidades sin fichas en la fuente de superficie: {sin_sup}")

    for a in avisos:
        print(f'⚠ NO CUADRA — {a}')
    if not avisos:
        print(f"✔ Cuadre contra fuentes únicas: declarada "
              f"{tot['superficie_declarada_ha']:,.2f} ha · "
              f"{tot['fichas']:,} fichas · {tot['regantes']:,} regantes · "
              f"caudal {caudal['totales']['caudal_sistema_ls']:,.2f} l/s "
              f"({q_ref:,.2f} de comunidades + "
              f"{caudal['totales']['caudal_individual_ls']:,.2f} individuales)")
    return avisos


# ─── Mapas por sector (dibujados desde las capas publicadas) ─────────────────
# Decisión de JAVIKO (29-ago-2026): los mapas se GENERAN desde
# comunidades.geojson, no se capturan de la web — una captura manual se
# desactualiza en silencio con cada sincronización y trae la interfaz encima.
# Cuatro mapas en total (uno general + uno por sector), incrustados solo en el
# HTML (el conversor a Word no maneja imágenes). La numeración de las
# etiquetas es la del listado oficial, la misma de los cuadros.

COLOR_SECTOR = {'Sector 1': '#2e7d4f', 'Sector 2': '#1e4d8c',
                'Sector 3': '#e0a800'}
# fondos suaves para distinguir comunidades vecinas dentro de un sector
PALETA_COM = ['#bfe3cc', '#c7d9f2', '#f6dcb5', '#dcc9e8', '#f2c9c9',
              '#cde8e4', '#e6e2b8', '#d9d2c2']


def _anillos(geom):
    """Anillos exteriores de un (Multi)Polygon GeoJSON, como listas [x, y]."""
    if geom['type'] == 'Polygon':
        return [geom['coordinates'][0]]
    return [p[0] for p in geom['coordinates']]


def _centro_etiqueta(anillos):
    """Punto para la etiqueta: promedio de vértices del anillo más grande
    (suficiente para estos polígonos; no hace falta un centroide exacto)."""
    mayor = max(anillos, key=len)
    xs = [p[0] for p in mayor]
    ys = [p[1] for p in mayor]
    return sum(xs) / len(xs), sum(ys) / len(ys)


def _separar_etiquetas(comunidades, d_min=650.0):
    """Separa las etiquetas que quedarían una sobre otra (Armando detectó el
    16 montado sobre el 1, 30-ago-2026). Repulsión simple por pares en
    metros Mercator: si dos centros quedan a menos de `d_min`, se empujan a
    partes iguales sobre la línea que los une, unas pocas pasadas. La
    posición ajustada va a `c['etiqueta']`; el centro real no se toca."""
    import math
    pos = [list(c['centro']) for c in comunidades]
    for _ in range(40):
        movido = False
        for i in range(len(pos)):
            for j in range(i + 1, len(pos)):
                dx = pos[j][0] - pos[i][0]
                dy = pos[j][1] - pos[i][1]
                d = math.hypot(dx, dy)
                if d < d_min:
                    if d < 1:
                        dx, dy, d = 1.0, 0.0, 1.0
                    e = (d_min - d) / 2 / d
                    pos[i][0] -= dx * e
                    pos[i][1] -= dy * e
                    pos[j][0] += dx * e
                    pos[j][1] += dy * e
                    movido = True
        if not movido:
            break
    for c, p in zip(comunidades, pos):
        c['etiqueta'] = tuple(p)


def _mercator(lon, lat):
    """WGS84 → Web Mercator (EPSG:3857), el sistema de los mosaicos. A la
    latitud del proyecto (~0°) el metro Mercator coincide con el real, así
    que la barra de escala sigue siendo válida."""
    import math
    R = 6378137.0
    return (math.radians(lon) * R,
            math.log(math.tan(math.pi / 4 + math.radians(lat) / 2)) * R)


def descargar_base_satelital(limites, ancho_px=4200):
    """Mosaico satelital de Esri World Imagery para la extensión del sistema
    (la misma base del mapa web del proyecto). Se descarga UNA vez y se
    reutiliza en los cuatro mapas. Si no hay red, se avisa y los mapas salen
    sobre fondo blanco — el documento nunca deja de generarse por esto.
    Devuelve (imagen PIL, extensión mercator (x0, x1, y0, y1)) o (None, None).
    """
    import io as _io
    import math
    import urllib.request
    from concurrent.futures import ThreadPoolExecutor
    try:
        from PIL import Image
    except ImportError:
        print('⚠ Mapas sin base satelital: falta Pillow')
        return None, None

    x0, x1, y0, y1 = limites
    for z in range(10, 17):
        tx0 = int((x0 + 180) / 360 * 2 ** z)
        tx1 = int((x1 + 180) / 360 * 2 ** z)
        if (tx1 - tx0 + 1) * 256 >= ancho_px:
            break

    def fila(lat):
        lr = math.radians(lat)
        return int((1 - math.log(math.tan(lr) + 1 / math.cos(lr)) / math.pi)
                   / 2 * 2 ** z)

    ty0, ty1 = fila(y1), fila(y0)  # el norte es la fila menor
    tiles = [(tx, ty) for ty in range(ty0, ty1 + 1)
             for tx in range(tx0, tx1 + 1)]
    mosaico = Image.new('RGB', ((tx1 - tx0 + 1) * 256, (ty1 - ty0 + 1) * 256),
                        (238, 242, 247))

    def traer(t):
        tx, ty = t
        url = ('https://server.arcgisonline.com/ArcGIS/rest/services/'
               f'World_Imagery/MapServer/tile/{z}/{ty}/{tx}')
        req = urllib.request.Request(url, headers={'User-Agent': 'padron-app'})
        # un mosaico perdido deja un rectángulo gris en pleno mapa: 3 intentos
        for _ in range(3):
            try:
                with urllib.request.urlopen(req, timeout=25) as r:
                    img = Image.open(_io.BytesIO(r.read())).convert('RGB')
                mosaico.paste(img, ((tx - tx0) * 256, (ty - ty0) * 256))
                return True
            except Exception:
                continue
        return False

    with ThreadPoolExecutor(max_workers=8) as ex:
        ok = sum(ex.map(traer, tiles))
    if ok < len(tiles) * 0.7:
        print(f'⚠ Base satelital incompleta ({ok}/{len(tiles)} mosaicos): '
              'los mapas salen sobre fondo blanco')
        return None, None

    R = 6378137.0
    mundo = 2 * math.pi * R
    ext = (tx0 / 2 ** z * mundo - mundo / 2,
           (tx1 + 1) / 2 ** z * mundo - mundo / 2,
           mundo / 2 - (ty1 + 1) / 2 ** z * mundo,
           mundo / 2 - ty0 / 2 ** z * mundo)
    print(f'✔ Base satelital Esri: {ok}/{len(tiles)} mosaicos (zoom {z})')
    return mosaico, ext


def generar_mapas(catalogo, pdf_ruta=None):
    """Imagen en base64 por mapa: {'general': ..., 'Sector 1': ..., ...}.
    Con `pdf_ruta`, además escribe las mismas cuatro láminas en un PDF (una
    por página, texto vectorial): el anexo cartográfico. Se generan aquí y no
    con el Diseñador de Impresión de la web porque su exportación a PDF está
    rota al 29-ago-2026 (mapa gris sin base y extensión mal compuesta).
    Se dibuja en Web Mercator sobre el mosaico satelital de Esri (la misma
    base del mapa web del proyecto, decisión de JAVIKO del 29-ago-2026);
    sin red, los mapas salen sobre fondo blanco y el documento no se detiene.
    Los cuatro mapas comparten extensión, para que siempre se vea dónde cae
    cada sector dentro del sistema.

    POLÍGONOS: la capa oficial de comunas entregada por el GADM, ya recortada
    al área de estudio (public/geo/comunas_oficiales.geojson, 53 de 117) —
    decisión de JAVIKO del 31-ago-2026; antes se dibujaba el dissolve de los
    predios (comunidades.geojson), que ahora solo aporta la ubicación de las
    etiquetas 1-50 de las organizaciones de riego. El sector de cada comuna se
    asigna por CRUCE ESPACIAL (mayor área de intersección con las
    organizaciones del sector): el cruce por nombre está prohibido en el
    proyecto — una comuna puede contener varias organizaciones y hay nombres
    iguales que designan sitios distintos."""
    import base64
    import io
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patheffects as pe
    from shapely.geometry import shape
    from shapely.ops import unary_union

    with open(os.path.join(GEO, 'comunidades.geojson'), encoding='utf-8') as f:
        capa = json.load(f)['features']

    cat_por_key = {c['key']: c for c in catalogo}
    comunidades = []
    for feat in capa:
        key = canonica(feat['properties'].get('comunidad') or '')
        c = cat_por_key.get(key)
        if not c:
            print(f'⚠ Mapa: comunidad de la capa sin catálogo: {key}')
            continue
        anillos = [[_mercator(*p) for p in a]
                   for a in _anillos(feat['geometry'])]
        comunidades.append({'n': c['n'], 'sector': c['sector'],
                            'oficial': c['oficial'], 'anillos': anillos,
                            'centro': _centro_etiqueta(anillos),
                            'geom': shape(feat['geometry'])})

    _separar_etiquetas(comunidades)

    # ── capa oficial de comunas, con su sector asignado espacialmente ──
    with open(os.path.join(GEO, 'comunas_oficiales.geojson'),
              encoding='utf-8') as f:
        capa_comunas = json.load(f)['features']
    geom_sector = {s: unary_union([c['geom'] for c in comunidades
                                   if c['sector'] == s])
                   for s in COLOR_SECTOR}
    comunas = []
    sin_sector = []
    for feat in capa_comunas:
        g = shape(feat['geometry'])
        sector, area = None, 0.0
        for s, sg in geom_sector.items():
            a = g.intersection(sg).area
            if a > area:
                sector, area = s, a
        anillos = [[_mercator(*p) for p in a_]
                   for a_ in _anillos(feat['geometry'])]
        comunas.append({'nombre': feat['properties'].get('comuna') or '',
                        'sector': sector, 'anillos': anillos})
        if sector is None:
            sin_sector.append(feat['properties'].get('comuna') or '?')
    if sin_sector:
        print(f'ℹ Mapa: {len(sin_sector)} comunas sin organización de riego '
              f'encima (van en gris neutro): {sin_sector}')

    # extensión: en grados para pedir mosaicos, en mercator para dibujar.
    # Manda la capa dibujada (las comunas); los centros de las etiquetas
    # entran también por si alguna organización quedara fuera del recorte.
    todos = [p for c in comunas for a in c['anillos'] for p in a] + \
            [c['centro'] for c in comunidades]
    xs, ys = [p[0] for p in todos], [p[1] for p in todos]
    margen = 900  # metros
    limites = (min(xs) - margen, max(xs) + margen,
               min(ys) - margen, max(ys) + margen)
    import math
    R = 6378137.0
    grados = (math.degrees(limites[0] / R), math.degrees(limites[1] / R),
              math.degrees(2 * math.atan(math.exp(limites[2] / R)) - math.pi / 2),
              math.degrees(2 * math.atan(math.exp(limites[3] / R)) - math.pi / 2))
    base, base_ext = descargar_base_satelital(grados)

    halo = [pe.withStroke(linewidth=2.4, foreground='white')]
    halo_blanco = [pe.withStroke(linewidth=2.2, foreground='#1a1a1a')]

    def dibujar(nombre_sector, archivo=None):
        con_leyenda = nombre_sector is not None
        fig, ax = plt.subplots(figsize=(7.4, 6.4), dpi=150)
        fig.subplots_adjust(left=0.01, right=0.70 if con_leyenda else 0.99,
                            top=0.99, bottom=0.01)
        if base is not None:
            ax.imshow(base, extent=(base_ext[0], base_ext[1],
                                    base_ext[2], base_ext[3]),
                      interpolation='bilinear', zorder=0)

        destacadas = [c for c in comunidades
                      if nombre_sector is None or c['sector'] == nombre_sector]
        # Los polígonos son las COMUNAS OFICIALES; se destacan las del sector
        # del capítulo. Dentro del sector cada comuna toma un fondo distinto
        # de la paleta para distinguir vecinas, igual que hacía el dissolve.
        idx_paleta = 0
        for c in comunas:
            es = (nombre_sector is None and c['sector'] is not None) or \
                 c['sector'] == nombre_sector
            if c['sector'] is None:
                # comuna del recorte sin organización de riego encima
                relleno, alfa = '#d7dce3', 0.45 if base is not None else 1.0
            elif nombre_sector is None:
                relleno, alfa = COLOR_SECTOR[c['sector']], 0.5
            elif es:
                relleno = PALETA_COM[idx_paleta % len(PALETA_COM)]
                idx_paleta += 1
                alfa = 0.55
            else:
                # el resto del sistema, atenuado sobre el satélite
                relleno, alfa = '#ffffff', 0.4 if base is not None else 1.0
            # el borde de comuna va marcado (pedido de Armando, 30-ago):
            # blanco sobre el relleno de sector en el mapa general, azul
            # oscuro en el mapa del sector
            for a in c['anillos']:
                ax.fill([p[0] for p in a], [p[1] for p in a],
                        facecolor=relleno, alpha=alfa,
                        edgecolor=('white' if nombre_sector is None
                                   else '#24405e') if es else '#8896a8',
                        linewidth=0.9 if es else 0.3, zorder=2 if es else 1)
        for c in destacadas:
            ax.annotate(str(c['n']), c['etiqueta'], ha='center', va='center',
                        fontsize=7.5, fontweight='bold', color='#24405e',
                        zorder=3, path_effects=halo)

        # barra de escala de 2 km, norte y crédito de la base
        x0, x1, y0, y1 = limites
        bx, by = x0 + (x1 - x0) * 0.04, y0 + (y1 - y0) * 0.045
        ax.plot([bx, bx + 2000], [by, by], color='white' if base is not None
                else '#24405e', linewidth=2.5, zorder=4,
                path_effects=halo_blanco if base is not None else [])
        ax.annotate('2 km', (bx + 1000, by + (y1 - y0) * 0.014), ha='center',
                    fontsize=7, fontweight='bold', zorder=4,
                    color='white' if base is not None else '#24405e',
                    path_effects=halo_blanco if base is not None else [])
        nx = x1 - (x1 - x0) * 0.05
        ax.annotate('N', (nx, y1 - (y1 - y0) * 0.10), ha='center', fontsize=9,
                    fontweight='bold', zorder=4,
                    color='white' if base is not None else '#24405e',
                    path_effects=halo_blanco if base is not None else [])
        ax.annotate('', xy=(nx, y1 - (y1 - y0) * 0.035),
                    xytext=(nx, y1 - (y1 - y0) * 0.085), zorder=4,
                    arrowprops=dict(arrowstyle='-|>',
                                    color='white' if base is not None
                                    else '#24405e'))
        if base is not None:
            ax.annotate('Base: Esri World Imagery · Límites: comunas '
                        'oficiales (GADM Cayambe)', (x1 - (x1 - x0) * 0.02,
                        y0 + (y1 - y0) * 0.02), ha='right', fontsize=5.5,
                        color='white', zorder=4, path_effects=halo_blanco)

        if con_leyenda:
            filas = sorted(destacadas, key=lambda c: c['n'])
            fig.text(0.72, 0.97, nombre_sector, fontsize=9,
                     fontweight='bold', color='#1e4d8c', va='top')
            for i, c in enumerate(filas):
                fig.text(0.72, 0.93 - i * 0.038, f"{c['n']}. {c['oficial']}",
                         fontsize=7, color='#1a1a1a', va='top')
        else:
            for i, (sec, col) in enumerate(COLOR_SECTOR.items()):
                fig.text(0.03 + i * 0.16, 0.965, '■', color=col, fontsize=10)
                fig.text(0.05 + i * 0.16, 0.965, sec, fontsize=8,
                         color='#1a1a1a')

        ax.set_xlim(x0, x1)
        ax.set_ylim(y0, y1)
        ax.set_aspect('equal')
        ax.axis('off')
        buf = io.BytesIO()
        # JPEG: la foto satelital en PNG pesaría megas por mapa. La web va a
        # 150 dpi; el archivo en disco —el que acaba en el Word y en el anexo
        # PDF— sale a 300 dpi sin submuestreo de color, para que se pueda
        # imprimir y ampliar sin que se deshaga (JAVIKO, 31-ago-2026).
        fig.savefig(buf, format='jpg', dpi=150, pil_kwargs={'quality': 88})
        if archivo:
            fig.savefig(archivo, format='jpg', dpi=300,
                        pil_kwargs={'quality': 96, 'subsampling': 0})
        plt.close(fig)
        return base64.b64encode(buf.getvalue()).decode('ascii')

    # el anexo se arma con Pillow desde los mismos JPEG: PdfPages de
    # matplotlib incrusta el satélite sin comprimir y la salida pesa 12 MB
    # Los mapas en disco: el HTML los lleva embebidos en base64, pero el
    # Markdown (y por tanto el Word) necesita archivos a los que apuntar, y el
    # anexo PDF se arma con esos mismos archivos de alta resolución.
    carpeta = os.path.join(BASE, 'docs', DIR_MAPAS)
    rutas = {}
    if pdf_ruta:
        os.makedirs(carpeta, exist_ok=True)

    def ruta_de(clave):
        if not pdf_ruta:
            return None
        r = os.path.join(carpeta, clave.lower().replace(' ', '-') + '.jpg')
        rutas[clave] = r
        return r

    mapas = {'general': dibujar(None, ruta_de('general'))}
    for sec in ('Sector 1', 'Sector 2', 'Sector 3'):
        mapas[sec] = dibujar(sec, ruta_de(sec))

    if pdf_ruta:
        from PIL import Image
        orden = ('general', 'Sector 1', 'Sector 2', 'Sector 3')
        paginas = [Image.open(rutas[k]) for k in orden]
        paginas[0].save(pdf_ruta, save_all=True, append_images=paginas[1:],
                        resolution=300.0,
                        title='Anexo cartográfico — Informe por Comunidad',
                        author=E.PIE_INSTITUCION)
        mb = os.path.getsize(pdf_ruta) / 1e6
        print(f'✔ {os.path.relpath(pdf_ruta, BASE)} ({mb:.1f} MB, 300 dpi)')
    return mapas


def figura_mapa(b64, pie):
    return ('<div class="evitar-corte" style="margin:14px 0">'
            f'<img src="data:image/jpeg;base64,{b64}" alt="{pie}" '
            'style="width:100%;border:1px solid #dbe3ee;border-radius:7px">'
            f'<p class="sub" style="margin-top:4px">{pie}</p></div>')


def barra_pct(p):
    """Barra de porcentaje con la cifra FUERA de la barra. El componente de la
    casa (informe_estilo.barra) pone la cifra sobre la barra y solo la pasa a
    blanco desde el 88 %: en estos cuadros comparativos los porcentajes de
    60–90 % son la norma y la cifra quedaría sobre el verde, ilegible. Aquí la
    cifra va a la izquierda y la barra al lado, con los mismos colores."""
    return ('<div style="display:flex;align-items:center;gap:6px;min-width:110px">'
            f'<span style="width:44px;text-align:right;font-size:8.5pt;'
            f'color:#24405e;flex:none">{p:,.1f}%</span>'
            f'<div class="barra" style="flex:1">'
            f'<span style="width:{min(p, 100):.0f}%"></span></div></div>')


# ─── Cuadros: una definición, tres salidas (HTML, MD, XLSX) ──────────────────
# Cada cuadro es (titulo, descripcion, columnas, filas, fila_total, pies).
# Las filas ya vienen como texto formateado; el XLSX usa sus propias matrices
# crudas (ver hojas_xlsx) para que el sociólogo pueda recalcular.

class Cuadro:
    def __init__(self, titulo, descripcion, columnas, filas, fila_total, pies,
                 barras=()):
        """`barras`: índices de columna cuyas celdas llegan como float (un
        porcentaje) y se dibujan con la barra del estilo de la casa en el
        HTML; en el Markdown/Word salen como número."""
        self.titulo = titulo
        self.descripcion = descripcion
        self.columnas = columnas
        self.filas = filas
        self.fila_total = fila_total
        self.pies = [p for p in pies if p]
        self.barras = set(barras)

    def _celda_html(self, i, v):
        if i in self.barras and isinstance(v, float):
            return f'<td>{barra_pct(v)}</td>'
        return f'<td{" class=\"n\"" if i else ""}>{v}</td>'

    def _celda_txt(self, i, v):
        if i in self.barras and isinstance(v, float):
            return f'{v:,.1f} %'
        return str(v)

    def html(self):
        h = [f'<h3>{self.titulo}</h3>']
        if self.descripcion:
            h.append(f'<p>{self.descripcion}</p>')
        h.append('<table class="evitar-corte"><tr>' +
                 ''.join(f'<th{" class=\"n\"" if i else ""}>{c}</th>'
                         for i, c in enumerate(self.columnas)) + '</tr>')
        for fila in self.filas:
            h.append('<tr>' + ''.join(self._celda_html(i, v)
                                      for i, v in enumerate(fila)) + '</tr>')
        if self.fila_total:
            h.append('<tr class="dest">' + ''.join(
                self._celda_html(i, v)
                for i, v in enumerate(self.fila_total)) + '</tr>')
        h.append('</table>')
        for p in self.pies:
            h.append(f'<p class="sub" style="margin-top:-8px">{p}</p>')
        return '\n'.join(h)

    def md(self):
        m = [f'### {self.titulo}', '']
        if self.descripcion:
            m += [self.descripcion, '']
        m.append('<!-- tabla-completa -->')
        m.append('| ' + ' | '.join(self.columnas) + ' |')
        m.append('|' + '|'.join('---' for _ in self.columnas) + '|')
        for fila in self.filas:
            m.append('| ' + ' | '.join(self._celda_txt(i, v)
                                       for i, v in enumerate(fila)) + ' |')
        if self.fila_total:
            m.append('| ' + ' | '.join(f'**{self._celda_txt(i, v)}**'
                                       for i, v in enumerate(self.fila_total)) + ' |')
        m.append('')
        for p in self.pies:
            m += [f'*{p}*', '']
        return '\n'.join(m)


def et(c):
    """Etiqueta de comunidad: número y nombre del listado oficial del
    consorcio (mismo formato que usa la aplicación web)."""
    return f"{c['n']}. {c['oficial']}"


def cuadros_sector(coms, sec_sup, caudal):
    """Los cuadros comparativos de un sector, en el orden de los bloques:
    A tierra y producción · B el agua · C las personas."""
    C = []
    nota_alpaka = any(c['key'] == 'ALPAKA' for c in coms)

    # A1 — padrón (misma terminología que las cards del Dashboard, 31-ago)
    C.append(Cuadro(
        'Fichas principales, adicionales y totales',
        'Cada titular entrevistado llenó una ficha principal. Sobre un mismo '
        'predio puede haber varias fichas principales —los predios '
        'familiares, donde cada heredero llena la suya— y una ficha '
        'principal puede tener fichas adicionales: los otros predios del '
        'mismo titular, levantados aparte. Las fichas totales suman las dos. '
        'Los predios catastrales son los polígonos distintos del catastro '
        'municipal que todas esas fichas ocupan: por esos dos efectos no '
        'coinciden ni con las principales ni con las totales, y no tienen '
        'por qué hacerlo.',
        ['Comunidad', 'Fichas principales', 'Fichas adicionales',
         'Fichas totales', 'Predios catastrales'],
        [[et(c), f0(c['regantes']), f0(c['adicionales']), f0(c['fichas']),
          f0(c['predios_catastrales'])] for c in coms],
        ['Total del sector', f0(sum(c['regantes'] for c in coms)),
         f0(sum(c['adicionales'] for c in coms)),
         f0(sum(c['fichas'] for c in coms)),
         f0(sum(c['predios_catastrales'] for c in coms))],
        ['Fuente: superficie_por_comunidad.json (fuente única del padrón).']))

    # A2 — superficie declarada, con la referencia catastral al pie
    cat_sector = sum(c['sup_catastral'] for c in coms)
    C.append(Cuadro(
        'Superficie declarada: con riego, sin riego y total',
        'La superficie que los comuneros declararon en la entrevista, en '
        'hectáreas. Es el dato del universo declarado: describe lo que las '
        'familias dicen tener y regar.',
        ['Comunidad', 'Con riego (ha)', 'Sin riego (ha)', 'Total declarado (ha)',
         '% con riego', 'Ref. catastral (ha)'],
        [[et(c), f2(c['sup_riego_decl']), f2(c['sup_sin_riego_decl']),
          f2(c['sup_declarada']), pct(c['sup_riego_decl'], c['sup_declarada']),
          f2(c['sup_catastral'])] for c in coms],
        ['Total del sector', f2(sum(c['sup_riego_decl'] for c in coms)),
         f2(sum(c['sup_sin_riego_decl'] for c in coms)),
         f2(sum(c['sup_declarada'] for c in coms)),
         pct(sum(c['sup_riego_decl'] for c in coms),
             sum(c['sup_declarada'] for c in coms)), f2(cat_sector)],
        [NOTA_UNIV_TODAS,
         'La referencia catastral mide los polígonos del catastro municipal '
         '(cada predio una sola vez). Son dos mediciones distintas del mismo '
         'territorio y NO se suman entre sí.',
         NOTA_P001 if nota_alpaka else None],
        barras=[4]))

    # A3 — tecnificación
    C.append(Cuadro(
        'Riego tecnificado frente a riego por inundación',
        'Cómo se aplica el agua en el predio, según el porcentaje que declaró '
        'cada titular entrevistado. Tecnificado agrupa aspersión y goteo; inundación es el '
        'riego por gravedad.',
        ['Comunidad', 'Inundación (gravedad) %', 'Aspersión %', 'Goteo %',
         'Tecnificado %', 'Fichas con dato'],
        [[et(c), f1(c['met_gravedad']), f1(c['met_aspersion']),
          f1(c['met_goteo']),
          (c['met_tecnificado'] + 0.0 if c['met_n'] >= MIN_RESPUESTAS
           else f1(c['met_tecnificado']) + ' %'), f0(c['met_n'])]
         for c in coms],
        None,
        [NOTA_UNIV_PRI + ' Promedio simple de los porcentajes declarados por '
         'ficha con al menos un método.', NOTA_POCAS],
        barras=[4]))

    # A4 — cultivos top 5
    def celda_cultivo(c, i):
        if i < len(c['cultivos_top']):
            nombre, ha = c['cultivos_top'][i]
            return f'{nombre} ({f1(ha)} ha)'
        return '—'
    C.append(Cuadro(
        'Los cinco cultivos relevantes',
        'Los cinco cultivos con más superficie declarada en cada comunidad; '
        'el resto se agrupa como «otros».',
        ['Comunidad', '1.º', '2.º', '3.º', '4.º', '5.º', 'Otros (ha)',
         'Total cultivado (ha)'],
        [[et(c)] + [celda_cultivo(c, i) for i in range(5)] +
         [f1(c['cultivos_otros_ha']), f1(c['cultivos_total_ha'])] for c in coms],
        ['Total del sector', '', '', '', '', '',
         f1(sum(c['cultivos_otros_ha'] for c in coms)),
         f1(sum(c['cultivos_total_ha'] for c in coms))],
        [NOTA_UNIV_TODAS,
         ('ALPAKA no registra cultivos: ' + NOTA_ALPAKA_LOTEO)
         if nota_alpaka else None]))

    # A5 — uso pecuario: pasto mejorado × inventario pecuario
    granja = sum(c['aves_granja_excluidas'] for c in coms)
    C.append(Cuadro(
        'Superficie de uso pecuario e inventario de animales',
        'La superficie de uso pecuario se deriva del cultivo «Pasto mejorado» '
        'declarado en cada comunidad, cruzado con el inventario pecuario de '
        'las mismas fichas. La carga bovina resulta de dividir los bovinos '
        'entre las hectáreas de pasto mejorado.',
        ['Comunidad', 'Pasto mejorado (ha)', 'Bovinos', 'Ovinos y caprinos',
         'Porcinos', 'Cuyes y conejos', 'Aves', 'Carga bovina (cab/ha)'],
        [[et(c), f1(c['pasto_mejorado_ha']), f0(c['bovinos']),
          f0(c['pecuario'].get('Ovinos y caprinos', 0)),
          f0(c['pecuario'].get('Porcinos', 0)),
          f0(c['pecuario'].get('Cuyes y conejos', 0)),
          f0(c['pecuario'].get('Aves', 0)),
          f1(c['carga_bovina']) if c['carga_bovina'] is not None else '—']
         for c in coms],
        ['Total del sector', f1(sum(c['pasto_mejorado_ha'] for c in coms)),
         f0(sum(c['bovinos'] for c in coms)),
         f0(sum(c['pecuario'].get('Ovinos y caprinos', 0) for c in coms)),
         f0(sum(c['pecuario'].get('Porcinos', 0) for c in coms)),
         f0(sum(c['pecuario'].get('Cuyes y conejos', 0) for c in coms)),
         f0(sum(c['pecuario'].get('Aves', 0) for c in coms)), ''],
        [NOTA_UNIV_TODAS,
         (f'Se excluyen {f0(granja)} aves de una granja avícola de Asociación '
          'Rosalía (registros de 10.000 aves por titular sobre el mismo '
          'predio, en verificación con terceros).') if granja else None]))

    # A6 — tenencia
    C.append(Cuadro(
        'Tenencia del predio',
        'La forma en que cada titular entrevistado dice tener su predio.',
        ['Comunidad', 'Escritura / título %', 'Posesión sin título %',
         'Otras formas %', 'Respuestas'],
        [[et(c), pct_o_conteo(c['ten_escritura'], c['ten_n']),
          f1(pct(c['ten_posesion'], c['ten_n'])),
          f1(pct(c['ten_otras'], c['ten_n'])), f0(c['ten_n'])] for c in coms],
        None,
        [NOTA_UNIV_PRI, NOTA_POCAS, NOTA_ALPAKA_LOTEO if nota_alpaka else None],
        barras=[1]))

    # B1 — el agua: caudal, turnos, días
    q_sector = sum(c['caudal_ls'] or 0 for c in coms if not c['caudal_heredado_de'])
    def celda_caudal(c):
        if c['caudal_ls'] is None:
            return '—'
        v = f1(c['caudal_ls'])
        if c['caudal_heredado_de']:
            return f'{v} *'
        return v
    C.append(Cuadro(
        'El agua que recibe cada comunidad: caudal, turnos y días de riego',
        'Los litros por segundo que recibe la comunidad (calculados por la '
        'moda de sus fichas, nunca sumando ficha a ficha), cuánto dura el '
        'turno de cada titular y cuántos días a la semana riega.',
        ['Comunidad', 'Caudal (l/s)', 'Cómo se obtuvo', 'Turno (horas, mediana)',
         'Días de riego por semana (mediana)'],
        [[et(c), celda_caudal(c), c['caudal_origen'].capitalize() or '—',
          f1(c['horas_turno_med']) if c['horas_turno_med'] is not None else '—',
          f1(c['dias_riego_med']) if c['dias_riego_med'] is not None else '—']
         for c in coms],
        ['Total del sector', f1(q_sector), '', '', ''],
        ['Fuente del caudal: caudal_por_comunidad.json (fuente única; regla '
         'del proyecto: el caudal no se suma ficha a ficha).',
         '* Caudal heredado de otra comunidad (comparten la misma llave); se '
         'muestra pero no se suma al total del sector.',
         'Turnos y días se agregan de las fichas principales (mediana).',
         ('Valores de más de 7 días provienen de fichas que declararon el '
          'dato en otra escala (por ejemplo, por mes); se reportan tal como '
          'fueron declarados.'
          if any((c['dias_riego_med'] or 0) > 7 for c in coms) else None),
         NOTA_P001 if nota_alpaka else None]))

    # B2 — frecuencia
    cats_frec = ['Semanal', 'Quincenal', 'Mensual', 'Permanente', 'No tiene riego']
    C.append(Cuadro(
        'Frecuencia de riego',
        'Cada cuánto le llega el agua al titular, en porcentaje de las '
        'respuestas de cada comunidad.',
        ['Comunidad'] + [f'{f} %' for f in cats_frec] + ['Respuestas'],
        [[et(c)] + [f1(pct(c['frec'].get(f, 0), c['frec_n'])) for f in cats_frec] +
         [f0(c['frec_n'])] for c in coms],
        None,
        [NOTA_UNIV_PRI]))

    # B3 — tarifas (mediana por modalidad, nunca mezclando USD/mes con USD/año)
    def fila_tarifa(c):
        if c['key'] == 'ALPAKA':
            return [et(c), '—', '—', '—', '—']
        return [et(c),
                f2(c['tarifa_mensual_med']) if c['tarifa_mensual_med'] is not None else '—',
                f0(c['tarifa_mensual_n']),
                f2(c['tarifa_anual_med']) if c['tarifa_anual_med'] is not None else '—',
                f0(c['tarifa_anual_n'])]
    C.append(Cuadro(
        'Tarifas por el agua',
        'Lo que el titular entrevistado dice pagar por el riego, según su modalidad (fija '
        'mensual o fija anual). Se reporta la mediana, no el promedio: unas '
        'pocas fichas con valores extremos desplazarían el promedio hacia una '
        'cifra que no representa a nadie.',
        ['Comunidad', 'Mediana mensual (USD)', 'Pagan mensual',
         'Mediana anual (USD)', 'Pagan anual'],
        [fila_tarifa(c) for c in coms],
        None,
        [NOTA_UNIV_PRI,
         ('ALPAKA declara tarifas de 672 y 308 USD «mensuales» en 491 fichas, '
          'muy por encima de cualquier tarifa del sistema: no es la tarifa de '
          'riego sino otro concepto del fraccionamiento, y se excluye del '
          'análisis económico como anomalía en verificación.')
         if nota_alpaka else None]))

    # B4 — reservorios
    C.append(Cuadro(
        'Reservorios',
        'Si el titular entrevistado cuenta con un reservorio de agua y de qué tipo.',
        ['Comunidad', 'Comunitario %', 'Privado %', 'No tiene %', 'Respuestas'],
        [[et(c), pct_o_conteo(c['res_comunitario'], c['res_n']),
          f1(pct(c['res_privado'], c['res_n'])),
          f1(pct(c['res_no'], c['res_n'])), f0(c['res_n'])] for c in coms],
        None,
        [NOTA_UNIV_PRI, NOTA_POCAS],
        barras=[1]))

    # C1 — instrucción
    cats_ins = ['Ninguno', 'Alfabetizado', 'Primaria', 'Secundaria', 'Superior']
    C.append(Cuadro(
        'Nivel de instrucción del titular',
        'El nivel de estudios que declaró cada titular entrevistado, en '
        'porcentaje de las respuestas de su comunidad.',
        ['Comunidad'] + [f'{i} %' for i in cats_ins] + ['Respuestas'],
        [[et(c)] + [f1(pct(c['instruccion'].get(i, 0), c['ins_n']))
                    for i in cats_ins] + [f0(c['ins_n'])] for c in coms],
        None,
        [NOTA_UNIV_PRI, NOTA_POCAS,
         NOTA_ALPAKA_LOTEO if nota_alpaka else None]))

    # C2 — hijos
    C.append(Cuadro(
        'Hijos por familia',
        'Cuántos hijos e hijas declaró cada titular, en promedio por '
        'comunidad.',
        ['Comunidad', 'Hijos por familia (promedio)', 'Respuestas'],
        [[et(c), f1(c['hijos_prom']) if c['hijos_prom'] is not None else '—',
          f0(c['hijos_n'])] for c in coms],
        None,
        [NOTA_UNIV_PRI]))

    # C3 — servicios básicos (denominador = viviendas, no todas las fichas)
    n_viv = sum(c['con_vivienda'] for c in coms)
    n_pri_sec = sum(c['n_pri'] for c in coms)
    C.append(Cuadro(
        'Vivienda y servicios básicos',
        'Cuántas fichas principales declaran una vivienda en el predio y qué '
        'servicios tiene esa vivienda. Los porcentajes de agua y energía se '
        'calculan sobre las viviendas, no sobre todas las fichas: un predio '
        'sin casa no tiene agua ni luz porque no hay vivienda, no porque le '
        'falte el servicio. El teléfono es dato de contacto de la persona y '
        'se calcula sobre todas las fichas principales.',
        ['Comunidad', 'Con vivienda', '% con vivienda', 'Agua de consumo',
         '% de las viviendas', 'Energía eléctrica', '% de las viviendas',
         'Teléfono', '% de las fichas', 'Fichas principales'],
        [[et(c), f0(c['con_vivienda']), pct(c['con_vivienda'], c['n_pri']),
          f0(c['agua_consumo']), pct(c['agua_consumo'], c['con_vivienda']),
          f0(c['energia']), pct(c['energia'], c['con_vivienda']),
          f0(c['telefono']), pct(c['telefono'], c['n_pri']),
          f0(c['n_pri'])] for c in coms],
        ['Total del sector', f0(n_viv), pct(n_viv, n_pri_sec),
         f0(sum(c['agua_consumo'] for c in coms)),
         pct(sum(c['agua_consumo'] for c in coms), n_viv),
         f0(sum(c['energia'] for c in coms)),
         pct(sum(c['energia'] for c in coms), n_viv),
         f0(sum(c['telefono'] for c in coms)),
         pct(sum(c['telefono'] for c in coms), n_pri_sec), f0(n_pri_sec)],
        [NOTA_UNIV_PRI,
         'La vivienda se identifica por el material de construcción '
         'declarado. Criterio del cliente (9 de agosto de 2026): sin material '
         'de construcción no hay vivienda, y entonces agua y energía vacías '
         'son la respuesta correcta, no un dato faltante.'],
        barras=[2, 4, 6, 8]))

    # C3b — material de la vivienda (del capítulo de servicios, integrado)
    MATS = ['BLOQUE', 'HORMIGÓN ARMADO', 'TAPIA', 'LADRILLO', 'ADOBE',
            'MIXTA', 'MADERA']
    def otros_mat(c):
        return sum(v for k, v in c['materiales'].items() if k not in MATS)
    C.append(Cuadro(
        'Material de la vivienda',
        'De qué está construida la vivienda declarada, en número de fichas '
        'por material. Describe el hábitat de las familias del sistema.',
        ['Comunidad'] + [m.capitalize() for m in MATS] + ['Otros',
                                                          'Con vivienda'],
        [[et(c)] + [f0(c['materiales'].get(m, 0)) for m in MATS] +
         [f0(otros_mat(c)), f0(c['con_vivienda'])] for c in coms],
        ['Total del sector'] +
        [f0(sum(c['materiales'].get(m, 0) for c in coms)) for m in MATS] +
        [f0(sum(otros_mat(c) for c in coms)),
         f0(sum(c['con_vivienda'] for c in coms))],
        [NOTA_UNIV_PRI,
         'Solo las fichas que declaran vivienda; los predios sin '
         'construcción no aparecen en este cuadro.']))

    # C3c — altitud de los predios (del capítulo de servicios, integrado)
    C.append(Cuadro(
        'Altitud de los predios',
        'A qué altura sobre el nivel del mar están los predios de cada '
        'comunidad. La altitud condiciona qué se puede sembrar y cómo llega '
        'el agua, y es el dato de hábitat que ordena el sistema de arriba '
        'hacia abajo.',
        ['Comunidad', 'Mínima (m s. n. m.)', 'Mediana (m s. n. m.)',
         'Máxima (m s. n. m.)', 'Fichas con dato'],
        [[et(c),
          f0(c['cota_min']) if c['cota_min'] is not None else '—',
          f0(c['cota_med']) if c['cota_med'] is not None else '—',
          f0(c['cota_max']) if c['cota_max'] is not None else '—',
          f0(c['cota_n'])] for c in coms],
        ['Total del sector',
         f0(min(c['cota_min'] for c in coms if c['cota_min'] is not None)),
         '', f0(max(c['cota_max'] for c in coms if c['cota_max'] is not None)),
         f0(sum(c['cota_n'] for c in coms))],
        [NOTA_UNIV_PRI]))

    # C4 — presa
    C.append(Cuadro(
        'Conocimiento de la presa',
        'Si el titular entrevistado sabe de la construcción de la presa del proyecto.',
        ['Comunidad', 'Conoce la presa %', 'Respuestas'],
        [[et(c), pct_o_conteo(c['presa_si'], c['presa_n']), f0(c['presa_n'])]
         for c in coms],
        None,
        [NOTA_UNIV_PRI, NOTA_POCAS],
        barras=[1]))

    # C5 — capacitación
    def celda_temas(c):
        if not c['temas_top']:
            return '—'
        return '; '.join(f'{t} ({f0(n)})' for t, n in c['temas_top'])
    C.append(Cuadro(
        'Capacitación recibida y demandada',
        'Si el titular entrevistado recibió capacitación, si le gustaría recibirla, y los '
        'temas que más pide (agrupados desde las respuestas libres de la '
        'entrevista).',
        ['Comunidad', 'Recibió %', 'Le gustaría %', 'Temas más pedidos'],
        [[et(c), pct_o_conteo(c['cap_si'], c['cap_n']),
          pct_o_conteo(c['des_si'], c['des_n']), celda_temas(c)] for c in coms],
        None,
        [NOTA_UNIV_PRI, NOTA_POCAS],
        barras=[1, 2]))

    return C


# ─── Narrativa por sector (generada desde los mismos agregados) ──────────────

def narrativa_sector(nombre, coms, sec_sup):
    n_com = len(coms)
    fichas = sum(c['fichas'] for c in coms)
    regantes = sum(c['regantes'] for c in coms)
    decl = sum(c['sup_declarada'] for c in coms)
    riego = sum(c['sup_riego_decl'] for c in coms)
    q = sum(c['caudal_ls'] or 0 for c in coms if not c['caudal_heredado_de'])

    ha_cult = defaultdict(float)
    for c in coms:
        for nombre_c, ha in c['cultivos_top']:
            ha_cult[nombre_c] += ha
        ha_cult['(otros)'] += c['cultivos_otros_ha']
    top = sorted(((k, v) for k, v in ha_cult.items() if k != '(otros)'),
                 key=lambda kv: -kv[1])[:3]

    tec = [c for c in coms if c['met_n']]
    tec_prom = st.mean(c['met_tecnificado'] for c in tec) if tec else 0
    grav_prom = st.mean(c['met_gravedad'] for c in tec) if tec else 0
    mayor_tec = max(tec, key=lambda c: c['met_tecnificado']) if tec else None

    ten_n = sum(c['ten_n'] for c in coms)
    ten_esc = sum(c['ten_escritura'] for c in coms)
    ins_n = sum(c['ins_n'] for c in coms)
    ins_prim = sum(c['instruccion'].get('Primaria', 0) for c in coms)
    ins_ning = sum(c['instruccion'].get('Ninguno', 0) for c in coms)
    presa_n = sum(c['presa_n'] for c in coms)
    presa_si = sum(c['presa_si'] for c in coms)
    des_n = sum(c['des_n'] for c in coms)
    des_si = sum(c['des_si'] for c in coms)

    mayor = max(coms, key=lambda c: c['sup_declarada'])
    menor = min((c for c in coms if c['fichas']), key=lambda c: c['fichas'],
                default=None)

    p = []
    p.append(
        f'El {nombre} agrupa {n_com} comunidades, con {f0(fichas)} fichas '
        f'catastrales: {f0(regantes)} principales y '
        f'{f0(fichas - regantes)} adicionales. En conjunto sus comuneros '
        f'declaran '
        f'{f2(decl)} hectáreas, de las cuales {f2(riego)} '
        f'({f1(pct(riego, decl))} %) se riegan. La comunidad de mayor '
        f'superficie declarada del sector es {mayor["oficial"]} '
        f'({f2(mayor["sup_declarada"])} ha)' +
        (f' y la de menor número de fichas es {menor["oficial"]} '
         f'({f0(menor["fichas"])} ficha{"s" if menor["fichas"] != 1 else ""}).'
         if menor else '.'))
    # la frase del método se deriva del dato: cada sector puede ser
    # mayoritariamente tecnificado o mayoritariamente de inundación
    if grav_prom >= tec_prom:
        frase_met = (f'El riego del sector es mayoritariamente por '
                     f'inundación: en promedio, el {f1(grav_prom)} % del agua '
                     f'se aplica por gravedad y el {f1(tec_prom)} % con '
                     f'métodos tecnificados (aspersión y goteo)')
    else:
        frase_met = (f'El riego del sector es mayoritariamente tecnificado: '
                     f'en promedio, el {f1(tec_prom)} % del agua se aplica '
                     f'por aspersión o goteo y el {f1(grav_prom)} % por '
                     f'gravedad (inundación)')
    p.append(
        f'Por sus llaves entra un caudal de {f1(q)} litros por segundo '
        f'(sumando la moda de cada comunidad, nunca ficha a ficha). ' +
        frase_met +
        (f'; la comunidad más tecnificada del sector es '
         f'{mayor_tec["oficial"]}, con {f1(mayor_tec["met_tecnificado"])} % '
         f'de sus porcentajes declarados.' if mayor_tec else '.'))
    if top:
        frase_top = ', '.join(f'{k} ({f1(v)} ha)' for k, v in top)
        p.append(
            f'En la tierra dominan {frase_top}, sobre un total cultivado de '
            f'{f1(sum(ha_cult.values()))} hectáreas declaradas en el sector. '
            f'El pasto mejorado, cruzado con el inventario pecuario, marca la '
            f'superficie de uso pecuario de cada comunidad: '
            f'{f1(sum(c["pasto_mejorado_ha"] for c in coms))} ha que '
            f'sostienen {f0(sum(c["bovinos"] for c in coms))} bovinos.')
    p.append(
        f'En el perfil de las personas, {f1(pct(ten_esc, ten_n))} % de los '
        f'titulares con respuesta tiene escritura o título de propiedad. '
        f'{f1(pct(ins_prim, ins_n))} % alcanzó la primaria y '
        f'{f1(pct(ins_ning, ins_n))} % no tiene instrucción formal. El '
        f'{f1(pct(presa_si, presa_n))} % conoce la presa del proyecto y el '
        f'{f1(pct(des_si, des_n))} % quisiera recibir capacitación.')
    return p


# ─── Documento HTML y MD ─────────────────────────────────────────────────────

def fecha_es(d):
    return f'{d.day} de {MESES[d.month - 1]} de {d.year}'


def introduccion_general(sup, caudal, corte_txt, hoy_txt):
    tot = sup['total']
    q = caudal['totales']
    parrafos = [
        'Este informe presenta, comunidad por comunidad, los resultados de la '
        'ficha catastral del padrón de usuarios del sistema de riego '
        'Guanguilquí–Porotog. Está organizado por sector de investigación: '
        'cada sector abre con una lectura general de sus resultados y '
        'continúa con cuadros comparativos donde cada pregunta de la ficha '
        'se responde con una fila por comunidad.',
        f'El universo del informe es LO DECLARADO POR LOS COMUNEROS en la '
        f'entrevista: {f0(tot["fichas"])} fichas catastrales en 50 '
        f'comunidades —{f0(tot["regantes"])} principales y '
        f'{f0(tot["fichas"] - tot["regantes"])} adicionales— que declaran '
        f'{f2(tot["superficie_declarada_ha"])} hectáreas. Como referencia, '
        f'el catastro municipal mide {f2(tot["superficie_catastral_ha"])} '
        f'hectáreas en los mismos predios: son dos mediciones distintas del '
        f'mismo territorio (lo que las familias declaran y lo que los '
        f'polígonos miden) y en ningún cuadro se suman entre sí.',
        f'El caudal del sistema es de {f2(q["caudal_sistema_ls"])} litros '
        f'por segundo: {f2(q["caudal_comunidades_ls"])} l/s que reciben las '
        f'comunidades (calculados por la moda de las fichas de cada una, '
        f'nunca sumando ficha a ficha) más {f2(q["caudal_individual_ls"])} '
        f'l/s de {f0(q["fichas_individuales"])} tomas individuales.',
        # Glosario fijado con JAVIKO el 31-ago-2026: en cifras, columnas y
        # títulos se cuentan fichas (nunca «regantes»); en la prosa, cuando
        # el sujeto es la persona que responde, se dice «titular
        # entrevistado» — una ficha no tiene hijos ni conoce la presa.
        'CÓMO SE CUENTA EN ESTE INFORME. Cada titular entrevistado llenó una '
        '<b>ficha principal</b>. Sobre un mismo predio puede haber varias '
        'fichas principales: en los predios familiares cada heredero llena '
        'la suya. Una ficha principal puede tener <b>fichas adicionales</b>, '
        'que son los otros predios del mismo titular, levantados aparte; una '
        'ficha adicional es un predio más, no una persona más. Las <b>fichas '
        'totales</b> suman las dos. Los <b>predios catastrales</b> son los '
        'polígonos distintos del catastro municipal que esas fichas ocupan y '
        'no coinciden con ninguno de los dos conteos anteriores, por los '
        'mismos dos efectos.',
        'Dos universos conviven en los cuadros y cada uno nombra el suyo al '
        'pie: los datos de tierra y producción salen de TODAS las fichas '
        '(cada ficha es un predio); los datos de las personas (tenencia, '
        'instrucción, hijos, capacitación) salen SOLO de las fichas '
        'principales, porque las adicionales pertenecen al mismo titular y '
        'duplicarían su respuesta. El respaldo ficha por ficha existe en el '
        'Excel catastral del proyecto; aquí no se publican listados '
        'nominales.',
        # el corte tiene dos fechas y conviene decirlo: la última ficha de
        # campo es del 5-ago, pero las cifras llevan la depuración posterior
        f'Las cifras corresponden al padrón al {hoy_txt}, la misma fecha '
        'de referencia de los demás documentos entregados al consorcio. El '
        'levantamiento de campo está cerrado: todas las secciones de la '
        'ficha se completaron y desde entonces no se ha incorporado ninguna '
        'ficha nueva al padrón.',
    ]
    return parrafos


def construir_documento(comunidades, sup, caudal, corte_txt, fichas,
                        solo_sector=None):
    sectores = ['Sector 1', 'Sector 2', 'Sector 3']
    if solo_sector:
        sectores = [f'Sector {solo_sector}']

    titulo = 'Informe por Comunidad'
    subtitulo = ('Resultados de la ficha catastral por comunidad y sector de '
                 'investigación · Universo declarado por los comuneros')
    if solo_sector:
        subtitulo += f' · MUESTRA: Sector {solo_sector}'

    hoy_txt = FECHA_CORTE
    corte_linea = f'Datos al {FECHA_CORTE}'

    H = [E.cabecera(titulo, subtitulo)]
    M = [f'# {titulo}', '', f'*{subtitulo}*', '',
         f'*{corte_linea}.*', '']

    total = sup['total']
    H.append(E.kpis([
        (f0(total['fichas']), 'fichas catastrales'),
        (f0(total['regantes']), 'fichas principales'),
        (f2(total['superficie_declarada_ha']) + ' ha', 'superficie declarada'),
        (f2(caudal['totales']['caudal_sistema_ls']) + ' l/s',
         'caudal del sistema'),
    ]))

    mapas = generar_mapas(
        comunidades,
        pdf_ruta=None if solo_sector else
        os.path.join(BASE, 'docs', 'ANEXO-MAPAS-sociologo.pdf'))

    H.append('<h2>Presentación</h2>')
    M += ['## Presentación', '']
    for p in introduccion_general(sup, caudal, corte_txt, hoy_txt):
        H.append(f'<p>{p}</p>')
        # el glosario lleva <b> para el HTML; en Markdown va como **negrita**
        M += [p.replace('<b>', '**').replace('</b>', '**'), '']
    pie_general = ('El área de estudio sobre imagen satelital (Esri World '
                   'Imagery): límites oficiales de comunas entregados por el '
                   'GADM Cayambe, recortados al sistema y coloreados por '
                   'sector de investigación (asignación por cruce espacial '
                   'con las organizaciones de riego). La numeración 1–50 es '
                   'la del listado oficial de organizaciones, la misma de '
                   'los cuadros; mapa generado junto con el documento.')
    H.append(figura_mapa(mapas['general'], pie_general))
    M += [f'![{pie_general}]({DIR_MAPAS}/general.jpg)', '']

    for sec in sectores:
        coms = [c for c in comunidades if c['sector'] == sec]
        sec_sup = sup['sectores'].get(sec, {})
        H.append(f'<h2>{sec} — {len(coms)} comunidades</h2>')
        M += [f'## {sec} — {len(coms)} comunidades', '']
        q_sec = sum(c['caudal_ls'] or 0 for c in coms
                    if not c['caudal_heredado_de'])
        H.append(E.kpis([
            (f0(sum(c['fichas'] for c in coms)), 'fichas catastrales'),
            (f0(sum(c['regantes'] for c in coms)), 'fichas principales'),
            (f2(sum(c['sup_declarada'] for c in coms)) + ' ha',
             'superficie declarada'),
            (f1(q_sec) + ' l/s', 'caudal de sus comunidades'),
        ]))
        pie_sec = (f'Las comunas oficiales del {sec} (límites del GADM '
                   'Cayambe, asignadas al sector por cruce espacial). La '
                   'numeración es la del listado oficial de organizaciones '
                   'del consorcio, la misma de los cuadros; el resto del '
                   'sistema queda en gris.')
        H.append(figura_mapa(mapas[sec], pie_sec))
        M += [f'![{pie_sec}]({DIR_MAPAS}/{sec.lower().replace(" ", "-")}.jpg)', '']
        for p in narrativa_sector(sec, coms, sec_sup):
            H.append(f'<p>{p}</p>')
            M += [p, '']
        for cu in cuadros_sector(coms, sec_sup, caudal):
            H.append(cu.html())
            M += [cu.md(), '']

    H.append(E.pie(corte_linea.lower()))
    M += ['---', '',
          f'*{E.PIE_INSTITUCION} · {corte_linea}. Documento generado '
          f'por scripts/generar_informe_sociologo.py; especificaciones del '
          f'cliente, 24-ago-2026.*', '']

    html = E.documento(f'{titulo} — Padrón Guanguilquí–Porotog', '\n'.join(H))
    md = '\n'.join(M)
    return html, md


# ─── Excel de matrices crudas (una hoja por bloque temático) ─────────────────

def escribir_xlsx(comunidades, ruta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    cab_font = Font(bold=True, color='FFFFFF')
    cab_fill = PatternFill('solid', fgColor='1E4D8C')

    def hoja(nombre, columnas, filas):
        ws = wb.create_sheet(nombre[:31])
        ws.append(columnas)
        for cel in ws[1]:
            cel.font = cab_font
            cel.fill = cab_fill
            cel.alignment = Alignment(vertical='center', wrap_text=True)
        for fila in filas:
            ws.append(fila)
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            ancho = max(len(str(c.value or '')) for c in col[:40])
            ws.column_dimensions[col[0].column_letter].width = min(max(ancho + 2, 10), 42)

    base = lambda c: [c['n'], c['oficial'], c['sector']]
    cols_base = ['N.º', 'Comunidad (listado oficial)', 'Sector']

    hoja('Padron', cols_base + ['Fichas principales',
                                'Fichas adicionales', 'Fichas totales',
                                'Predios catastrales'],
         [base(c) + [c['regantes'], c['adicionales'], c['fichas'],
                     c['predios_catastrales']] for c in comunidades])

    hoja('Superficie declarada',
         cols_base + ['Con riego (ha)', 'Sin riego (ha)', 'Total declarado (ha)',
                      'Referencia catastral (ha)'],
         [base(c) + [round(c['sup_riego_decl'], 2),
                     round(c['sup_sin_riego_decl'], 2),
                     round(c['sup_declarada'], 2),
                     round(c['sup_catastral'], 2)] for c in comunidades])

    hoja('Metodos de riego',
         cols_base + ['Gravedad %', 'Aspersion %', 'Goteo %', 'Tecnificado %',
                      'Fichas con dato'],
         [base(c) + [round(c['met_gravedad'], 1), round(c['met_aspersion'], 1),
                     round(c['met_goteo'], 1), round(c['met_tecnificado'], 1),
                     c['met_n']] for c in comunidades])

    filas_cult = []
    for c in comunidades:
        f = base(c)
        for i in range(5):
            if i < len(c['cultivos_top']):
                f += [c['cultivos_top'][i][0], round(c['cultivos_top'][i][1], 2)]
            else:
                f += [None, None]
        f += [round(c['cultivos_otros_ha'], 2), round(c['cultivos_total_ha'], 2)]
        filas_cult.append(f)
    hoja('Cultivos top 5',
         cols_base + sum(([f'Cultivo {i}', f'ha {i}'] for i in range(1, 6)), []) +
         ['Otros (ha)', 'Total (ha)'], filas_cult)

    hoja('Pecuario',
         cols_base + ['Pasto mejorado (ha)', 'Bovinos', 'Ovinos y caprinos',
                      'Porcinos', 'Cuyes y conejos', 'Aves',
                      'Carga bovina (cab/ha)', 'Aves granja excluidas'],
         [base(c) + [round(c['pasto_mejorado_ha'], 2), c['bovinos'],
                     c['pecuario'].get('Ovinos y caprinos', 0),
                     c['pecuario'].get('Porcinos', 0),
                     c['pecuario'].get('Cuyes y conejos', 0),
                     c['pecuario'].get('Aves', 0),
                     round(c['carga_bovina'], 2) if c['carga_bovina'] is not None else None,
                     c['aves_granja_excluidas']] for c in comunidades])

    hoja('Tenencia',
         cols_base + ['Escritura/titulo', 'Posesion sin titulo', 'Otras',
                      'Respuestas'],
         [base(c) + [c['ten_escritura'], c['ten_posesion'], c['ten_otras'],
                     c['ten_n']] for c in comunidades])

    hoja('Agua caudal turnos',
         cols_base + ['Caudal (l/s)', 'Origen', 'Heredado de',
                      'Turno (h, mediana)', 'Dias riego/semana (mediana)',
                      'Frecuencia predominante'],
         [base(c) + [c['caudal_ls'], c['caudal_origen'],
                     c['caudal_heredado_de'] or '',
                     c['horas_turno_med'], c['dias_riego_med'],
                     c['frec_predominante'] or ''] for c in comunidades])

    hoja('Tarifas y reservorios',
         cols_base + ['Mediana mensual (USD)', 'Pagan mensual',
                      'Mediana anual (USD)', 'Pagan anual',
                      'Reservorio comunitario', 'Privado', 'No tiene',
                      'Respuestas reservorio'],
         [base(c) + [c['tarifa_mensual_med'], c['tarifa_mensual_n'],
                     c['tarifa_anual_med'], c['tarifa_anual_n'],
                     c['res_comunitario'], c['res_privado'], c['res_no'],
                     c['res_n']] for c in comunidades])

    cats_ins = ['Ninguno', 'Alfabetizado', 'Primaria', 'Secundaria', 'Superior']
    hoja('Perfil',
         cols_base + cats_ins + ['Respuestas instruccion',
                                 'Hijos por familia (prom)', 'Respuestas hijos'],
         [base(c) + [c['instruccion'].get(i, 0) for i in cats_ins] +
          [c['ins_n'], round(c['hijos_prom'], 2) if c['hijos_prom'] is not None else None,
           c['hijos_n']] for c in comunidades])

    hoja('Vivienda y servicios',
         cols_base + ['Con vivienda', 'Agua de consumo', 'Energia electrica',
                      'Telefono', 'Fichas principales'],
         [base(c) + [c['con_vivienda'], c['agua_consumo'], c['energia'],
                     c['telefono'], c['n_pri']] for c in comunidades])

    hoja('Material de vivienda',
         cols_base + ['Bloque', 'Hormigon armado', 'Tapia', 'Ladrillo',
                      'Adobe', 'Mixta', 'Madera', 'Con vivienda'],
         [base(c) + [c['materiales'].get(m, 0) for m in
                     ('BLOQUE', 'HORMIGÓN ARMADO', 'TAPIA', 'LADRILLO',
                      'ADOBE', 'MIXTA', 'MADERA')] + [c['con_vivienda']]
          for c in comunidades])

    hoja('Altitud',
         cols_base + ['Cota minima', 'Cota mediana', 'Cota maxima',
                      'Fichas con dato'],
         [base(c) + [c['cota_min'], c['cota_med'], c['cota_max'], c['cota_n']]
          for c in comunidades])

    hoja('Presa y capacitacion',
         cols_base + ['Conoce presa', 'Respuestas presa', 'Recibio capacitacion',
                      'Respuestas', 'Le gustaria', 'Respuestas deseo',
                      'Temas mas pedidos'],
         [base(c) + [c['presa_si'], c['presa_n'], c['cap_si'], c['cap_n'],
                     c['des_si'], c['des_n'],
                     '; '.join(f'{t} ({n})' for t, n in c['temas_top'])]
          for c in comunidades])

    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    wb.save(ruta)


# ─── main ────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[1])
    ap.add_argument('--sector', type=int, choices=[1, 2, 3],
                    help='genera SOLO ese sector como muestra (sufijo -MUESTRA)')
    args = ap.parse_args()

    comunidades, sup, caudal, corte_txt, fichas = agregar_todo()
    verificar(comunidades, sup, caudal, fichas)

    html, md = construir_documento(comunidades, sup, caudal, corte_txt, fichas,
                                   solo_sector=args.sector)

    if args.sector:
        html_out = HTML_OUT.replace('.html', f'-MUESTRA-sector{args.sector}.html')
        with open(html_out, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f'✔ Muestra escrita: {os.path.relpath(html_out, BASE)}')
        print('  (el documento completo, el .md y el Excel se generan sin --sector)')
        return

    with open(HTML_OUT, 'w', encoding='utf-8') as f:
        f.write(html)
    with open(MD_OUT, 'w', encoding='utf-8') as f:
        f.write(md)
    escribir_xlsx(comunidades, XLSX_OUT)
    print(f'✔ {os.path.relpath(HTML_OUT, BASE)}')
    print(f'✔ {os.path.relpath(MD_OUT, BASE)}')
    print(f'✔ {os.path.relpath(XLSX_OUT, BASE)}')
    print('  Word: python scripts/md_a_docx.py docs/INFORME-SOCIOLOGO-por-comunidad.md')


if __name__ == '__main__':
    main()
