# -*- coding: utf-8 -*-
"""
Generar reporte Excel catastral premium ejecutivo de entrega para el cliente.
Lee del GeoJSON depurado de la web para garantizar consistencia del 100% con el Dashboard.
Estructura las pestañas con hipervínculos cruzados bidireccionales, formatos corporativos y fórmulas.
"""

import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Directorio del proyecto web y archivos JSON depurados
PROJECT_DIR = r"C:\Users\HP\OneDrive\Escritorio\CAYAMBE CATASTRO RIEGO\padron-app"
GEO_DIR = os.path.join(PROJECT_DIR, "public", "geo")

FICHAS_GEOJSON = os.path.join(GEO_DIR, "fichas_predios.geojson")
CULTIVOS_JSON = os.path.join(GEO_DIR, "cultivos.json")
ANIMALES_JSON = os.path.join(GEO_DIR, "animales.json")
ADICIONALES_JSON = os.path.join(GEO_DIR, "predios_adicionales.json")

# Guardar en Escritorio y Descargas
OUTPUT_XLSX_DESKTOP = r"C:\Users\HP\OneDrive\Escritorio\CAYAMBE CATASTRO RIEGO\padron_usuarios_riego_consolidado.xlsx"
OUTPUT_XLSX_DOWNLOADS = r"C:\Users\HP\Downloads\padron_usuarios_riego_consolidado.xlsx"

# Mapeo de comunidades oficiales y sectores de constants.ts
COMUNIDADES_POR_SECTOR = {
    'Sector 1': [
        "ASOCIACIÓN 17 DE JUNIO", "ASOCIACIÓN POROTOG", "AVELLANEDA",
        "CARRERA", "CHAMBITOLA", "COCHAPAMBA", "COMUNA IZACATA",
        "COMUNA POROTOG", "CORDILLERAS DE LOS ANDES", "IZACATA GRANDE",
        "JESÚS GRAN PODER", "LA CANDELARIA", "LA LIBERTAD",
        "LARCACHACA", "LOMA GORDA", "LOS ANDES IZACATA",
        "MATÍAS IMBAGO", "MILAGRO", "SAN ANTONIO", "SAN JACINTO",
        "SAN JOSÉ", "SANTA BÁRBARA"
    ],
    'Sector 2': [
        "ALPAKA", "ASOC. PITANA BAJO", "ASOC. SAN VICENTE ALTO",
        "ASOC. SAN VICENTE BAJO", "ASOCIACIÓN ROSALÍA", "ASOCIACIÓN SAN PEDRO",
        "CUARTO LOTE", "PAMBAMARCA", "PITANA ALTO", "PROMEJ. PITANA BAJO",
        "PUCARÁ", "SANTA MARIANITA DE PINGULMI", "SANTA ROSA DE PACCHA",
        "SANTA ROSA DE PINGULMI"
    ],
    'Sector 3': [
        "ASOCIACIÓN ROSALÍA", "CANGAHUA PUNGO", "CHAUPIESTANCIA",
        "CHINCHINLOMA", "EL MANZANO", "HDA. GUANGUILQUI",
        "HDA. SAN FRANSISCO", "JUNTA SAN LUIS", "MONTESERÍN BAJO",
        "MONTESERRÍN ALTO", "OTONCITO", "PAMBAMARQUITO", "PUEBLO DE ASCÁZUBI",
        "PUEBLO DE OTÓN", "SAN VICENTE DE GUAYLLABAMBA", "SR. COLOMA",
        "SR. HERNÁN TIMPE"
    ]
}

# Metas de regantes planificadas de constants.ts
META_COMUNEROS = {
    # Sector 1
    "LARCACHACA": 103,
    "LA LIBERTAD": 125,
    "SAN ANTONIO": 95,
    "SAN JOSÉ": 120,
    "MILAGRO": 38,
    "ASOCIACIÓN 17 DE JUNIO": 32,
    "CHAMBITOLA": 120,
    "LA CANDELARIA": 170,
    "CARRERA": 280,
    "MATÍAS IMBAGO": 1,
    "COCHAPAMBA": 244,
    "JESÚS GRAN PODER": 56,
    "SANTA BÁRBARA": 6,
    "ASOCIACIÓN POROTOG": 48,
    "COMUNA POROTOG": 80,
    "CORDILLERAS DE LOS ANDES": 43,
    "COMUNA IZACATA": 65,
    "IZACATA GRANDE": 43,
    "LOS ANDES IZACATA": 48,
    "LOMA GORDA": 45,
    "SAN JACINTO": 6,
    
    # Sector 2
    "CUARTO LOTE": 46,
    "ASOC. SAN VICENTE BAJO": 80,
    "SANTA ROSA DE PACCHA": 46,
    "ASOC. SAN VICENTE ALTO": 59,
    "PUCARÁ": 231,
    "ASOCIACIÓN SAN PEDRO": 24,
    "PITANA ALTO": 99,
    "ALPAKA": 15,
    "ASOC. PITANA BAJO": 42,
    "PROMEJ. PITANA BAJO": 180,
    "SANTA ROSA DE PINGULMI": 117,
    "SANTA MARIANITA DE PINGULMI": 189,
    "PAMBAMARCA": 61,
    
    # Sector 3
    "OTONCITO": 66,
    "PAMBAMARQUITO": 100,
    "MONTESERRÍN ALTO": 27,
    "CHAUPIESTANCIA": 150,
    "PUEBLO DE OTÓN": 152,
    "CANGAHUA PUNGO": 147,
    "CHINCHINLOMA": 80,
    "ASOCIACIÓN ROSALÍA": 41,
    "SR. COLOMA": 16,
    "MONTESERÍN BAJO": 118,
    "HDA. GUANGUILQUI": 15,
    "PUEBLO DE ASCÁZUBI": 16,
    "EL MANZANO": 19,
    "JUNTA SAN LUIS": 33,
    "SAN VICENTE DE GUAYLLABAMBA": 453,
    
    # Comunidades sin datos meta pero existentes
    "AVELLANEDA": 0,
    "HDA. SAN FRANSISCO": 0,
    "SR. HERNÁN TIMPE": 0
}

MAPEO_TECNICOS = {
    'u0_a314': 'Melany Jara', 'u0_a319': 'Melany Jara', 'jvk-editor': 'Melany Jara',
    'u0_a504': 'Adriana Cuascota', 'jvk-editor6': 'Adriana Cuascota',
    'u0_a279': 'Huguito Ipial', 'jvk-editor2': 'Huguito Ipial',
    'u0_a70': 'Pablo Barrionuevo', 'jvk-editor5': 'Pablo Barrionuevo',
    'u0_a330': 'Mayra Benavides', 'mayralisseth201': 'Mayra Benavides',
    'u0_a362': 'Martha Simbaña', 'u0_a335': 'Martha Simbaña', 'jvk-editor4': 'Martha Simbaña',
    'u0_a302': 'Dylan Chavez', 'jvk-editor3': 'Dylan Chavez',
    'u0_a200': 'Melanie2', 'JVK-DIGITALIZACION': 'Oficina Digitalización'
}

def get_tecnico_name(usr):
    if not usr: return "Desconocido"
    return MAPEO_TECNICOS.get(usr, str(usr))

def main():
    if not os.path.exists(FICHAS_GEOJSON):
        print(f"[ERROR] Archivo GeoJSON de fichas no encontrado en: {FICHAS_GEOJSON}")
        return

    print("=" * 85)
    print(" GENERANDO EXCEL CATASTRAL CONSISTENTE CON LA WEB (LEENDO DE GEOJSON)")
    print("=" * 85)

    # 1. Cargar fichas del GeoJSON depurado
    with open(FICHAS_GEOJSON, 'r', encoding='utf-8') as f:
        geojson_data = json.load(f)
    properties_list = [feat['properties'] for feat in geojson_data['features']]
    df_fichas = pd.DataFrame(properties_list)
    print(f"Cargados {len(df_fichas)} predios depurados de {FICHAS_GEOJSON}")

    # 2. Cargar cultivos
    df_cultivos = pd.read_json(CULTIVOS_JSON)
    print(f"Cargados {len(df_cultivos)} cultivos depurados de {CULTIVOS_JSON}")

    # 3. Cargar animales
    df_animales = pd.read_json(ANIMALES_JSON)
    print(f"Cargados {len(df_animales)} animales depurados de {ANIMALES_JSON}")

    # 4. Cargar predios adicionales
    df_adicionales = pd.read_json(ADICIONALES_JSON)
    print(f"Cargados {len(df_adicionales)} predios adicionales depurados de {ADICIONALES_JSON}")

    # ─────────────────────────────────────────────────────────────────────────
    # PRE-PROCESAMIENTO Y DEPURACIÓN DE DATOS (CLIENTE)
    # ─────────────────────────────────────────────────────────────────────────
    # Asegurar que todas las columnas clave existan en el DataFrame (evita errores si vienen vacías)
    columnas_requeridas = [
        'id', 'codigo_final', 'propietario', 'apellidos', 'nombres', 'cedula', 'clave_catastral',
        'parroquia', 'comunidad', 'sector_comunidad', 'canal', 'sector', 'area_total', 'area_riego',
        'area_sin_riego', 'tiene_reservorio', 'dias_riego', 'horas_turno', 'agua_consumo',
        'energia_electrica', 'frecuencia_riego', 'metodo_gravedad_pct', 'metodo_aspersion_pct',
        'metodo_goteo_pct', 'valor_tarifa', 'tipo_tarifa', 'creado_por', 'fecha_creacion',
        'sector_investigacion', 'hijos_hombres', 'hijos_mujeres', 'nivel_instruccion', 'tenencia_predio'
    ]
    for col in columnas_requeridas:
        if col not in df_fichas.columns:
            df_fichas[col] = None

    df_fichas['propietario'] = df_fichas.apply(
        lambda r: f"{r['apellidos'] or ''} {r['nombres'] or ''}".strip().upper() or str(r['propietario'] or '').strip().upper(),
        axis=1
    )
    df_fichas['cedula'] = df_fichas['cedula'].astype(str).str.strip()
    df_fichas['comunidad'] = df_fichas['comunidad'].fillna('').str.strip()
    
    def normalize_com(c):
        if not c: return ""
        c = str(c).upper().strip()
        for a, b in [('Á','A'), ('É','E'), ('Í','I'), ('Ó','O'), ('Ú','U'), ('Ñ','N')]:
            c = c.replace(a, b)
        return c

    df_fichas['comunidad_norm'] = df_fichas['comunidad'].apply(normalize_com)
    
    # Derivar sector_investigacion priorizando la comunidad (idéntico al frontend)
    def derivar_sector(row):
        c_norm = normalize_com(row['comunidad'])
        
        # Para Asociación Rosalía respetamos el sector ingresado en SQLite/GeoJSON si existe
        if c_norm == 'ASOCIACION ROSALIA':
            val = row['sector_investigacion']
            if isinstance(val, str) and val.strip() != '':
                return val.strip()
            return 'Sector 3'
            
        # Para las demás, la comunidad manda sobre el sector de investigación
        for sec, coms in COMUNIDADES_POR_SECTOR.items():
            if any(normalize_com(c) == c_norm for c in coms):
                return sec
                
        # Fallback al valor físico o por defecto
        val = row['sector_investigacion']
        if isinstance(val, str) and val.strip() != '':
            return val.strip()
        return 'Sector 1'
        
    df_fichas['sector_investigacion'] = df_fichas.apply(derivar_sector, axis=1)

    # ─────────────────────────────────────────────────────────────────────────
    # CREACIÓN DEL LIBRO DE EXCEL PREMIUM
    # ─────────────────────────────────────────────────────────────────────────
    wb = Workbook()
    
    # Fuentes y estilos corporativos
    font_title = Font(name='Calibri', size=14, bold=True, color='0F5132')
    font_section = Font(name='Calibri', size=12, bold=True, color='0F5132')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Calibri', size=10)
    font_bold = Font(name='Calibri', size=10, bold=True)
    font_link = Font(name='Calibri', size=10, color='0A58CA', underline='single')
    
    fill_header = PatternFill(start_color='0F5132', end_color='0F5132', fill_type='solid')
    fill_total = PatternFill(start_color='D1E7DD', end_color='D1E7DD', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    border_thin = Side(style='thin', color='CBD5E1')
    border_double = Side(style='double', color='0F5132')
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double)

    # ─────────────────────────────────────────────────────────────────────────
    # PESTAÑA 1: RESUMEN GENERAL (DASHBOARD)
    # ─────────────────────────────────────────────────────────────────────────
    ws_dashboard = wb.active
    ws_dashboard.title = "Resumen Ejecutivo"
    ws_dashboard.views.sheetView[0].showGridLines = True
    
    ws_dashboard.cell(row=2, column=2, value="SISTEMA DE RIEGO COMUNITARIO GUANGUILQUI POROTOG").font = font_title
    ws_dashboard.cell(row=3, column=2, value="MONITOREO Y COBERTURA DE CATASTRO DE RIEGO Y ENCUESTAS").font = Font(name='Calibri', size=11, italic=True)
    
    ws_dashboard.cell(row=5, column=2, value="CATASTRO GLOBAL: RESUMEN CONSOLIDADO POR SECTORES").font = font_section
    headers_dashboard = ["SECTOR", "COMUNIDADES", "CATASTRO BASE (# PLANIFICADO)", "ENCUESTAS LEVANTADAS", "COBERTURA GLOBAL"]
    
    for c_idx, h in enumerate(headers_dashboard, start=2):
        cell = ws_dashboard.cell(row=6, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = cell_border
        
    r_idx = 7
    for sectorName in ['Sector 1', 'Sector 2', 'Sector 3']:
        coms = COMUNIDADES_POR_SECTOR[sectorName]
        planificado = sum(META_COMUNEROS.get(c, 0) for c in coms)
        
        # Conteo robusto de fichas levantadas insensible a tildes
        coms_norm = [normalize_com(c) for c in coms]
        levantado = len(df_fichas[(df_fichas['comunidad_norm'].isin(coms_norm)) & (df_fichas['sector_investigacion'] == sectorName)])
        
        ws_dashboard.cell(row=r_idx, column=2, value=sectorName.upper()).font = font_bold
        ws_dashboard.cell(row=r_idx, column=3, value=f"{len(coms)} comunidades").alignment = align_center
        ws_dashboard.cell(row=r_idx, column=4, value=planificado).number_format = '#,##0'
        ws_dashboard.cell(row=r_idx, column=5, value=levantado).number_format = '#,##0'
        
        pct_cell = ws_dashboard.cell(row=r_idx, column=6, value=f"=E{r_idx}/D{r_idx}")
        pct_cell.number_format = '0.0%'
        pct_cell.font = font_bold
        
        for c in range(2, 7):
            ws_dashboard.cell(row=r_idx, column=c).border = cell_border
            ws_dashboard.cell(row=r_idx, column=c).font = font_data
        r_idx += 1
        
    ws_dashboard.cell(row=r_idx, column=2, value="TOTAL GLOBAL").font = font_bold
    ws_dashboard.cell(row=r_idx, column=3, value="53 comunidades").alignment = align_center
    ws_dashboard.cell(row=r_idx, column=3).font = font_bold
    
    tot_plan = ws_dashboard.cell(row=r_idx, column=4, value="=SUM(D7:D9)")
    tot_plan.number_format = '#,##0'
    tot_plan.font = font_bold
    
    tot_lev = ws_dashboard.cell(row=r_idx, column=5, value="=SUM(E7:E9)")
    tot_lev.number_format = '#,##0'
    tot_lev.font = font_bold
    
    tot_pct = ws_dashboard.cell(row=r_idx, column=6, value="=E10/D10")
    tot_pct.number_format = '0.0%'
    tot_pct.font = Font(name='Calibri', size=11, bold=True, color='0F5132')
    
    for c in range(2, 7):
        cell = ws_dashboard.cell(row=r_idx, column=c)
        cell.fill = fill_total
        cell.border = total_border

    # ─────────────────────────────────────────────────────────────────────────
    # PESTAÑA 2: PADRÓN GENERAL (CATASTRO)
    # ─────────────────────────────────────────────────────────────────────────
    ws_padron = wb.create_sheet(title="Padrón General")
    ws_padron.views.sheetView[0].showGridLines = True
    
    headers_padron = [
        "N° REGISTRO", "CÉDULA / RUC", "CLAVE CATASTRAL", "APELLIDOS Y NOMBRES", 
        "HIJOS HOMBRES", "HIJOS MUJERES", "NIVEL INSTRUCCIÓN", "TENENCIA PREDIO",
        "PARROQUIA", "COMUNIDAD", "SECTOR / CANAL", 
        "ÁREA TOTAL (m²)", "ÁREA CON RIEGO (m²)", "ÁREA SIN RIEGO (m²)",
        "TIENE RESERVORIO", "DÍAS DE RIEGO", "HORAS DE TURNO", "CONSTRUCCIÓN: AGUA", "CONSTRUCCIÓN: ENERGÍA",
        "FRECUENCIA", "GRAVEDAD %", "ASPERSIÓN %", "GOTEO %", "TARIFA ($)", "TIPO TARIFA",
        "TÉCNICO", "FECHA REGISTRO", "CULTIVOS", "ANIMALES", "LOTES ADIC."
    ]
    
    for col_idx, h in enumerate(headers_padron, start=1):
        cell = ws_padron.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = cell_border
        
    padron_indices = {}
    
    print("Escribiendo pestaña de Padrón General...")
    for idx, row in df_fichas.iterrows():
        r_num = idx + 2
        uuid_val = row['id']
        padron_indices[uuid_val] = r_num
        
        ws_padron.cell(row=r_num, column=1, value=idx + 1).alignment = align_center
        
        # Cédula / RUC
        ced_cell = ws_padron.cell(row=r_num, column=2, value=row['cedula'])
        ced_cell.number_format = '@'
        ced_cell.alignment = align_center
        
        # Clave Catastral
        cc_cell = ws_padron.cell(row=r_num, column=3, value=row['clave_catastral'] or '')
        cc_cell.number_format = '@'
        cc_cell.alignment = align_center
        
        ws_padron.cell(row=r_num, column=4, value=row['propietario']).alignment = align_left
        
        # Hijos hombres/mujeres
        ws_padron.cell(row=r_num, column=5, value=row['hijos_hombres'] or 0).number_format = '0'
        ws_padron.cell(row=r_num, column=6, value=row['hijos_mujeres'] or 0).number_format = '0'
        
        # Nivel Instrucción y Tenencia Predio
        ws_padron.cell(row=r_num, column=7, value=row['nivel_instruccion'] or '').alignment = align_center
        ws_padron.cell(row=r_num, column=8, value=row['tenencia_predio'] or '').alignment = align_center
        
        # Parroquia, Comunidad, Sector/Canal
        ws_padron.cell(row=r_num, column=9, value=row['parroquia']).alignment = align_center
        ws_padron.cell(row=r_num, column=10, value=row['comunidad']).alignment = align_left
        ws_padron.cell(row=r_num, column=11, value=row['canal'] or row['sector'] or '').alignment = align_left
        
        # Áreas
        ws_padron.cell(row=r_num, column=12, value=row['area_total'] or 0.0).number_format = '#,##0.00'
        ws_padron.cell(row=r_num, column=13, value=row['area_riego'] or 0.0).number_format = '#,##0.00'
        ws_padron.cell(row=r_num, column=14, value=row['area_sin_riego'] or 0.0).number_format = '#,##0.00'
        
        # Reservorio y turnos de riego
        ws_padron.cell(row=r_num, column=15, value=row['tiene_reservorio'] or '').alignment = align_center
        ws_padron.cell(row=r_num, column=16, value=row['dias_riego'] or 0).number_format = '0'
        ws_padron.cell(row=r_num, column=17, value=row['horas_turno'] or 0).number_format = '0'
        
        # Construcción: Agua / Energía
        agua_str = "SÍ" if row['agua_consumo'] in [1, True] else "NO"
        energia_str = "SÍ" if row['energia_electrica'] in [1, True] else "NO"
        ws_padron.cell(row=r_num, column=18, value=agua_str).alignment = align_center
        ws_padron.cell(row=r_num, column=19, value=energia_str).alignment = align_center
        
        # Frecuencia y métodos de riego
        ws_padron.cell(row=r_num, column=20, value=row['frecuencia_riego'] or '').alignment = align_center
        ws_padron.cell(row=r_num, column=21, value=row['metodo_gravedad_pct'] or 0).number_format = '0'
        ws_padron.cell(row=r_num, column=22, value=row['metodo_aspersion_pct'] or 0).number_format = '0'
        ws_padron.cell(row=r_num, column=23, value=row['metodo_goteo_pct'] or 0).number_format = '0'
        
        # Tarifa
        ws_padron.cell(row=r_num, column=24, value=row['valor_tarifa'] or 0.0).number_format = '$#,##0.00'
        ws_padron.cell(row=r_num, column=25, value=row['tipo_tarifa'] or '').alignment = align_center
        ws_padron.cell(row=r_num, column=26, value=get_tecnico_name(row['creado_por'])).alignment = align_left
        
        # Fecha
        fecha_val = str(row['fecha_creacion'] or '')[:10]
        ws_padron.cell(row=r_num, column=27, value=fecha_val).alignment = align_center
        
        # Bordes y fuente
        for c in range(1, 31):
            ws_padron.cell(row=r_num, column=c).border = cell_border
            ws_padron.cell(row=r_num, column=c).font = font_data

    ws_padron.auto_filter.ref = f"A1:AD{len(df_fichas) + 1}"
    ws_padron.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────────────────
    # PESTAÑA 3: CULTIVOS Y PRODUCCIÓN
    # ─────────────────────────────────────────────────────────────────────────
    ws_cultivos = wb.create_sheet(title="Cultivos y Producción")
    ws_cultivos.views.sheetView[0].showGridLines = True
    
    headers_cultivos = ["CÓDIGO PREDIO", "PROPIETARIO / REGANTE", "PARROQUIA", "COMUNIDAD", "CULTIVO REGISTRADO", "SUPERFICIE (m²)", "PRIORIDAD", "VOLVER AL PADRÓN"]
    for col_idx, h in enumerate(headers_cultivos, start=1):
        cell = ws_cultivos.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = cell_border
        
    cultivos_indices = {}
    r_idx = 2
    
    print("Escribiendo pestaña de Cultivos...")
    df_fichas_lookup = df_fichas.set_index('id')
    
    for idx, row in df_cultivos.iterrows():
        f_id = row['ficha_id']
        if f_id in df_fichas_lookup.index:
            ficha = df_fichas_lookup.loc[f_id]
            codigo_final = ficha['codigo_final']
            propietario = ficha['propietario']
            parroquia = ficha['parroquia']
            comunidad = ficha['comunidad']
            
            if f_id not in cultivos_indices:
                cultivos_indices[f_id] = r_idx
                
            ws_cultivos.cell(row=r_idx, column=1, value=codigo_final).alignment = align_center
            ws_cultivos.cell(row=r_idx, column=2, value=propietario).alignment = align_left
            ws_cultivos.cell(row=r_idx, column=3, value=parroquia).alignment = align_center
            ws_cultivos.cell(row=r_idx, column=4, value=comunidad).alignment = align_left
            ws_cultivos.cell(row=r_idx, column=5, value=str(row['tipo_cultivo']).upper().strip()).alignment = align_left
            ws_cultivos.cell(row=r_idx, column=6, value=row['superficie_m2'] or 0.0).number_format = '#,##0.00'
            
            prioridad = "Principal" if row['es_principal'] else "Secundario"
            ws_cultivos.cell(row=r_idx, column=7, value=prioridad).alignment = align_center
            
            r_padron = padron_indices.get(f_id, 2)
            ret_cell = ws_cultivos.cell(row=r_idx, column=8, value="Volver")
            ret_cell.hyperlink = f"#'Padrón General'!D{r_padron}"
            ret_cell.font = font_link
            ret_cell.alignment = align_center
            
            for c in range(1, 9):
                ws_cultivos.cell(row=r_idx, column=c).border = cell_border
                if c != 8:
                    ws_cultivos.cell(row=r_idx, column=c).font = font_data
            r_idx += 1

    ws_cultivos.auto_filter.ref = f"A1:H{r_idx - 1}"
    ws_cultivos.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────────────────
    # PESTAÑA 4: INVENTARIO PECUARIO
    # ─────────────────────────────────────────────────────────────────────────
    ws_animales = wb.create_sheet(title="Inventario Pecuario")
    ws_animales.views.sheetView[0].showGridLines = True
    
    headers_animales = ["CÓDIGO PREDIO", "PROPIETARIO / REGANTE", "COMUNIDAD", "ESPECIE ANIMAL", "CANTIDAD (UNIDADES)", "VOLVER AL PADRÓN"]
    for col_idx, h in enumerate(headers_animales, start=1):
        cell = ws_animales.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = cell_border
        
    animales_indices = {}
    r_idx = 2
    
    print("Escribiendo pestaña de Especies Pecuarias...")
    for idx, row in df_animales.iterrows():
        f_id = row['ficha_id']
        if f_id in df_fichas_lookup.index:
            ficha = df_fichas_lookup.loc[f_id]
            codigo_final = ficha['codigo_final']
            propietario = ficha['propietario']
            comunidad = ficha['comunidad']
            
            if f_id not in animales_indices:
                animales_indices[f_id] = r_idx
                
            ws_animales.cell(row=r_idx, column=1, value=codigo_final).alignment = align_center
            ws_animales.cell(row=r_idx, column=2, value=propietario).alignment = align_left
            ws_animales.cell(row=r_idx, column=3, value=comunidad).alignment = align_left
            ws_animales.cell(row=r_idx, column=4, value=str(row['especie']).upper().strip()).alignment = align_left
            ws_animales.cell(row=r_idx, column=5, value=row['cantidad'] or 0).number_format = '#,##0'
            
            r_padron = padron_indices.get(f_id, 2)
            ret_cell = ws_animales.cell(row=r_idx, column=6, value="Volver")
            ret_cell.hyperlink = f"#'Padrón General'!D{r_padron}"
            ret_cell.font = font_link
            ret_cell.alignment = align_center
            
            for c in range(1, 7):
                ws_animales.cell(row=r_idx, column=c).border = cell_border
                if c != 6:
                    ws_animales.cell(row=r_idx, column=c).font = font_data
            r_idx += 1

    ws_animales.auto_filter.ref = f"A1:F{r_idx - 1}"
    ws_animales.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────────────────
    # PESTAÑA 5: LOTES SECUNDARIOS (PREDIOS ADICIONALES)
    # ─────────────────────────────────────────────────────────────────────────
    ws_predios = wb.create_sheet(title="Lotes Adicionales")
    ws_predios.views.sheetView[0].showGridLines = True
    
    headers_predios = ["CÓDIGO PREDIO", "PROPIETARIO / REGANTE", "COMUNIDAD", "CÓDIGO LOTE ADICIONAL", "ÁREA LOTE (m²)", "VOLVER AL PADRÓN"]
    for col_idx, h in enumerate(headers_predios, start=1):
        cell = ws_predios.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = cell_border
        
    predios_indices = {}
    r_idx = 2
    
    print("Escribiendo pestaña de Lotes Adicionales...")
    for idx, row in df_adicionales.iterrows():
        f_id = row['ficha_id']
        if f_id in df_fichas_lookup.index:
            ficha = df_fichas_lookup.loc[f_id]
            codigo_final = ficha['codigo_final']
            propietario = ficha['propietario']
            comunidad = ficha['comunidad']
            
            if f_id not in predios_indices:
                predios_indices[f_id] = r_idx
                
            ws_predios.cell(row=r_idx, column=1, value=codigo_final).alignment = align_center
            ws_predios.cell(row=r_idx, column=2, value=propietario).alignment = align_left
            ws_predios.cell(row=r_idx, column=3, value=comunidad).alignment = align_left
            ws_predios.cell(row=r_idx, column=4, value=str(row['clave_catastral_otro'] or '').upper().strip()).alignment = align_center
            ws_predios.cell(row=r_idx, column=5, value=row['area_total_otro'] or 0.0).number_format = '#,##0.00'
            
            r_padron = padron_indices.get(f_id, 2)
            ret_cell = ws_predios.cell(row=r_idx, column=6, value="Volver")
            ret_cell.hyperlink = f"#'Padrón General'!D{r_padron}"
            ret_cell.font = font_link
            ret_cell.alignment = align_center
            
            for c in range(1, 7):
                ws_predios.cell(row=r_idx, column=c).border = cell_border
                if c != 6:
                    ws_predios.cell(row=r_idx, column=c).font = font_data
            r_idx += 1

    ws_predios.auto_filter.ref = f"A1:F{r_idx - 1}"
    ws_predios.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────────────────
    # ESCRIBIR HIPERVÍNCULOS EN EL PADRÓN GENERAL (DESPLAZAMIENTO CRUZADO)
    # ─────────────────────────────────────────────────────────────────────────
    print("Inyectando hipervínculos de navegación interactiva bidireccionales...")
    for idx, row in df_fichas.iterrows():
        r_num = idx + 2
        f_id = row['id']
        
        # Enlace a Cultivos (Columna 28)
        col_cultivos = ws_padron.cell(row=r_num, column=28)
        if f_id in cultivos_indices:
            col_cultivos.value = "Ver Cultivos"
            col_cultivos.hyperlink = f"#'Cultivos y Producción'!A{cultivos_indices[f_id]}"
            col_cultivos.font = font_link
            col_cultivos.alignment = align_center
        else:
            col_cultivos.value = "-"
            col_cultivos.alignment = align_center
            
        # Enlace a Animales (Columna 29)
        col_animales = ws_padron.cell(row=r_num, column=29)
        if f_id in animales_indices:
            col_animales.value = "Ver Animales"
            col_animales.hyperlink = f"#'Inventario Pecuario'!A{animales_indices[f_id]}"
            col_animales.font = font_link
            col_animales.alignment = align_center
        else:
            col_animales.value = "-"
            col_animales.alignment = align_center
            
        # Enlace a Adicionales (Columna 30)
        col_predios = ws_padron.cell(row=r_num, column=30)
        if f_id in predios_indices:
            col_predios.value = "Ver Lotes"
            col_predios.hyperlink = f"#'Lotes Adicionales'!A{predios_indices[f_id]}"
            col_predios.font = font_link
            col_predios.alignment = align_center
        else:
            col_predios.value = "-"
            col_predios.alignment = align_center

    # ─────────────────────────────────────────────────────────────────────────
    # AUTO-AJUSTAR ANCHO DE COLUMNAS DE TODAS LAS PESTAÑAS (EVITA ###)
    # ─────────────────────────────────────────────────────────────────────────
    print("Auto-ajustando el ancho de columnas de forma dinámica...")
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith('='):
                    val = "TOTAL_FORMULA"
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = col[0].column_letter
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Ajustar dimensiones de la pestaña de resumen a mano para que luzca perfecta
    ws_dashboard.column_dimensions['B'].width = 45
    ws_dashboard.column_dimensions['C'].width = 20
    ws_dashboard.column_dimensions['D'].width = 30
    ws_dashboard.column_dimensions['E'].width = 25
    ws_dashboard.column_dimensions['F'].width = 20

    # Guardar los archivos de forma segura
    saved_desktop = False
    saved_downloads = False
    
    try:
        wb.save(OUTPUT_XLSX_DESKTOP)
        print(f"  ✓ Guardado en Escritorio: {OUTPUT_XLSX_DESKTOP}")
        saved_desktop = True
    except PermissionError:
        alt_path = OUTPUT_XLSX_DESKTOP.replace(".xlsx", "_CERRAR_EXCEL_Y_RENOMBRAR.xlsx")
        try:
            wb.save(alt_path)
            print(f"  ⚠️ Escritorio bloqueado (Excel abierto). Guardado alternativo en:\n    -> {alt_path}")
            saved_desktop = True
        except Exception as e:
            print(f"  ❌ Error al guardar archivo alternativo en Escritorio: {e}")
            
    try:
        wb.save(OUTPUT_XLSX_DOWNLOADS)
        print(f"  ✓ Guardado en Descargas: {OUTPUT_XLSX_DOWNLOADS}")
        saved_downloads = True
    except PermissionError:
        alt_path = OUTPUT_XLSX_DOWNLOADS.replace(".xlsx", "_CERRAR_EXCEL_Y_RENOMBRAR.xlsx")
        try:
            wb.save(alt_path)
            print(f"  ⚠️ Descargas bloqueado (Excel abierto). Guardado alternativo en:\n    -> {alt_path}")
            saved_downloads = True
        except Exception as e:
            print(f"  ❌ Error al guardar archivo alternativo en Descargas: {e}")

    if saved_desktop or saved_downloads:
        print(f"\n[OK] Excel catastral premium generado con éxito.")
    else:
        print(f"\n[ERROR] No se pudo guardar el archivo de Excel en ninguna ruta.")
    print("=" * 85)

if __name__ == "__main__":
    main()
