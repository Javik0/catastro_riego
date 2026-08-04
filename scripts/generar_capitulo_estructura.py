# -*- coding: utf-8 -*-
"""
Capítulo del informe técnico: "Estructura del padrón".

Explica de qué está hecho el padrón —personas, predios y fichas— y por qué esas
tres cifras no coinciden. Cubre las secciones 6 y 7 de la ficha (observaciones y
otros predios del regante) y el estado del levantamiento.

POR QUÉ ESTE CAPÍTULO
---------------------
El padrón maneja tres números distintos que se confunden con facilidad:
6.823 fichas, 4.301 entrevistas y ~4.082 personas. Cada uno responde a una
pregunta diferente y usarlos indistintamente produce cifras contradictorias
entre capítulos. Aquí se fijan las definiciones.

SOBRE LA CONCENTRACIÓN DE LA TIERRA
-----------------------------------
Los mayores tenedores del padrón NO son personas acaparando tierra: son la
comunidad de Monteserrín Bajo (729 ha), una empresa (601 ha) y varios comités
pro mejoras. Presentar la concentración sin separar los titulares colectivos
daría a entender un acaparamiento privado que no existe, así que se calcula por
separado para personas naturales y para organizaciones.

SALIDAS
  docs/CAPITULO-estructura-del-padron.html
  build_entrega/Estructura_del_Padron.xlsx
"""

import os
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comunidades_canon import canonica, nombre_publico, normalizar  # noqa: E402
import informe_estilo as E  # noqa: E402

GPKG = r"C:\Users\HP\QField\cloud\porotog_levantamiento_offline\data.gpkg"
T = 'Fichas_Predios_880eb10d_d887_4fc6_99a2_8af3ac63877e'
BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
HTML = os.path.join(BASE, 'docs', 'CAPITULO-estructura-del-padron.html')
XLSX = os.path.join(BASE, 'build_entrega', 'Estructura_del_Padron.xlsx')
MESES = ('enero febrero marzo abril mayo junio julio agosto septiembre '
         'octubre noviembre diciembre').split()

# Palabras que identifican a un titular COLECTIVO (no una persona natural).
COLECTIVOS = ('COMUNA', 'COMITE', 'COMITÉ', 'ASOCIACION', 'ASOCIACIÓN', 'JUNTA',
              'COMUNIDAD', 'PROMEJORAS', 'PRO MEJORAS', 'HACIENDA', 'HDA',
              'S.A', 'CIA', 'COOPERATIVA', 'FUNDACION', 'FUNDACIÓN')


def sin_tildes(s):
    s = unicodedata.normalize('NFD', (s or '').upper())
    return re.sub(r'\s+', ' ', ''.join(c for c in s
                                       if unicodedata.category(c) != 'Mn')).strip()


def num(f, k):
    try:
        return float(f.get(k) or 0)
    except (TypeError, ValueError):
        return 0.0


def pct(a, b):
    return 100.0 * a / b if b else 0.0


def es_hija(f):
    return f.get('es_ficha_hija') in (1, True)


def ident(f):
    """Identidad del titular: cédula si existe, si no el nombre normalizado."""
    ced = (f.get('cedula') or '').strip()
    if ced:
        return 'CI:' + ced
    return 'NOM:' + sin_tildes(f"{f.get('apellidos') or ''} {f.get('nombres') or ''}")


def es_colectivo(f):
    n = sin_tildes(f"{f.get('apellidos') or ''} {f.get('nombres') or ''}")
    return any(k in n for k in COLECTIVOS)


def main():
    con = sqlite3.connect(GPKG)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(f'SELECT * FROM "{T}"')
    todas = [dict(r) for r in cur.fetchall()]
    cur.execute(f'SELECT MAX(fecha_creacion), MAX(fecha_completado) FROM "{T}"')
    f1, f2 = cur.fetchone()
    con.close()

    corte = max(str(f1 or '')[:10], str(f2 or '')[:10])
    corte_txt = (f'{int(corte[8:10])} de {MESES[int(corte[5:7]) - 1]} de {corte[:4]}'
                 if corte else 'la fecha de generación')

    por_id = {f['id']: f for f in todas}
    pri = [f for f in todas if not es_hija(f)]
    hij = [f for f in todas if es_hija(f)]
    N = len(todas)

    # ── personas ──
    grupos = defaultdict(list)
    for f in todas:
        grupos[ident(f)].append(f)
    personas = len(grupos)
    dist = Counter(len(v) for v in grupos.values())
    multi = sum(v for k, v in dist.items() if k > 1)

    # personas que solo existen como predio adicional de otro titular
    ids_pri = {ident(f) for f in pri}
    solo_adic = {ident(f) for f in hij
                 if por_id.get(f.get('ficha_madre_id'))
                 and ident(f) != ident(por_id[f['ficha_madre_id']])
                 and ident(f) not in ids_pri}

    # ── concentración, separando titulares colectivos ──
    sup_por_titular = {k: sum(num(f, 'area_total') for f in v) for k, v in grupos.items()}
    colectivos = {k for k, v in grupos.items() if es_colectivo(v[0])}
    sup_col = sum(s for k, s in sup_por_titular.items() if k in colectivos)
    sup_tot = sum(sup_por_titular.values())
    naturales = sorted((s for k, s in sup_por_titular.items() if k not in colectivos),
                       reverse=True)
    sup_nat = sum(naturales)

    def concentra(lista, p):
        n = max(1, int(len(lista) * p / 100))
        return n, pct(sum(lista[:n]), sum(lista))

    top_col = sorted(((s, k) for k, s in sup_por_titular.items() if k in colectivos),
                     reverse=True)[:6]

    # ── estado del levantamiento ──
    estados = Counter(str(f.get('estado_investigacion') or 'sin estado') for f in hij)
    obs = sum(1 for f in todas if (f.get('observaciones') or '').strip())

    # ── documento ──
    B = []
    A = B.append
    A(E.cabecera('Estructura del padrón',
                 'Personas, predios y fichas · Capítulo del informe técnico'))
    A(E.aviso_corte(corte_txt, len(pri), estados.get('pendiente_produccion', 0)))
    A(E.kpis([
        (f'{N:,}', 'fichas de predio'),
        (f'{len(pri):,}', 'regantes entrevistados'),
        (f'{personas:,}', 'personas en el padrón'),
        (f'{pct(multi, personas):.0f}%', 'con más de un predio'),
    ]))

    A('<h2>1. Tres cifras que no son lo mismo</h2>')
    A('<p>El padrón maneja tres números distintos que responden a preguntas '
      'diferentes. Confundirlos produce contradicciones entre capítulos, de modo '
      'que conviene fijarlos:</p>')
    A('<table class="evitar-corte"><tr><th>Cifra</th><th class="n">Valor</th>'
      '<th>Qué responde</th></tr>')
    A(f'<tr><td><b>Fichas de predio</b></td><td class="n">{N:,}</td>'
      '<td>¿Cuántas parcelas se levantaron? Es la unidad de la producción y del catastro</td></tr>')
    A(f'<tr class="dest"><td><b>Regantes entrevistados</b></td><td class="n">{len(pri):,}</td>'
      '<td>¿A cuántas personas se aplicó la encuesta? Es la unidad de los capítulos sociales</td></tr>')
    A(f'<tr><td><b>Personas del padrón</b></td><td class="n">{personas:,}</td>'
      '<td>¿Cuántos titulares distintos hay? Incluye a quienes solo constan como predio adicional</td></tr>')
    A('</table>')
    A(f'<p>La diferencia entre las dos primeras se explica por los '
      f'<b>{len(hij):,} predios adicionales</b>: un mismo regante puede tener varias '
      'parcelas y se le entrevista una sola vez. La tercera cifra incorpora además '
      f'<b>{len(solo_adic)} personas</b> que figuran únicamente como titulares de un '
      'predio adicional declarado por otro y que no tienen ficha propia.</p>')
    A('<div class="nota"><b>Regla de uso.</b> Para hablar de <i>personas</i> '
      '(escolaridad, conocimiento, capacitación) se usan las fichas principales. '
      'Para hablar de <i>territorio o producción</i> (superficie, cultivos, ganado) '
      'se usan todas las fichas, porque cada predio produce. Nunca deben sumarse '
      'ambos universos.</div>')

    A('<h2>2. Cuántos predios tiene cada regante</h2>')
    A('<table class="evitar-corte"><tr><th>Predios por titular</th>'
      '<th class="n">Titulares</th><th>Peso</th></tr>')
    for k in sorted(dist):
        if k > 6:
            continue
        A(f'<tr><td>{k} predio{"s" if k > 1 else ""}</td>'
          f'<td class="n">{dist[k]:,}</td><td>{E.barra(pct(dist[k], personas))}</td></tr>')
    mas = sum(v for k, v in dist.items() if k > 6)
    A(f'<tr><td>7 o más</td><td class="n">{mas:,}</td>'
      f'<td>{E.barra(pct(mas, personas))}</td></tr>')
    A('</table>')
    A(f'<p>El <b>{pct(dist[1], personas):.1f} % de los titulares tiene un solo '
      f'predio</b>. El resto —{multi:,} personas— posee dos o más parcelas, con un '
      f'máximo de {max(dist)} predios en un mismo titular. Esta dispersión explica '
      'por qué el número de fichas supera al de regantes y por qué el registro de '
      'predios adicionales fue necesario.</p>')

    A('<h2>3. Distribución de la tierra</h2>')
    A('<div class="nota"><b>Advertencia de lectura.</b> Los mayores tenedores del '
      'padrón <b>no son personas acumulando tierra</b>: son la propia comunidad de '
      'Monteserrín Bajo, comités pro mejoras, comunas y una empresa. Presentar la '
      'concentración sin separarlos sugeriría un acaparamiento privado que no '
      'existe. Por eso se calculan por separado.</div>')
    A('<h3>Titulares colectivos</h3>')
    A(f'<p><b>{len(colectivos)} titulares</b> del padrón son organizaciones —comunas, '
      f'comités, asociaciones, haciendas o empresas— y reúnen <b>{sup_col / 10000:,.1f} '
      f'ha, el {pct(sup_col, sup_tot):.1f} % de la superficie</b>:</p>')
    A('<table class="evitar-corte"><tr><th>Titular colectivo</th>'
      '<th class="n">Superficie (ha)</th></tr>')
    for s, k in top_col:
        v = grupos[k][0]
        nom = f"{v.get('apellidos') or ''} {v.get('nombres') or ''}".strip()
        A(f'<tr><td>{nom[:52]}</td><td class="n">{s / 10000:,.1f}</td></tr>')
    A('</table>')
    A('<h3>Personas naturales</h3>')
    A(f'<p>Excluidos los titulares colectivos, quedan <b>{len(naturales):,} personas '
      f'naturales</b> con <b>{sup_nat / 10000:,.1f} ha</b>. Su distribución:</p>')
    A('<table class="evitar-corte"><tr><th>Tramo</th><th class="n">Personas</th>'
      '<th>Superficie que concentran</th></tr>')
    for p in (1, 5, 10, 25, 50):
        n, c = concentra(naturales, p)
        A(f'<tr><td>El {p} % con más tierra</td><td class="n">{n:,}</td>'
          f'<td>{E.barra(c)}</td></tr>')
    A('</table>')
    n10, c10 = concentra(naturales, 10)
    A(f'<p>Entre personas naturales la concentración sigue siendo alta: el '
      f'<b>{10} % con más tierra reúne el {c10:.1f} %</b> de la superficie privada. '
      'Coherente con el minifundio descrito en el capítulo del predio: la mayoría '
      'de los regantes trabaja parcelas pequeñas mientras unos pocos concentran '
      'las extensiones mayores.</p>')

    A('<h2>4. Estado del levantamiento</h2>')
    A(f'<p>De los <b>{len(hij):,} predios adicionales</b> registrados:</p>')
    A('<table class="evitar-corte"><tr><th>Estado</th><th class="n">Predios</th>'
      '<th>Peso</th></tr>')
    ET = {'completada': 'Con producción levantada',
          'pendiente_produccion': 'Pendiente de levantar la producción',
          'en_revision': 'En revisión'}
    for k, n in estados.most_common():
        A(f'<tr><td>{ET.get(k, k)}</td><td class="n">{n:,}</td>'
          f'<td>{E.barra(pct(n, len(hij)))}</td></tr>')
    A('</table>')
    pend = estados.get('pendiente_produccion', 0)
    A(f'<p>Quedan <b>{pend:,} predios adicionales</b> por completar '
      f'({pct(pend, len(hij)):.1f} % de los adicionales). Es el trabajo de campo '
      'pendiente a la fecha de corte y la razón por la que las cifras de este '
      'informe son provisionales.</p>')
    A(f'<p>Adicionalmente, <b>{obs:,} fichas</b> incluyen observaciones escritas por '
      'el técnico, una fuente cualitativa que documenta casos particulares del '
      'levantamiento.</p>')

    A('<h2>5. Conclusiones</h2>')
    A('<ul>')
    A(f'<li>El padrón registra <b>{N:,} predios</b> pertenecientes a '
      f'<b>{personas:,} titulares</b>, de los cuales {len(pri):,} fueron '
      'entrevistados.</li>')
    A(f'<li><b>{pct(multi, personas):.0f} % de los titulares posee más de un '
      'predio</b>, lo que hace imprescindible distinguir entre contar personas y '
      'contar parcelas.</li>')
    A(f'<li><b>{len(colectivos)} titulares colectivos</b> (comunas, comités, '
      f'haciendas) concentran el {pct(sup_col, sup_tot):.1f} % de la superficie: '
      'la propiedad colectiva es una característica estructural del sistema, no '
      'una anomalía.</li>')
    A(f'<li>Entre personas naturales, el 10 % mayor reúne el {c10:.0f} % de la '
      'tierra privada.</li>')
    A(f'<li>Restan <b>{pend:,} predios adicionales</b> por completar para cerrar '
      'el levantamiento.</li>')
    A('</ul>')
    A(E.pie(corte_txt))

    os.makedirs(os.path.dirname(HTML), exist_ok=True)
    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(E.documento('Estructura del padrón — Padrón Guanguilquí–Porotog',
                            '\n'.join(B)))
    print(f'  capítulo: {os.path.relpath(HTML, BASE)}  (corte: {corte_txt})')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    azul, blanco = PatternFill('solid', fgColor='1e4d8c'), Font(color='FFFFFF', bold=True)

    def hoja(nombre, cab, filas):
        ws = wb.create_sheet(nombre)
        ws.append(cab)
        for c in ws[1]:
            c.fill, c.font = azul, blanco
        for f_ in filas:
            ws.append(f_)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                12, min(46, max(len(str(c.value or '')) for c in col) + 2))

    hoja('Resumen', ['Indicador', 'Valor'], [
        ['Fichas de predio', N], ['Regantes entrevistados', len(pri)],
        ['Predios adicionales', len(hij)], ['Personas distintas', personas],
        ['Personas solo como predio adicional', len(solo_adic)],
        ['% con más de un predio', round(pct(multi, personas), 1)],
        ['Titulares colectivos', len(colectivos)],
        ['Superficie de titulares colectivos (ha)', round(sup_col / 10000, 1)],
        ['% superficie colectiva', round(pct(sup_col, sup_tot), 1)],
        ['Predios adicionales pendientes', pend],
    ])
    hoja('Predios por titular', ['Predios', 'Titulares'],
         [[k, dist[k]] for k in sorted(dist)])
    hoja('Titulares colectivos', ['Titular', 'Superficie (ha)', 'Predios'],
         [[f"{grupos[k][0].get('apellidos') or ''} {grupos[k][0].get('nombres') or ''}".strip(),
           round(s / 10000, 2), len(grupos[k])]
          for s, k in sorted(((sup_por_titular[k], k) for k in colectivos), reverse=True)])
    hoja('Concentración', ['Tramo (personas naturales)', 'Personas', '% superficie'],
         [[f'El {p} % con más tierra', concentra(naturales, p)[0],
           round(concentra(naturales, p)[1], 1)] for p in (1, 5, 10, 25, 50)])
    hoja('Estado adicionales', ['Estado', 'Predios'],
         [[ET.get(k, k), n] for k, n in estados.most_common()])

    del wb['Sheet']
    os.makedirs(os.path.dirname(XLSX), exist_ok=True)
    wb.save(XLSX)
    print(f'  excel   : {os.path.relpath(XLSX, BASE)}')
    print(f'\n  {N:,} fichas | {personas:,} personas | {pct(multi, personas):.0f}% multi-predio | '
          f'{len(colectivos)} titulares colectivos ({pct(sup_col, sup_tot):.1f}% del área)')


if __name__ == '__main__':
    main()
